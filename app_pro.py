import streamlit as st
import pandas as pd
import numpy as np
from scipy import stats
from scipy.special import gamma, gammainc
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import rcParams
import io
import warnings
import calendar
import plotly.graph_objects as go
from scipy.optimize import brentq
from lang import (fmt_c, cur_symbol, cur_decimal, CURRENCIES, LANG_DEFAULTS,
                  T, set_lang, get_lang,
                  BIZ_SUBSCRIPTION, BIZ_SPOT,
                  BILLING_CALENDAR_MONTHLY, BILLING_ANNUAL_365, BILLING_CUSTOM_DAYS,
                  BILLING_FIXED_30, BILLING_DAILY_SPOT)
warnings.filterwarnings('ignore')

# ── Page config ──────────────────────────────────────────────
st.set_page_config(
    page_title="LTV Analyzer Demo" if st.secrets.get("MODE", "demo").lower() == "demo" else "LTV Analyzer Advanced" if st.secrets.get("MODE", "demo").lower() == "advanced" else "LTV Analyzer Standard",
    page_icon="◆",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Mode & Authentication ─────────────────────────────────────
# st.secrets に MODE ("demo"/"standard"/"advanced") と PASSWORD を設定
# demo: パスワード不要、サンプルデータのみ
# standard/advanced: パスワード必要、CSVアップロード可
APP_MODE = st.secrets.get("MODE", "demo").lower()

if APP_MODE in ("standard", "advanced") and "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if APP_MODE in ("standard", "advanced") and not st.session_state.authenticated:
    st.set_page_config is None  # already called above
    _auth_col1, _auth_col2, _auth_col3 = st.columns([1, 1.5, 1])
    with _auth_col2:
        st.markdown("<div style='padding-top: 120px;'></div>", unsafe_allow_html=True)
        st.markdown("""
        <div style='text-align: center; margin-bottom: 24px;'>
            <div style='font-size: 1.4rem; font-weight: 500; color: #c8d0d8; letter-spacing: -0.02em;'>LTV Analyzer</div>
            <div style='font-size: 0.75rem; color: #3a6a7a; margin-top: 4px;'>Enter your password to continue</div>
        </div>
        """, unsafe_allow_html=True)
        _pw = st.text_input("Password", type="password", label_visibility="collapsed", placeholder="Password")
        if st.button("Enter", use_container_width=True):
            if _pw == st.secrets.get("PASSWORD", ""):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password")
    st.stop()

# ── CSS ───────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=BIZ+UDPGothic:wght@400;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

/* ── Base font ── */
html {
    font-size: 13px;
}
.stApp *:not(code):not(pre):not(.js-plotly-plot):not(.js-plotly-plot *):not([class*="material"]):not([class*="icon"]):not([data-testid="stIconMaterial"]) {
    font-family: 'BIZ UDPGothic', sans-serif !important;
}
.stApp { background-color: #0a0e14; color: #c8d0d8; }

/* ── Sidebar divider lines ── */
[data-testid="stSidebar"] hr,
[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] hr,
[data-testid="stSidebarContent"] hr { display: none !important; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background-color: #0d1117;
    border-right: 1px solid #1c2430;
    color: #c8d0d8;
}
[data-testid="stSidebar"] .stMarkdown h3 {
    font-size: 0.8rem;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #8ab4c4;
    margin: 24px 0 8px 0;
    padding-bottom: 6px;
    border-bottom: 1px solid #1c2430;
}

/* ── Metric cards ── */
.metric-card {
    background: #0d1520;
    border: 1px solid #1a2a3a;
    border-radius: 8px;
    padding: 22px 18px;
    text-align: center;
    height: 130px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    transition: border-color 0.2s;
}
.metric-card:hover { border-color: #2a4a5a; }
.metric-value {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.75rem;
    font-weight: 500;
    color: #56b4d3;
    line-height: 1.1;
    word-break: break-all;
    letter-spacing: -0.02em;
}
.metric-label {
    font-size: 0.68rem;
    font-weight: 500;
    color: #7a9aaa;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-top: 8px;
}
.metric-desc {
    font-size: 0.82rem;
    color: #8aaabb;
    margin-top: 3px;
    line-height: 1.3;
}

/* ── Section titles ── */
.section-title {
    font-size: 0.68rem;
    font-weight: 600;
    color: #7ab4c4;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    border-bottom: 1px solid #1a2a3a;
    padding-bottom: 8px;
    margin: 36px 0 18px 0;
}

/* ── Prompt box ── */
.prompt-box {
    background: #0d1117;
    border: 1px solid #1c2430;
    border-left: 2px solid #56b4d3;
    border-radius: 6px;
    padding: 16px;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.82rem;
    color: #8a9ab0;
    white-space: pre-wrap;
    word-break: break-word;
    line-height: 1.7;
}

/* ── Help / info box ── */
.help-box {
    background: #0d1117;
    border: 1px solid #1c2430;
    border-radius: 6px;
    padding: 14px 16px;
    font-size: 0.82rem;
    color: #5a7a8a;
    margin-top: 8px;
    line-height: 1.6;
}

/* ── Tag ── */
.tag {
    display: inline-block;
    background: #0d1520;
    border: 1px solid #1a2a3a;
    border-radius: 4px;
    padding: 2px 8px;
    font-size: 0.68rem;
    font-weight: 500;
    color: #4a7a8a;
    letter-spacing: 0.04em;
    margin: 2px;
}

/* ── Dataframe ── */
[data-testid="stDataFrame"] { border-radius: 6px; overflow: hidden; }

/* ── Expander ── */
[data-testid="stExpander"] {
    border: 1px solid #1a2a3a;
    border-radius: 6px;
    margin-bottom: 6px;
}

/* ── Tabs ── */
[data-testid="stTabs"] [data-baseweb="tab"] {
    font-size: 0.78rem;
    font-weight: 500;
    letter-spacing: 0.04em;
}
/* ── Text / Number input fields ── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {
    background-color: #0d1a28 !important;
    color: #c8d0d8 !important;
    border: 1px solid #1c3a4a !important;
    border-radius: 6px !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {
    border-color: #56b4d3 !important;
    box-shadow: 0 0 0 1px #56b4d3 !important;
}
[data-testid="stNumberInput"] button {
    background-color: #0d1a28 !important;
    border-color: #1c3a4a !important;
    color: #6a9aaa !important;
}

/* ── Radio: 選択時の赤を青に ── */
div[data-baseweb="radio"] [data-checked="true"] { border-color: #56b4d3 !important; }
div[data-baseweb="radio"] [data-checked="true"] div { background-color: #56b4d3 !important; }
div[data-baseweb="radio"] input:checked ~ * { border-color: #56b4d3 !important; }
/* フォーカスリング */
div[data-baseweb="radio"] [data-focused="true"] { box-shadow: 0 0 0 3px rgba(86,180,211,0.3) !important; }



/* ── Download buttons ── */
div.stDownloadButton > button {
    width: 72px !important;
    height: 30px !important;
    padding: 0 !important;
    background: #0d1f2d !important;
    color: #a8dadc !important;
    border: 1.5px solid #56b4d3 !important;
    border-radius: 6px !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.04em !important;
    text-align: center !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    transition: background 0.2s, color 0.2s !important;
}
div.stDownloadButton > button:hover {
    background: #56b4d3 !important;
    color: #0d1f2d !important;
}
div.stDownloadButton {
    display: inline-block !important;
    margin-right: 0.2rem !important;
}

/* ── Sample selectbox font ── */
[data-testid="stSidebar"] [data-testid="stSelectbox"] select,
[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
    font-size: 0.82rem !important;
}

/* ── Radio & Slider accent color (override Streamlit red) ── */
/* ── Radio label ── */
[data-testid="stRadio"] > label {
    color: #56b4d3 !important;
    font-size: 0.78rem !important;
}
[data-testid="stRadio"] label div p {
    color: #c8d0d8 !important;
    font-size: 0.82rem !important;
}

/* Radio: 選択済みの塗り・ボーダー */
div[data-baseweb="radio"] div { background-color: #56b4d3 !important; border-color: #56b4d3 !important; }
div[data-baseweb="radio"] [data-checked="true"] div { background-color: #56b4d3 !important; border-color: #56b4d3 !important; }
div[data-baseweb="radio"] input:checked + div { background-color: #56b4d3 !important; border-color: #56b4d3 !important; }
[data-testid="stRadio"] [aria-checked="true"] > div { background-color: #56b4d3 !important; border-color: #56b4d3 !important; }

/* Slider: トラック・ハンドル・塗り済みトラック */
div[data-baseweb="slider"] [role="slider"] { background-color: #56b4d3 !important; border-color: #56b4d3 !important; }
div[data-baseweb="slider"] [data-testid="stSlider"] div { background-color: #56b4d3 !important; }
div[data-baseweb="slider"] div[class*="Track"] > div { background-color: #56b4d3 !important; }
div[data-baseweb="slider"] div[class*="InnerTrack"] { background-color: #56b4d3 !important; }
/* Streamlit 1.x のスライダー塗り */
.stSlider [data-baseweb="slider"] > div > div > div:nth-child(2) { background: #56b4d3 !important; }
.stSlider [data-baseweb="slider"] > div > div > div:last-child { background-color: #56b4d3 !important; border-color: #56b4d3 !important; }

/* ── Sidebar text colors ── */
[data-testid="stSidebar"] label { color: #c8d0d8 !important; }
[data-testid="stSidebar"] p { color: #c8d0d8 !important; }
[data-testid="stSidebar"] .stCaption p { color: #3a6a7a !important; font-size: 0.78rem !important; }
[data-testid="stSidebar"] .stRadio label p { color: #c8d0d8 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #56b4d3 !important; }

/* ── Toggle font size ── */
[data-testid="stSidebar"] [data-testid="stToggle"] label p,
[data-testid="stSidebar"] [data-testid="stToggle"] > label,
[data-testid="stSidebar"] [data-testid="stToggle"] span:not([data-testid]),
[data-testid="stSidebar"] .stToggle label {
    font-size: 0.82rem !important;
    color: #c8d0d8 !important;
}

/* ── Main area background ── */
.stApp { background-color: #0a0e14 !important; }

/* ── Header / toolbar / top decoration ── */
[data-testid="stHeader"] {
    background-color: #0a0e14 !important;
    border-bottom: 1px solid #1a2430 !important;
}
[data-testid="stToolbar"] { background-color: #0a0e14 !important; }
[data-testid="stDecoration"] { background-color: #0a0e14 !important; display: none; }
.stAppHeader, header[data-testid="stHeader"] { background: #0a0e14 !important; }

/* ── Download buttons: sidebar (sample CSV) ── */
[data-testid="stSidebar"] [data-testid="stDownloadButton"] > button,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button {
    background-color: #0d1a28 !important;
    color: #a8c8d8 !important;
    border: 1px solid #1c3a4a !important;
    border-radius: 8px !important;
    width: 100% !important;
    height: auto !important;
    font-size: 0.78rem !important;
    font-weight: 400 !important;
    line-height: 1.4 !important;
    padding: 6px 10px !important;
    letter-spacing: 0 !important;
    text-align: center !important;
    display: block !important;
    white-space: normal !important;
}
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button p,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button span,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button div {
    font-size: 0.78rem !important;
    font-weight: 400 !important;
    letter-spacing: 0 !important;
    border: none !important;
    background: none !important;
    padding: 0 !important;
    margin: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stDownloadButton"] > button:hover,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button:hover {
    background-color: #112030 !important;
    border-color: #56b4d3 !important;
    color: #56b4d3 !important;
}

/* ── Download buttons: main area (export) ── */
.main [data-testid="stDownloadButton"] > button,
.main [data-testid="stDownloadButton"] button {
    background-color: #0d1a28 !important;
    color: #a8c8d8 !important;
    border: 1px solid #1c3a4a !important;
    border-radius: 8px !important;
}
.main [data-testid="stDownloadButton"] > button:hover,
.main [data-testid="stDownloadButton"] button:hover {
    background-color: #112030 !important;
    border-color: #56b4d3 !important;
    color: #56b4d3 !important;
}

/* ── File uploader ── */
section[data-testid="stFileUploader"],
[data-testid="stFileUploader"],
[data-testid="stFileUploader"] > div,
[data-testid="stFileUploader"] > div > div,
[data-testid="stFileUploader"] > section {
    background-color: #0d1520 !important;
    border-radius: 8px !important;
}
[data-testid="stFileUploadDropzone"] {
    background-color: #0d1520 !important;
    border: 1px dashed #1c3a4a !important;
    border-radius: 8px !important;
}
[data-testid="stFileUploadDropzone"] p,
[data-testid="stFileUploadDropzone"] span,
[data-testid="stFileUploadDropzone"] small { color: #6a9aaa !important; }
[data-testid="stFileUploadDropzone"] svg { fill: #3a6a7a !important; }
[data-testid="stFileUploaderFile"] {
    background-color: #0d1a28 !important;
    border-radius: 6px !important;
}
[data-testid="stFileUploaderFileName"] { color: #a8c8d8 !important; }

/* ── Inline code badges (end_date, last_purchase_date etc.) ── */
code, .stMarkdown code {
    background-color: #0d1f2d !important;
    color: #56b4d3 !important;
    border: 1px solid #1a3a4a !important;
    border-radius: 4px !important;
    padding: 1px 5px !important;
    font-size: 0.82em !important;
}

/* ── Browse files button ── */
[data-testid="stFileUploadDropzone"] button {
    background-color: #0d2030 !important;
    color: #56b4d3 !important;
    border: 1px solid #1c3a4a !important;
    border-radius: 6px !important;
}

/* ── Plotly chart top padding reset ── */
[data-testid="stPlotlyChart"] {
    margin-top: -12px !important;
}

/* ── section-title直後のPlotlyグラフ用 ── */
.section-title-tight {
    font-size: 0.68rem;
    font-weight: 600;
    color: #7ab4c4;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    border-bottom: 1px solid #1a2a3a;
    padding-bottom: 8px;
    margin: 36px 0 0 0;
}
</style>
""", unsafe_allow_html=True)

# ── Matplotlib theme ──────────────────────────────────────────
plt.style.use('dark_background')
rcParams['font.family'] = 'DejaVu Sans'
for _k, _v in {
    'figure.facecolor': '#111820', 'axes.facecolor': '#111820',
    'axes.edgecolor': '#1a3040', 'axes.labelcolor': '#7ab8cc',
    'xtick.color': '#4a8a9a', 'ytick.color': '#4a8a9a',
    'grid.color': '#1a3040', 'grid.linewidth': 0.5,
    'axes.unicode_minus': False,
}.items():
    rcParams[_k] = _v


ACCENT  = '#56b4d3'
ACCENT2 = '#a8dadc'
ACCENT3 = '#1d6fa4'

# ══════════════════════════════════════════════════════════════
# Analysis functions
# ══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def compute_km(duration_arr, event_arr):
    """Kaplan-Meier推定（配列を受け取りキャッシュ対応）"""
    S, records = 1.0, []
    for t in np.sort(np.unique(duration_arr)):
        n_risk   = (duration_arr >= t).sum()
        n_events = ((duration_arr == t) & (event_arr == 1)).sum()
        if n_risk > 0:
            S *= (1 - n_events / n_risk)
        records.append({'t': t, 'S': S, 'n_events': n_events, 'n_risk': n_risk})
    return pd.DataFrame(records)

@st.cache_data(show_spinner=False)
def fit_weibull(t_arr, S_arr):
    """Weibullフィッティング（配列を受け取りキャッシュ対応）"""
    mask = (t_arr > 0) & (S_arr > 0) & (S_arr < 1)
    t_f, S_f = t_arr[mask], S_arr[mask]
    ln_t        = np.log(t_f)
    ln_neg_ln_S = np.log(-np.log(S_f))
    valid = np.isfinite(ln_t) & np.isfinite(ln_neg_ln_S)
    ln_t, ln_neg_ln_S = ln_t[valid], ln_neg_ln_S[valid]
    if len(ln_t) < 3:
        return None, None, None, None
    k, b, r, *_ = stats.linregress(ln_t, ln_neg_ln_S)
    lam = np.exp(-b / k)
    fd = pd.DataFrame({'ln_t': ln_t, 'ln_neg_ln_S': ln_neg_ln_S})
    return k, lam, r**2, fd

def _compute_km_df(df):
    """DataFrameからKM計算（内部ヘルパー）"""
    return compute_km(df['duration'].values, df['event'].values)

def _fit_weibull_df(km_df):
    """KM DataFrameからWeibullフィット（内部ヘルパー）"""
    return fit_weibull(km_df['t'].values.astype(float), km_df['S'].values.astype(float))

@st.cache_data(show_spinner=False)
def load_and_preprocess_csv(file_bytes, dormancy_days, billing_cycle, business_type):
    """CSVの読み込みと前処理をキャッシュ化（同じファイル・設定なら再計算しない）"""
    import io as _io
    df_raw = pd.read_csv(_io.BytesIO(file_bytes))
    df_raw.columns = df_raw.columns.str.strip().str.lower()

    col_map = {}
    for c in df_raw.columns:
        if any(x in c for x in ['id','customer','顧客']):       col_map.setdefault('customer_id', c)
        if any(x in c for x in ['start','開始','契約']):        col_map.setdefault('start_date', c)
        if any(x in c for x in ['end','解約','終了','cancel']):  col_map.setdefault('end_date', c)
        if any(x in c for x in ['revenue','売上','arpu','rev','price','amount']):
            col_map.setdefault('revenue', c)

    missing = [k for k in ['start_date','end_date','revenue'] if k not in col_map]
    if missing:
        return None, missing, None, None, None, None

    for c in df_raw.columns:
        if any(x in c for x in ['last','最終','purchase','購買','購入']):
            col_map.setdefault('last_purchase_date', c)

    df = df_raw.rename(columns={v: k for k, v in col_map.items()})
    df['start_date'] = pd.to_datetime(df['start_date'], errors='coerce')
    df['end_date']   = pd.to_datetime(df['end_date'],   errors='coerce')
    if 'last_purchase_date' in df.columns:
        df['last_purchase_date'] = pd.to_datetime(df['last_purchase_date'], errors='coerce')
    else:
        df['last_purchase_date'] = pd.NaT

    n_input = len(df)
    bad_dates = df['start_date'].isna().sum()
    df = df.dropna(subset=['start_date'])

    today = pd.Timestamp.today()
    dates_for_today = [df['start_date'].max()]
    if df['last_purchase_date'].notna().any():
        dates_for_today.append(df['last_purchase_date'].max())
    if df['end_date'].notna().any():
        dates_for_today.append(df['end_date'].max())
    ref_date = max(d for d in dates_for_today if pd.notna(d))
    if ref_date <= today:
        today = ref_date
    n_dormant = 0
    if dormancy_days is not None and df['last_purchase_date'].notna().any():
        dormant_mask = (
            df['end_date'].isna() &
            df['last_purchase_date'].notna() &
            ((today - df['last_purchase_date']).dt.days > dormancy_days)
        )
        n_dormant = dormant_mask.sum()
        # 離脱日 = last_purchase_date + dormancy_days（休眠判定日が実質の離脱日）
        df.loc[dormant_mask, 'end_date'] = (
            df.loc[dormant_mask, 'last_purchase_date'] + pd.Timedelta(days=dormancy_days)
        )

    def get_end(row):
        if pd.notna(row['end_date']):
            return row['end_date'], 1
        if dormancy_days is not None and pd.notna(row['last_purchase_date']):
            if (today - row['last_purchase_date']).days > dormancy_days:
                return row['last_purchase_date'] + pd.Timedelta(days=dormancy_days), 1
        return today, 0

    result = df.apply(get_end, axis=1, result_type='expand')
    df['end_resolved'] = result[0]
    df['event']        = result[1]
    df['duration']     = (df['end_resolved'] - df['start_date']).dt.days

    # サブスクの最低契約期間を保証
    if business_type != BIZ_SPOT:
        if billing_cycle == BILLING_CALENDAR_MONTHLY:
            min_dur = 30
        elif billing_cycle == BILLING_FIXED_30:
            min_dur = 30
        elif billing_cycle == BILLING_ANNUAL_365:
            min_dur = 365
        elif billing_cycle == BILLING_CUSTOM_DAYS:
            min_dur = custom_cycle_days if custom_cycle_days else 30
        else:
            min_dur = 30
        df['duration'] = df['duration'].clip(lower=min_dur)

    n_corrected = (df['duration'] == 0).sum()
    df['duration'] = df['duration'].clip(lower=1)
    n_excluded = (df['duration'] < 0).sum()
    df = df[df['duration'] > 0]

    df['revenue_total'] = pd.to_numeric(df['revenue'], errors='coerce')

    meta = {
        'n_input': n_input,
        'n_dormant': n_dormant,
        'n_corrected': n_corrected,
        'n_excluded': n_excluded,
        'bad_dates': bad_dates,
    }
    return df, None, meta, billing_cycle, business_type, col_map
    surv_int = lam * gamma(1 + 1/k)
    return surv_int * arpu, surv_int

def ltv_inf(k, lam, arpu):
    surv_int = lam * gamma(1 + 1/k)
    return surv_int * arpu, surv_int

def ltv_horizon(k, lam, arpu, h):
    x = (h / lam) ** k
    return lam * gamma(1 + 1/k) * gammainc(1 + 1/k, x) * arpu

def weibull_s(t, k, lam):
    return np.exp(-(t / lam) ** k)

# ══════════════════════════════════════════════════════════════
# Sidebar
# ══════════════════════════════════════════════════════════════

with st.sidebar:
    _lang_options = {'日本語': 'ja', 'English': 'en'}
    # ブラウザ言語が日本語の場合のみデフォルト日本語、それ以外は英語
    if '_browser_lang_checked' not in st.session_state:
        st.session_state['_browser_lang_checked'] = True
        st.components.v1.html(
            """<script>
            const lang = (navigator.language || navigator.userLanguage || 'en').toLowerCase();
            const isJa = lang.startsWith('ja');
            const params = new URLSearchParams(window.parent.location.search);
            if (!params.has('lang')) {
                params.set('lang', isJa ? 'ja' : 'en');
                window.parent.history.replaceState({}, '', '?' + params.toString());
                window.parent.location.reload();
            }
            </script>""",
            height=0,
        )
    _browser_lang = st.query_params.get('lang', 'en')
    _default_idx = 0 if _browser_lang == 'ja' else 1

    # ── デモモード案内（言語設定の上） ──
    if APP_MODE == 'demo':
        _demo_lang = 'ja' if _default_idx == 0 else 'en'
        _PURCHASE_URL = 'https://example.com/purchase'  # TODO: 購入ページURL確定後に差し替え
        if _demo_lang == 'ja':
            _demo_note = 'デモモード：サンプルデータで主要機能をお試しいただけます。一部の設定はサンプル用に固定されています。ご購入後はご自身のデータで全ての設定を自由にカスタマイズいただけます。'
            _purchase_label = '購入はこちら'
        else:
            _demo_note = 'Demo Mode: Try key features with sample data. Some settings are fixed for demo purposes. After purchase, all settings are fully customizable with your own data.'
            _purchase_label = 'Purchase'
        st.markdown(f"""<div style='background:linear-gradient(135deg, #0d1a28 0%, #142030 100%);
            border:1px solid #1c3a4a; border-radius:8px; padding:12px 14px; margin:8px 0 12px 0;
            font-size:0.73rem; color:#7ab4c4; line-height:1.6;'>
            {_demo_note}<br>
            <a href="{_PURCHASE_URL}" target="_blank" style="display:inline-block; margin-top:8px;
            padding:6px 18px; background:#56b4d3; color:#0a0e14; border-radius:5px;
            font-size:0.73rem; font-weight:600; text-decoration:none; letter-spacing:0.04em;">
            {_purchase_label}</a></div>""", unsafe_allow_html=True)

    _lang_sel = st.selectbox("Language", list(_lang_options.keys()), index=_default_idx)
    LANG = _lang_options[_lang_sel]
    set_lang(LANG)
    # ここから全てのT()が正しい言語で返る
    _cur_default = LANG_DEFAULTS.get(LANG, 'JPY')
    _cur_options = list(CURRENCIES.keys())
    _cur_idx = _cur_options.index(_cur_default) if _cur_default in _cur_options else 0
    if APP_MODE == 'demo':
        CUR = _cur_default  # Demo: 言語連動で固定
        st.selectbox(T('sidebar_cur_label'), [_cur_default], index=0, disabled=True)
    else:
        CUR = st.selectbox(T('sidebar_cur_label'), _cur_options, index=_cur_idx)

    st.markdown(T('sidebar_data_input'))

    # ══════════════════════════════════════════════════════
    # サンプルデータ生成
    # ① 動画学習プラットフォーム（サブスク・日割りOFF・月額¥9,800）
    # ② コワーキングスペース（サブスク・日割りON・月額¥19,800）
    # ③ D2C コスメ系EC（都度購入・平均¥5,800・休眠180日）
    # ══════════════════════════════════════════════════════
    n_sample  = 10000
    BASE_DATE = pd.Timestamp('2025-12-31')
    today_ts  = BASE_DATE

    # ── ① 動画学習プラットフォーム（サブスク・日割りOFF）─────
    # 月額¥9,800固定、k=0.85系（初期離脱型）
    # セグメント: channel, age_group, device
    np.random.seed(42)
    SUB1_START = pd.Timestamp('2023-01-01')
    _s1_dates = list(pd.date_range(SUB1_START, BASE_DATE, periods=n_sample))
    np.random.shuffle(_s1_dates)
    s1_starts = _s1_dates[:n_sample]
    MONTHLY_1 = 49 if CUR != 'JPY' else 9800

    # セグメント割り当て（チャネルごとにWeibullパラメータを変える）
    if get_lang() == 'en':
        _s1_ch_vals  = ['Paid Search', 'Social Ads', 'Word of Mouth', 'Organic Search', 'Other']
    else:
        _s1_ch_vals  = ['検索広告', 'SNS広告', '口コミ', '自然検索', 'その他']
    _s1_ch_prob  = [0.28, 0.30, 0.18, 0.19, 0.05]
    _s1_ch_k     = dict(zip(_s1_ch_vals, [0.93, 0.90, 0.98, 1.00, 0.92]))
    _s1_ch_lam   = dict(zip(_s1_ch_vals, [280, 250, 310, 330, 270]))
    s1_channel   = np.random.choice(_s1_ch_vals, n_sample, p=_s1_ch_prob)

    if get_lang() == 'en':
        _s1_age_vals = ['18-24', '25-34', '35-44', '45+', 'Unknown']
    else:
        _s1_age_vals = ['18-24', '25-34', '35-44', '45+', '不明']
    _s1_age_prob = [0.15, 0.35, 0.28, 0.17, 0.05]
    _s1_age_k    = dict(zip(_s1_age_vals, [0.92, 0.95, 0.96, 0.94, 0.95]))
    _s1_age_lam  = dict(zip(_s1_age_vals, [250, 300, 320, 280, 190]))
    s1_age       = np.random.choice(_s1_age_vals, n_sample, p=_s1_age_prob)

    if get_lang() == 'en':
        _s1_dev_vals = ['PC', 'Mobile', 'Tablet', 'Other']
    else:
        _s1_dev_vals = ['PC', 'スマホ', 'タブレット', 'その他']
    _s1_dev_prob = [0.35, 0.45, 0.15, 0.05]
    s1_device    = np.random.choice(_s1_dev_vals, n_sample, p=_s1_dev_prob)

    # 解約率 85%
    s1_churned = np.random.random(n_sample) < 0.85
    s1_end     = []
    s1_rev     = []
    for i in range(n_sample):
        sd  = s1_starts[i]
        ch  = s1_channel[i]
        ag  = s1_age[i]
        k_i = (_s1_ch_k[ch] + _s1_age_k[ag]) / 2
        l_i = (_s1_ch_lam[ch] + _s1_age_lam[ag]) / 2
        surv = np.random.weibull(k_i) * l_i
        if s1_churned[i]:
            ed = sd + pd.Timedelta(days=max(1, int(surv)))
            if ed <= BASE_DATE:
                s1_end.append(ed.strftime('%Y-%m-%d'))
                months = max(1, (ed - sd).days // 30 + 1)
            else:
                s1_end.append('')
                months = max(1, (BASE_DATE - sd).days // 30 + 1)
        else:
            s1_end.append('')
            months = max(1, (BASE_DATE - sd).days // 30 + 1)
        s1_rev.append(MONTHLY_1 * months)

    sample_elearn = pd.DataFrame({
        'customer_id': [f'EL{i:05d}' for i in range(1, n_sample+1)],
        'start_date':  [d.strftime('%Y-%m-%d') for d in s1_starts],
        'end_date':    s1_end,
        'revenue_total': s1_rev,
        'channel':     s1_channel,
        'age_group':   s1_age,
        'device':      s1_device,
    })

    # ── ② コワーキングスペース（サブスク・日割りON）──────
    # 月額¥19,800固定、k=1.1系（逓増離脱型）
    # セグメント: channel, age_group, occupation
    np.random.seed(43)
    SUB2_START = pd.Timestamp('2023-01-01')
    _s2_dates = list(pd.date_range(SUB2_START, BASE_DATE, periods=n_sample))
    np.random.shuffle(_s2_dates)
    s2_starts = _s2_dates[:n_sample]
    MONTHLY_2 = 199 if CUR != 'JPY' else 19800

    if get_lang() == 'en':
        _s2_ch_vals  = ['Paid Search', 'Social Ads', 'Referral', 'Google Maps', 'Other']
    else:
        _s2_ch_vals  = ['検索広告', 'SNS広告', '紹介', 'Googleマップ', 'その他']
    _s2_ch_prob  = [0.25, 0.22, 0.23, 0.25, 0.05]
    _s2_ch_k     = dict(zip(_s2_ch_vals, [1.08, 1.05, 1.15, 1.12, 1.06]))
    _s2_ch_lam   = dict(zip(_s2_ch_vals, [280, 250, 350, 320, 260]))
    s2_channel   = np.random.choice(_s2_ch_vals, n_sample, p=_s2_ch_prob)

    if get_lang() == 'en':
        _s2_age_vals = ['18-24', '25-34', '35-44', '45+', 'Unknown']
    else:
        _s2_age_vals = ['18-24', '25-34', '35-44', '45+', '不明']
    _s2_age_prob = [0.12, 0.35, 0.30, 0.18, 0.05]
    _s2_age_k    = dict(zip(_s2_age_vals, [1.03, 1.13, 1.16, 1.08, 0.98]))
    _s2_age_lam  = dict(zip(_s2_age_vals, [220, 320, 350, 280, 180]))
    s2_age       = np.random.choice(_s2_age_vals, n_sample, p=_s2_age_prob)

    if get_lang() == 'en':
        _s2_occ_vals = ['Employee', 'Self-employed', 'Student', 'Other']
    else:
        _s2_occ_vals = ['会社員', '経営者・自営など', '学生', 'その他']
    _s2_occ_prob = [0.35, 0.30, 0.20, 0.15]
    _s2_occ_k    = dict(zip(_s2_occ_vals, [1.10, 1.15, 1.00, 1.05]))
    _s2_occ_lam  = dict(zip(_s2_occ_vals, [290, 360, 180, 250]))
    s2_occupation = np.random.choice(_s2_occ_vals, n_sample, p=_s2_occ_prob)

    s2_churned = np.random.random(n_sample) < 0.85
    s2_end     = []
    s2_rev     = []
    for i in range(n_sample):
        sd  = s2_starts[i]
        ch  = s2_channel[i]
        occ = s2_occupation[i]
        ag  = s2_age[i]
        # チャネル・occupation・age_groupのパラメータを平均して使用
        k_i = (_s2_ch_k[ch] + _s2_occ_k[occ] + _s2_age_k[ag]) / 3
        l_i = (_s2_ch_lam[ch] + _s2_occ_lam[occ] + _s2_age_lam[ag]) / 3
        surv = np.random.weibull(k_i) * l_i
        if s2_churned[i]:
            ed = sd + pd.Timedelta(days=max(1, int(surv)))
            if ed <= BASE_DATE:
                s2_end.append(ed.strftime('%Y-%m-%d'))
                days = max(1, (ed - sd).days)
            else:
                s2_end.append('')
                days = max(1, (BASE_DATE - sd).days)
        else:
            s2_end.append('')
            days = max(1, (BASE_DATE - sd).days)
        # 日割り計算：月額 × 日数 / 30
        s2_rev.append(round(MONTHLY_2 * days / 30, 0))

    sample_cowork = pd.DataFrame({
        'customer_id': [f'CW{i:05d}' for i in range(1, n_sample+1)],
        'start_date':  [d.strftime('%Y-%m-%d') for d in s2_starts],
        'end_date':    s2_end,
        'revenue_total': s2_rev,
        'channel':     s2_channel,
        'age_group':   s2_age,
        'occupation':  s2_occupation,
    })

    # ── ③ D2C コスメ系EC（都度購入）────────────────
    # 平均単価¥5,800、k=0.75系（初期離脱型）、購入間隔平均60日（ばらつきあり）
    # 休眠判定180日推奨
    # セグメント: channel, age_group, gender
    np.random.seed(44)
    OBS_START = pd.Timestamp('2021-01-01')
    _s3_dates = list(pd.date_range(OBS_START, BASE_DATE, periods=n_sample))
    np.random.shuffle(_s3_dates)
    s3_starts = _s3_dates[:n_sample]
    _single_cutoff = BASE_DATE - pd.Timedelta(days=180)
    UNIT_PRICE_3 = 38 if CUR != 'JPY' else 5800

    if get_lang() == 'en':
        _s3_ch_vals  = ['Paid Search', 'Instagram Ads', 'Influencer Marketing', 'Organic Search', 'Other']
    else:
        _s3_ch_vals  = ['検索広告', 'Instagram広告', 'インフルエンサーマーケ', '自然検索', 'その他']
    _s3_ch_prob  = [0.25, 0.30, 0.15, 0.25, 0.05]
    _s3_ch_k     = dict(zip(_s3_ch_vals, [0.87, 0.83, 0.78, 0.95, 0.84]))
    _s3_ch_lam   = dict(zip(_s3_ch_vals, [260, 240, 180, 330, 245]))
    # チャネル別の単発率（インフルエンサー経由は単発が多い）
    _s3_ch_single = dict(zip(_s3_ch_vals, [0.50, 0.55, 0.70, 0.40, 0.55]))
    s3_channel   = np.random.choice(_s3_ch_vals, n_sample, p=_s3_ch_prob)

    if get_lang() == 'en':
        _s3_age_vals = ['18-24', '25-34', '35-44', '45-54', '55+', 'Unknown']
    else:
        _s3_age_vals = ['18-24', '25-34', '35-44', '45-54', '55+', '不明']
    _s3_age_prob = [0.10, 0.30, 0.28, 0.18, 0.09, 0.05]
    _s3_age_k    = dict(zip(_s3_age_vals, [0.82, 0.89, 0.92, 0.86, 0.83, 0.78]))
    _s3_age_lam  = dict(zip(_s3_age_vals, [210, 280, 300, 250, 220, 190]))
    s3_age       = np.random.choice(_s3_age_vals, n_sample, p=_s3_age_prob)

    if get_lang() == 'en':
        _s3_gen_vals = ['Female', 'Male', 'Not specified']
    else:
        _s3_gen_vals = ['女性', '男性', '未回答']
    _s3_gen_prob = [0.58, 0.37, 0.05]
    _s3_gen_k    = dict(zip(_s3_gen_vals, [0.90, 0.82, 0.85]))
    _s3_gen_lam  = dict(zip(_s3_gen_vals, [290, 210, 250]))
    s3_gender    = np.random.choice(_s3_gen_vals, n_sample, p=_s3_gen_prob)

    # リピートのうちアクティブ18%
    s3_active = np.random.random(n_sample) < 0.18

    s3_last_purchase = []
    s3_revenues      = []
    for i in range(n_sample):
        sd    = s3_starts[i]
        ch    = s3_channel[i]
        gen   = s3_gender[i]
        ag    = s3_age[i]
        k_i   = (_s3_ch_k[ch] + _s3_gen_k[gen] + _s3_age_k[ag]) / 3
        l_i   = (_s3_ch_lam[ch] + _s3_gen_lam[gen] + _s3_age_lam[ag]) / 3
        single_rate = _s3_ch_single[ch]
        is_single = np.random.random() < single_rate
        # 購入間隔：平均60日、標準偏差20日でばらつき
        avg_interval = max(20, int(np.random.normal(60, 20)))
        _price_var = 7 if CUR != 'JPY' else 1000
        _price_min = 20 if CUR != 'JPY' else 3000
        _price_max = 65 if CUR != 'JPY' else 10000
        price = UNIT_PRICE_3 + int(np.random.normal(0, _price_var))
        price = max(_price_min, min(_price_max, price))
        if is_single and sd <= _single_cutoff:
            lp        = sd
            purchases = 1
        elif s3_active[i] or (is_single and sd > _single_cutoff):
            days_since = np.random.randint(1, 180)
            lp         = BASE_DATE - pd.Timedelta(days=int(days_since))
            lp         = max(lp, sd + pd.Timedelta(days=1))
            purchases  = min(max(2, round((lp - sd).days / avg_interval)), 20)
        else:
            surv_days = max(1, int(np.random.weibull(k_i) * l_i))
            lp        = sd + pd.Timedelta(days=surv_days)
            lp        = min(lp, BASE_DATE - pd.Timedelta(days=1))
            lp        = max(lp, sd + pd.Timedelta(days=1))
            purchases = min(max(2, round((lp - sd).days / avg_interval)), 20)
        s3_last_purchase.append(lp.strftime('%Y-%m-%d'))
        s3_revenues.append(price * purchases)

    sample_skincare = pd.DataFrame({
        'customer_id':        [f'SC{i:05d}' for i in range(1, n_sample+1)],
        'start_date':         [d.strftime('%Y-%m-%d') for d in s3_starts],
        'end_date':           '',
        'last_purchase_date': s3_last_purchase,
        'revenue_total':      s3_revenues,
        'channel':            s3_channel,
        'age_group':          s3_age,
        'gender':             s3_gender,
    })

    import base64
    elearn_csv   = sample_elearn.to_csv(index=False).encode('utf-8-sig')
    cowork_csv   = sample_cowork.to_csv(index=False).encode('utf-8-sig')
    skincare_csv = sample_skincare.to_csv(index=False).encode('utf-8-sig')

    if APP_MODE == 'demo':
        st.markdown(f"<span style='color:#c8d0d8; font-size:0.78rem;'>{T('sidebar_sample_hint')}</span>", unsafe_allow_html=True)

    # サンプルデータをセッションステートで管理
    if 'sample_df' not in st.session_state:
        st.session_state.sample_df = None
    if 'sample_label' not in st.session_state:
        st.session_state.sample_label = None

    _sample_options = {
        T('sample_elearn'):   ('elearn', sample_elearn),
        T('sample_cowork'):   ('cowork', sample_cowork),
        T('sample_skincare'): ('skincare', sample_skincare),
    }

    if APP_MODE == 'demo':
        _btn_s = "display:block; width:100%; text-align:center; text-decoration:none; background:#0d1a28; color:#a8c8d8; border:1px solid #1c3a4a; border-radius:8px; padding:8px 6px; font-size:0.75rem; line-height:1.5; box-sizing:border-box;"
        _selected_sample = st.selectbox(
            T('sidebar_sample_select'),
            [T('sidebar_sample_placeholder')] + list(_sample_options.keys()),
            key='sample_select',
            label_visibility='collapsed'
        )
    if APP_MODE == 'demo':
        if _selected_sample != T('sidebar_sample_placeholder') and st.session_state.get('_prev_sample') != _selected_sample:
            st.session_state._prev_sample = _selected_sample
            _key, _df = _sample_options[_selected_sample]
            st.session_state.sample_df = _df
            st.session_state.sample_label = _selected_sample
            # サンプルに応じてデフォルト設定をセッションに保存
            if _key == 'elearn':
                st.session_state['_sample_biz']     = BIZ_SUBSCRIPTION
                st.session_state['_sample_prorate'] = False
                st.session_state['_sample_seg']     = 'channel, age_group, device'
                st.session_state['_sample_report_title']  = T('sample_elearn_title')
                st.session_state['_sample_client_name']   = T('sample_elearn_client')
                st.session_state['_sample_analyst_name']  = T('sample_elearn_analyst')
            elif _key == 'cowork':
                st.session_state['_sample_biz']     = BIZ_SUBSCRIPTION
                st.session_state['_sample_prorate'] = True
                st.session_state['_sample_seg']     = 'channel, age_group, occupation'
                st.session_state['_sample_report_title']  = T('sample_cowork_title')
                st.session_state['_sample_client_name']   = T('sample_cowork_client')
                st.session_state['_sample_analyst_name']  = T('sample_cowork_analyst')
            else:  # skincare
                st.session_state['_sample_biz']     = BIZ_SPOT
                st.session_state['_sample_prorate'] = False
                st.session_state['_sample_seg']     = 'channel, age_group, gender'
                st.session_state['_sample_report_title']  = T('sample_skincare_title')
                st.session_state['_sample_client_name']   = T('sample_skincare_client')
                st.session_state['_sample_analyst_name']  = T('sample_skincare_analyst')
            st.rerun()

        # サンプルデータCSVダウンロード
        _dl_map = {
            'elearn':   ('sample_elearn.csv',   elearn_csv),
            'cowork':   ('sample_cowork.csv',   cowork_csv),
            'skincare': ('sample_cosmetics.csv', skincare_csv),
        }
        _active_sample = _selected_sample if _selected_sample != T('sidebar_sample_placeholder') else st.session_state.get('_prev_sample', None)
        if _active_sample and _active_sample in _sample_options:
            _dl_key = _sample_options[_active_sample][0]
            _dl_fn, _dl_data = _dl_map[_dl_key]
            st.download_button(
                label=T('sidebar_sample_dl'),
                data=_dl_data,
                file_name=_dl_fn,
                mime='text/csv',
                key='sample_csv_dl'
            )
    else:
        _selected_sample = T('sidebar_sample_placeholder')

    if APP_MODE != 'demo':
        uploaded = st.file_uploader(T('sidebar_upload_csv'), type=['csv'])
    else:
        uploaded = None

    # サンプルボタン or アップロードでデータを確定
    if uploaded is not None:
        st.session_state.sample_df = None  # アップロード優先
        st.session_state.sample_label = None

    st.markdown(T('sidebar_outlier'))
    _oc1, _oc2 = st.columns(2)
    with _oc1:
        outlier_upper_pct = st.number_input(
            T('sidebar_outlier_upper'), min_value=0.0, max_value=20.0,
            value=0.0, step=0.5, format="%.1f",
            help=T('sidebar_outlier_upper_help'),
        )
    with _oc2:
        outlier_lower_pct = st.number_input(
            T('sidebar_outlier_lower'), min_value=0.0, max_value=20.0,
            value=0.0, step=0.5, format="%.1f",
            help=T('sidebar_outlier_lower_help'),
        )
    outlier_removal = (outlier_upper_pct > 0) or (outlier_lower_pct > 0)
    st.caption(T('sidebar_outlier_caption'))

    _demo_lock = (APP_MODE == 'demo')

    st.markdown(T('sidebar_biz_type'))
    _biz_options_map = {T('biz_subscription'): BIZ_SUBSCRIPTION, T('biz_spot'): BIZ_SPOT}
    _biz_display = list(_biz_options_map.keys())
    _biz_default = st.session_state.get('_sample_biz', BIZ_SUBSCRIPTION)
    # session_stateの内部キーから表示名を逆引き
    _biz_default_display = [k for k, v in _biz_options_map.items() if v == _biz_default]
    _biz_default_display = _biz_default_display[0] if _biz_default_display else _biz_display[0]
    _biz_idx = _biz_display.index(_biz_default_display) if _biz_default_display in _biz_display else 0
    _biz_selected_display = st.radio(
        T('sidebar_biz_type_label'),
        _biz_display,
        index=_biz_idx,
        disabled=_demo_lock,
    )
    business_type = _biz_options_map[_biz_selected_display]

    if business_type == BIZ_SUBSCRIPTION:
        st.caption(T('sidebar_sub_caption'))
        dormancy_days = None  # 休眠判定なし
        _billing_display_map = {
            T('billing_monthly_calendar'): BILLING_CALENDAR_MONTHLY,
            T('billing_annual_365'):       BILLING_ANNUAL_365,
            T('billing_custom_days'):      BILLING_CUSTOM_DAYS,
        }
        _billing_display_list = list(_billing_display_map.keys())
        billing_cycle_display_text = st.radio(
            T('sidebar_billing_period'),
            _billing_display_list,
            index=0,
            disabled=_demo_lock,
        )
        billing_cycle = _billing_display_map[billing_cycle_display_text]

        if billing_cycle == BILLING_CUSTOM_DAYS:
            custom_cycle_days = st.number_input(T('sidebar_custom_cycle_days'), min_value=1, max_value=365, value=30, disabled=_demo_lock)
        else:
            custom_cycle_days = None
        st.caption(T('sidebar_billing_caption'))

        st.markdown(f"<div style='font-size:0.82rem; color:#c8d0d8; margin-bottom:4px;'>{T('sidebar_prorate_label')}</div>", unsafe_allow_html=True)
        prorate_cancel = st.toggle(T('sidebar_prorate_label'), value=st.session_state.get("_sample_prorate", False), label_visibility="collapsed", disabled=_demo_lock)
        st.caption(T('sidebar_prorate_caption'))

    else:  # spot
        st.caption(T('sidebar_spot_caption'))
        billing_cycle = BILLING_DAILY_SPOT
        custom_cycle_days = None
        prorate_cancel = False
        _dormancy_map = {
            T('dormancy_180d'): 180,
            T('dormancy_365d'): 365,
            T('dormancy_730d'): 730,
            T('dormancy_custom'): None,
        }
        dormancy_option = st.radio(
            T('sidebar_dormancy_period'),
            list(_dormancy_map.keys()),
            index=0,
            disabled=_demo_lock,
        )
        st.caption(T('sidebar_dormancy_caption'))
        _dormancy_val = _dormancy_map[dormancy_option]
        if _dormancy_val is None:
            dormancy_days = st.number_input(T('sidebar_dormancy_days_label'), min_value=30, max_value=3650, value=180, disabled=_demo_lock)
        else:
            dormancy_days = _dormancy_val

    horizon_days = 730  # 内部計算用デフォルト

    st.markdown("### Gross Profit Margin (%)")
    # サンプルCSVのファイル名からデフォルトGPMを設定
    _gpm_default = 50
    _sample_label = st.session_state.get('sample_label', '') or ''
    _fn = (uploaded.name.lower() if uploaded is not None else '') + _sample_label.lower()
    if uploaded is not None or _sample_label:
        if 'saas' in _fn or 'subscription' in _fn:
            _gpm_default = 75
        elif 'gym' in _fn:
            _gpm_default = 50
        elif 'supplement' in _fn or 'supp' in _fn:
            _gpm_default = 60
        elif 'fec' in _fn or 'fashion' in _fn or 'spot' in _fn:
            _gpm_default = 40
    gpm = st.slider(T('sidebar_gpm_label'), 0, 100, _gpm_default, 1) / 100
    st.caption(T('sidebar_gpm_caption', gpm=f"{gpm:.0%}"))

    st.markdown(T('sidebar_cac'))
    cac_n = st.slider(T('sidebar_cac_slider'), 1.0, 10.0, 3.0, 0.5)
    cac_label = f"LTV:CAC = {cac_n}:1"
    cac_mode = 'LTV : CAC = N : 1'
    cac_recover_days = None
    st.caption(T('sidebar_cac_caption'))

    if APP_MODE != 'standard':
        st.markdown(T('sidebar_segment'))
        _seg_disabled = (APP_MODE == 'demo')
        segment_cols_input = st.text_input(
            T('sidebar_seg_input_label'),
            value=st.session_state.get('_sample_seg', ''),
            placeholder=T('sidebar_seg_placeholder'),
            disabled=_seg_disabled,
        )
        st.caption(T('sidebar_seg_caption'))
        st.markdown(T('sidebar_display_limit'))
        seg_display_limit = st.slider(
            T('sidebar_display_slider'),
            min_value=1, max_value=10, value=5,
        )
        st.caption(T('sidebar_display_caption'))
    else:
        segment_cols_input = ''
        seg_display_limit = 5

    cac_input = 0
    cac_known = False

    st.markdown("")
    st.markdown(T('sidebar_report_info'))
    # サンプル選択中は、言語切り替え時にデフォルト値を更新
    _active_key = None
    if st.session_state.get('sample_df') is not None and st.session_state.get('_prev_sample'):
        for _lbl, (_k, _) in _sample_options.items():
            if _lbl == st.session_state.get('_prev_sample'):
                _active_key = _k
                break
        if _active_key is None:
            # 言語切り替え後にラベルが変わった場合、キーで探す
            _prev = st.session_state.get('_prev_sample', '')
            for _lbl, (_k, _) in _sample_options.items():
                if _k in ['elearn', 'cowork', 'skincare']:
                    if st.session_state.get('_sample_biz') == BIZ_SPOT and _k == 'skincare':
                        _active_key = _k; break
                    elif st.session_state.get('_sample_prorate') == True and _k == 'cowork':
                        _active_key = _k; break
                    elif st.session_state.get('_sample_prorate') == False and st.session_state.get('_sample_biz') == BIZ_SUBSCRIPTION and _k == 'elearn':
                        _active_key = _k; break
    if _active_key:
        _title_key = f'sample_{_active_key}_title'
        _client_key = f'sample_{_active_key}_client'
        _analyst_key = f'sample_{_active_key}_analyst'
        st.session_state['_sample_report_title'] = T(_title_key)
        st.session_state['_sample_client_name'] = T(_client_key)
        st.session_state['_sample_analyst_name'] = T(_analyst_key)
    report_title = st.text_input(T('sidebar_report_title'), st.session_state.get('_sample_report_title', ''), placeholder=T('sidebar_report_title_ph'))
    client_name  = st.text_input(T('sidebar_client_name'), st.session_state.get('_sample_client_name', ''), placeholder=T('sidebar_client_name_ph'))
    analyst_name = st.text_input(T('sidebar_analyst_name'), st.session_state.get('_sample_analyst_name', ''), placeholder=T('sidebar_analyst_name_ph'))

# ══════════════════════════════════════════════════════════════
# Header
# ══════════════════════════════════════════════════════════════

st.markdown("""
<div style='padding: 16px 0 32px 0; border-bottom: 1px solid #1a2a3a; margin-bottom: 28px;'>
  <div style='font-family: 'BIZ UDPGothic', sans-serif; font-size: 0.8rem; font-weight: 600; letter-spacing: 0.16em; text-transform: uppercase; color: #3a6a7a; margin-bottom: 8px;'>Analytics Tool</div>
  <div style='font-family: 'IBM Plex Mono', monospace; font-size: 1.6rem; font-weight: 500; color: #c8d0d8; letter-spacing: -0.03em; line-height: 1;'>LTV Analyzer <span style='color: #56b4d3;'>""" + ("Demo" if APP_MODE == "demo" else "Standard" if APP_MODE == "standard" else "Advanced") + """</span></div>
  <div style='font-size: 0.78rem; color: #3a5a6a; margin-top: 8px; letter-spacing: 0.02em;'>Kaplan–Meier × Weibull — Segment-level LTV Intelligence &nbsp;·&nbsp; v358</div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# No file → instructions
# ══════════════════════════════════════════════════════════════

if uploaded is None and st.session_state.get('sample_df') is None:
    if APP_MODE == 'demo':
        _no_file_msg = 'サイドバーからサンプルデータを選択してください。3種類のサンプルで全機能をお試しいただけます。' if get_lang() == 'ja' else 'Select sample data from the sidebar. Try all features with 3 sample datasets.'
    else:
        _no_file_msg = T('main_no_file_info')
    st.info(_no_file_msg)

    st.markdown(f"<div class='section-title'>{T('main_csv_format_title')}</div>", unsafe_allow_html=True)
    if get_lang() == 'ja':
        st.markdown("""
| 列名 | 内容 | 形式 | 例 |
|------|------|------|----|
| `customer_id` | 顧客ID | 任意の文字列 | C0001 |
| `start_date` | 契約開始日 / 初回購入日 | YYYY-MM-DD | 2023-01-01 |
| `end_date` | 解約日（サブスク向け・継続中は**空欄**） | YYYY-MM-DD | 2024-03-15 |
| `last_purchase_date` | 最終購買日（都度購入向け・任意） | YYYY-MM-DD | 2024-06-01 |
| `revenue` | **累計売上** | 数値 | 48000 |
| `セグメント列`（任意の列名） | **Advanced機能**：プラン・チャネル・年齢層など | 文字列 | 月額300 |

> **Advanced版では必ずセグメント列を追加してください。**複数列追加可能です。\n
> 列名は完全一致でなくてもOKです。`start`・`end`・`last`・`revenue`を含む列名は自動認識します。\n
> ARPU daily はビジネスタイプに応じて自動計算されます。\n
> セグメント列は1列あたり最大50種類のユニーク値まで対応しています（都道府県47個も対応）。
    """)
    else:
        st.markdown("""
| Column | Description | Format | Example |
|--------|-------------|--------|---------| 
| `customer_id` | Customer ID | Any string | C0001 |
| `start_date` | Start date / first purchase date | YYYY-MM-DD | 2023-01-01 |
| `end_date` | End date (subscription: blank if active) | YYYY-MM-DD | 2024-03-15 |
| `last_purchase_date` | Last purchase date (spot purchase: optional) | YYYY-MM-DD | 2024-06-01 |
| `revenue` | **Cumulative revenue** | Numeric | 48000 |
| Segment columns (any name) | **Advanced**: plan, channel, age group, etc. | String | plan_A |

> **Always include segment columns for the Advanced version.** Multiple columns supported.\n
> Column names don't need to match exactly. Columns containing `start`, `end`, `last`, or `revenue` are auto-detected.\n
> Daily ARPU is automatically calculated based on business type.\n
> Each segment column supports up to 50 unique values.
    """)

    st.markdown(f"<div class='section-title'>{T('main_analysis_flow_title')}</div>", unsafe_allow_html=True)
    cols = st.columns(4)
    steps = [
        (T('main_step1_title'), T('main_step1_desc')),
        (T('main_step2_title'), T('main_step2_desc')),
        (T('main_step3_title'), T('main_step3_desc')),
        (T('main_step4_title'), T('main_step4_desc')),
    ]
    for col, (title, desc) in zip(cols, steps):
        with col:
            st.markdown(f"<div class='metric-card'><div class='metric-value' style='font-size:1.1rem'>{title}</div><div class='metric-label' style='font-size:0.75rem; color:#666; margin-top:8px;'>{desc}</div></div>", unsafe_allow_html=True)
    st.stop()

# query_paramsからサンプル選択を処理
_qp = st.query_params
if 'sample' in _qp and st.session_state.get('sample_df') is None:
    _s = _qp['sample']
    if _s == 'elearn':
        st.session_state.sample_df    = sample_elearn
        st.session_state.sample_label = T('sample_elearn')
    elif _s == 'cowork':
        st.session_state.sample_df    = sample_cowork
        st.session_state.sample_label = T('sample_cowork')
    elif _s == 'skincare':
        st.session_state.sample_df    = sample_skincare
        st.session_state.sample_label = T('sample_skincare')
    _keep_lang = st.query_params.get('lang', None)
    st.query_params.clear()
    if _keep_lang:
        st.query_params['lang'] = _keep_lang

# ══════════════════════════════════════════════════════════════
# Load & validate data
# ══════════════════════════════════════════════════════════════

# デフォルト値（tryブロックが途中終了した場合のフォールバック）
arpu_daily = None
gp_daily   = None

# sample_df優先、なければuploadedから読み込み
_active_df = st.session_state.get('sample_df', None)
if _active_df is not None:
    pass
try:
    if _active_df is not None:
        df_raw = _active_df.copy()
        df_raw.columns = df_raw.columns.str.strip().str.lower()
    else:
        df_raw = pd.read_csv(uploaded)
        df_raw.columns = df_raw.columns.str.strip().str.lower()

    col_map = {}
    for c in df_raw.columns:
        if any(x in c for x in ['id','customer','顧客']):      col_map.setdefault('customer_id', c)
        if any(x in c for x in ['start','開始','契約']):       col_map.setdefault('start_date', c)
        if any(x in c for x in ['end','解約','終了','cancel']): col_map.setdefault('end_date', c)
        if any(x in c for x in ['revenue','売上','arpu','rev','price','amount']):
            col_map.setdefault('revenue', c)

    missing = [k for k in ['start_date','end_date','revenue'] if k not in col_map]
    if missing:
        st.error(T('main_col_missing', missing=str(missing)))
        st.stop()

    # last_purchase_date 列の自動認識
    for c in df_raw.columns:
        if any(x in c for x in ['last','最終','purchase','購買','購入']):
            col_map.setdefault('last_purchase_date', c)

    df = df_raw.rename(columns={v: k for k, v in col_map.items()})
    df['start_date'] = pd.to_datetime(df['start_date'], errors='coerce')
    df['end_date']   = pd.to_datetime(df['end_date'], errors='coerce')
    if 'last_purchase_date' in df.columns:
        df['last_purchase_date'] = pd.to_datetime(df['last_purchase_date'], errors='coerce')
    else:
        df['last_purchase_date'] = pd.NaT

    bad_dates = df['start_date'].isna().sum()
    if bad_dates > 0:
        st.warning(T('main_bad_dates', n=f"{bad_dates:,}"))
    df = df.dropna(subset=['start_date'])

    today = pd.Timestamp.today()
    n_input = len(df)

    # 分析基準日：CSVの最新観測日（last_purchase_dateまたはend_dateの最大値）
    # → 同じCSVなら何日に分析しても結果が変わらない（再現性の確保）
    dates_for_today = [df['start_date'].max()]
    if df['last_purchase_date'].notna().any():
        dates_for_today.append(df['last_purchase_date'].max())
    if df['end_date'].notna().any():
        dates_for_today.append(df['end_date'].max())
    ref_date = max(d for d in dates_for_today if pd.notna(d))
    if ref_date <= today:
        today = ref_date

    # 休眠判定：end_date がなく last_purchase_date が休眠期間を超えている場合は離脱とみなす
    n_dormant = 0
    if dormancy_days is not None and df['last_purchase_date'].notna().any():
        dormant_mask = (
            df['end_date'].isna() &
            df['last_purchase_date'].notna() &
            ((today - df['last_purchase_date']).dt.days > dormancy_days)
        )
        n_dormant = dormant_mask.sum()
        # 離脱日 = last_purchase_date + dormancy_days（休眠判定日が実質の離脱日）
        df.loc[dormant_mask, 'end_date'] = (
            df.loc[dormant_mask, 'last_purchase_date'] + pd.Timedelta(days=dormancy_days)
        )

    # ── duration・event の確定 ──
    def get_end(row):
        if pd.notna(row['end_date']):
            return row['end_date'], 1  # 解約済み
        if dormancy_days is not None and pd.notna(row['last_purchase_date']):
            days_since = (today - row['last_purchase_date']).days
            if days_since > dormancy_days:
                # 離脱日 = last_purchase_date + dormancy_days
                return row['last_purchase_date'] + pd.Timedelta(days=dormancy_days), 1
        return today, 0  # 継続中

    result = df.apply(get_end, axis=1, result_type='expand')
    df['end_resolved'] = result[0]
    df['event']        = result[1]

    if business_type == BIZ_SPOT and 'last_purchase_date' in df.columns:
        # 都度購入型：duration = last_purchase_date - start_date（dormancy_days除外）
        # アクティブ顧客（event=0）：基準日 - start_date
        # 休眠離脱顧客（event=1）：last_purchase_date - start_date + dormancy_days
        def calc_duration_spot(row):
            if row['event'] == 0:
                return (today - row['start_date']).days
            else:
                if pd.notna(row['last_purchase_date']):
                    return (row['last_purchase_date'] - row['start_date']).days + (dormancy_days or 0)
                else:
                    return (row['end_resolved'] - row['start_date']).days
        df['duration'] = df.apply(calc_duration_spot, axis=1)
    else:
        df['duration'] = (df['end_resolved'] - df['start_date']).dt.days

    # 丸め前のdurationを保存（single_churn_rate計算用）
    df['duration_raw'] = df['duration']

    # サブスクの最低契約期間を保証（契約期間未満のdurationを引き上げる）
    if business_type != BIZ_SPOT:
        if billing_cycle == BILLING_ANNUAL_365:
            min_contract = 365
        elif billing_cycle == BILLING_CUSTOM_DAYS and custom_cycle_days:
            min_contract = custom_cycle_days
        else:
            min_contract = 30

        if not prorate_cancel:
            # 日割りなし：durationを契約更新日に丸める
            if billing_cycle == BILLING_CALENDAR_MONTHLY:
                # start_dateから月単位で次の更新日を計算
                import calendar as _cal
                def round_to_renewal(row):
                    sd = row['start_date']
                    dur = row['duration']
                    ed = sd + pd.Timedelta(days=dur)
                    months = 0
                    cur = sd
                    while True:
                        m = cur.month + 1 if cur.month < 12 else 1
                        y = cur.year if cur.month < 12 else cur.year + 1
                        max_d = _cal.monthrange(y, m)[1]
                        nxt = pd.Timestamp(y, m, min(sd.day, max_d))
                        if nxt > ed:
                            break
                        months += 1
                        cur = nxt
                    m2 = cur.month + 1 if cur.month < 12 else 1
                    y2 = cur.year if cur.month < 12 else cur.year + 1
                    max_d2 = _cal.monthrange(y2, m2)[1]
                    renewal = pd.Timestamp(y2, m2, min(sd.day, max_d2))
                    return max((renewal - sd).days, min_contract)
                df['duration'] = df.apply(round_to_renewal, axis=1)
            else:
                import numpy as _np
                df['duration'] = (_np.ceil(df['duration'] / min_contract) * min_contract).astype(int)

            # 最低契約期間未満で解約した顧客 → 打ち切り扱い（event=0）にしてdurationを引き上げ
            short_mask = df['duration'] < min_contract
            df.loc[short_mask, 'event'] = 0
            df.loc[short_mask, 'duration'] = min_contract

    # duration=0 を1日に自動補正
    n_corrected = (df['duration'] == 0).sum()
    df['duration'] = df['duration'].clip(lower=1)

    # duration<0（開始日が未来）は除外
    n_excluded = (df['duration'] < 0).sum()
    df = df[df['duration'] > 0]

    # ── ARPU_daily の計算 ──
    df['revenue_total'] = pd.to_numeric(df['revenue'], errors='coerce')

    def calc_arpu_daily(row):
        rev      = row['revenue_total']
        dur      = row['duration']
        start    = row['start_date']
        end_r    = row['end_resolved']
        if pd.isna(rev) or dur <= 0:
            return np.nan

        if billing_cycle == BILLING_DAILY_SPOT:
            # 都度購入：累計売上 ÷ 継続日数
            return rev / dur

        # 日割りONの場合：実際の日数で割る
        if prorate_cancel:
            return rev / max(dur, 1)

        elif billing_cycle == BILLING_CALENDAR_MONTHLY:
            # 契約開始日の「日」を基準に何ヶ月分更新されたかを数える
            s, e = start, end_r
            renewals = 0
            cur = s
            while True:
                # 翌月の同日を計算
                month = cur.month + 1 if cur.month < 12 else 1
                year  = cur.year if cur.month < 12 else cur.year + 1
                max_day = calendar.monthrange(year, month)[1]
                day   = min(cur.day, max_day)
                nxt   = pd.Timestamp(year, month, day)
                if nxt > e:
                    break
                renewals += 1
                cur = nxt
            renewals = max(renewals, 1)
            # 各月の実日数の平均で日次換算
            total_days = (pd.Timestamp(e.year, e.month,
                          calendar.monthrange(e.year, e.month)[1]) -
                          pd.Timestamp(s.year, s.month, 1)).days / renewals
            avg_days = max(total_days, 1)
            return (rev / renewals) / avg_days

        elif billing_cycle == BILLING_FIXED_30:
            import math
            renewals = max(math.ceil(dur / 30), 1)
            return (rev / renewals) / 30

        elif billing_cycle == BILLING_ANNUAL_365:
            import math
            renewals = max(math.ceil(dur / 365), 1)
            return (rev / renewals) / 365

        else:  # カスタム入力
            import math
            cycle = custom_cycle_days or 30
            renewals = max(math.ceil(dur / cycle), 1)
            return (rev / renewals) / cycle

    df['arpu_daily'] = df.apply(calc_arpu_daily, axis=1)
    df['gp_daily']   = df['arpu_daily'] * gpm
    df = df.dropna(subset=['arpu_daily'])
    df = df[df['arpu_daily'] > 0]

    # ── 異常値除外（オプション）──
    n_outlier = 0
    n_outlier_upper = 0
    n_outlier_lower = 0
    # 除外前の分布データを保持（ヒストグラム表示用）
    _rev_before_outlier = df['revenue_total'].copy()
    if outlier_removal:
        before = len(df)
        # 上位パーセンタイル除外
        if outlier_upper_pct > 0:
            upper_r = df['revenue_total'].quantile(1.0 - outlier_upper_pct / 100.0)
            n_outlier_upper = (df['revenue_total'] >= upper_r).sum()
        else:
            upper_r = df['revenue_total'].max() + 1  # 除外なし
            n_outlier_upper = 0
        # 下位パーセンタイル除外
        if outlier_lower_pct > 0:
            lower_r = df['revenue_total'].quantile(outlier_lower_pct / 100.0)
            n_outlier_lower = (df['revenue_total'] <= lower_r).sum()
        else:
            lower_r = df['revenue_total'].min() - 1  # 除外なし
            n_outlier_lower = 0
        df = df[(df['revenue_total'] > lower_r) & (df['revenue_total'] < upper_r)]
        n_outlier = before - len(df)

    # ARPU計算
    if billing_cycle == BILLING_DAILY_SPOT:
        # ── 都度購入型：ARPU_short / ARPU_long / ARPU_0-dormancy の3段階計算 ──
        _dorm = dormancy_days or 180

        # 基準日（today）
        # 単発顧客：first == last かつ 基準日-last >= dormancy_days（観測完結）
        if 'last_purchase_date' in df.columns:
            _gap_first_last = (df['last_purchase_date'] - df['start_date']).dt.days.fillna(-1)
            _days_since_last = (today - df['last_purchase_date']).dt.days.fillna(0)

            # 単発顧客マスク（first=last かつ観測完結）
            _single_mask = (_gap_first_last == 0) & (_days_since_last >= _dorm)
            # リピート顧客マスク（first≠last、基準日-last問わず）
            _long_mask   = _gap_first_last > 0

            # ARPU_short：単発顧客の総売上 ÷ 総顧客数 ÷ dormancy_days
            _single_df = df[_single_mask]
            if len(_single_df) > 0:
                arpu_short = _single_df['revenue_total'].sum() / len(_single_df) / _dorm
            else:
                arpu_short = 0.0

            # ARPU_long：リピート顧客の総売上 ÷ 総duration（加重平均）
            _long_df = df[_long_mask]
            if len(_long_df) > 0:
                arpu_long = _long_df['revenue_total'].sum() / _long_df['duration'].sum()
            else:
                arpu_long = df['revenue_total'].sum() / df['duration'].sum()

            # 比率計算
            _n_single = len(_single_df)
            _n_long   = len(_long_df)
            _n_total  = _n_single + _n_long
            if _n_total > 0:
                _w_short = _n_single / _n_total
                _w_long  = _n_long  / _n_total
            else:
                _w_short, _w_long = 0.5, 0.5

            # ARPU_0-dormancy：単発とロングの加重平均
            arpu_0_dorm = arpu_short * _w_short + arpu_long * _w_long

            # 全体arpu_dailyはARPU_longを採用（t=180以降の積分用）
            arpu_daily  = arpu_long
        else:
            # last_purchase_dateがない場合はフォールバック
            arpu_short  = df['revenue_total'].sum() / df['duration'].sum()
            arpu_long   = arpu_short
            arpu_0_dorm = arpu_short
            arpu_daily  = arpu_short
            _dorm       = dormancy_days or 180

    elif prorate_cancel:
        # 日割りON：算術平均（revenue_totalが既に日割り計算済みのため）
        arpu_daily  = df['arpu_daily'].mean()
        arpu_short  = arpu_daily
        arpu_long   = arpu_daily
        arpu_0_dorm = arpu_daily
    else:
        # サブスク日割りOFF：billing_cycleで月額÷月日数に正規化済みなので算術平均が正確
        arpu_daily  = df['arpu_daily'].mean()
        arpu_short  = arpu_daily
        arpu_long   = arpu_daily
        arpu_0_dorm = arpu_daily
    gp_daily = arpu_daily * gpm

    # ビジネスタイプ依存ラベル（データ読み込みブロック内で使用）
    acq_label  = T('common_first_purchase') if business_type == BIZ_SPOT else T('common_acquisition')
    date_label = T('common_first_purchase_date') if business_type == BIZ_SPOT else T('common_start_date')

    # ── 通知メッセージ ──
    if n_dormant > 0:
        st.info(T('main_dormant_converted', n=f"{n_dormant:,}", days=str(dormancy_days)))
    if n_corrected > 0:
        st.info(T('main_same_day_corrected', n=f"{n_corrected:,}", label=date_label))
    if n_excluded > 0:
        st.warning(T('main_future_excluded', n=f"{n_excluded:,}"))
    if n_outlier > 0:
        _parts = []
        if n_outlier_upper > 0:
            _parts.append(f"upper {n_outlier_upper:,} ({outlier_upper_pct:.1f}%)" if get_lang() == 'en' else f"上位{n_outlier_upper:,}件（上位{outlier_upper_pct:.1f}%）")
        if n_outlier_lower > 0:
            _parts.append(f"lower {n_outlier_lower:,} ({outlier_lower_pct:.1f}%)" if get_lang() == 'en' else f"下位{n_outlier_lower:,}件（下位{outlier_lower_pct:.1f}%）")
        _pct = n_outlier / (n_outlier + len(df)) * 100
        _join_char = ', ' if get_lang() == 'en' else '、'
        st.info(T('main_outlier_removed', n=f"{n_outlier:,}", parts=_join_char.join(_parts), pct=_pct, remaining=f"{len(df):,}"))

    # ── 売上分布ヒストグラム ──
    import plotly.graph_objects as go
    _rev_data = _rev_before_outlier
    _hist_fig = go.Figure()
    # ヒストグラム（除外前の全データ）
    _hist_fig.add_trace(go.Histogram(
        x=_rev_data, nbinsx=80,
        marker_color='rgba(180, 180, 180, 0.45)',
        marker_line=dict(color='rgba(200, 200, 200, 0.6)', width=0.5),
        name=T('hist_trace_name'),
    ))
    # カットライン・統計線の表示
    _cut_color = '#6CB4EE'
    _cutline_shapes = []
    _cutline_annots = []
    # 平均値・中央値のグレー縦線（先に追加＝背面に配置）
    _mean_val = _rev_data.mean()
    _median_val = _rev_data.quantile(0.50)
    _cutline_shapes.append(dict(
        type='line', x0=_mean_val, x1=_mean_val, y0=0, y1=1,
        yref='paper', line=dict(color='rgba(160,160,160,0.45)', width=1.5, dash='dot'),
    ))
    _cutline_annots.append(dict(
        x=_mean_val, y=1.08, yref='paper', text=T('hist_mean'),
        showarrow=False, font=dict(color='rgba(160,160,160,0.6)', size=10), xanchor='center',
    ))
    _cutline_shapes.append(dict(
        type='line', x0=_median_val, x1=_median_val, y0=0, y1=1,
        yref='paper', line=dict(color='rgba(160,160,160,0.45)', width=1.5, dash='dot'),
    ))
    _cutline_annots.append(dict(
        x=_median_val, y=1.08, yref='paper', text=T('hist_median'),
        showarrow=False, font=dict(color='rgba(160,160,160,0.6)', size=10), xanchor='center',
    ))
    # カットライン（後に追加＝前面に配置）
    if outlier_upper_pct > 0:
        _upper_val = _rev_data.quantile(1.0 - outlier_upper_pct / 100.0)
        _cutline_shapes.append(dict(
            type='line', x0=_upper_val, x1=_upper_val, y0=0, y1=1,
            yref='paper', line=dict(color=_cut_color, width=2, dash='dash'),
        ))
        _cutline_annots.append(dict(
            x=_upper_val, y=0.55, yref='paper',
            text=f'{T("hist_upper_pct", pct=outlier_upper_pct)}<br>{fmt_c(_upper_val, CUR)}',
            showarrow=False, font=dict(color=_cut_color, size=11), xanchor='left',
        ))
    if outlier_lower_pct > 0:
        _lower_val = _rev_data.quantile(outlier_lower_pct / 100.0)
        _cutline_shapes.append(dict(
            type='line', x0=_lower_val, x1=_lower_val, y0=0, y1=1,
            yref='paper', line=dict(color=_cut_color, width=2, dash='dash'),
        ))
        _cutline_annots.append(dict(
            x=_lower_val, y=0.55, yref='paper',
            text=f'{T("hist_lower_pct", pct=outlier_lower_pct)}<br>{fmt_c(_lower_val, CUR)}',
            showarrow=False, font=dict(color=_cut_color, size=11), xanchor='left',
        ))
    _hist_fig.update_layout(
        template='plotly_dark',
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        title=dict(text=T('chart_revenue_dist'), font=dict(size=14)),
        xaxis=dict(
            title=T('chart_cumulative_rev'), tickformat=',',
            gridcolor='rgba(255,255,255,0.06)',
            zeroline=False,
            showline=True, linewidth=1, linecolor='rgba(255,255,255,0.15)',
            mirror=False,
        ),
        yaxis=dict(
            title=T('chart_customer_count'),
            gridcolor='rgba(255,255,255,0.06)',
            zeroline=False,
            layer='below traces',
        ),
        shapes=_cutline_shapes,
        annotations=_cutline_annots,
        margin=dict(l=50, r=30, t=80, b=40),
        height=300,
        showlegend=False,
    )
    # サマリー指標（HTML統一）
    _excl_pct = n_outlier / (n_outlier + len(df)) * 100 if (n_outlier + len(df)) > 0 else 0
    _summary_items = [
        (T('hist_total'), f"{len(_rev_data):,}"),
        (T('hist_excluded'), f"{n_outlier:,}"),
        (T('hist_excl_rate'), f"{_excl_pct:.1f}%"),
        (T('hist_analyzed'), f"{len(df):,}"),
    ]
    _summary_html = ''.join(
        f'<span style="margin-right:2em;"><span style="color:#888;font-size:0.72rem;">{lbl}</span>'
        f'<br><span style="color:#ccc;font-size:0.92rem;font-weight:500;">{val}</span></span>'
        for lbl, val in _summary_items
    )
    st.markdown(
        f'<div style="display:flex;flex-wrap:nowrap;gap:0;padding:8px 0 4px 0;">{_summary_html}</div>',
        unsafe_allow_html=True,
    )
    # 分布統計（小フォントで1行表示）
    _stats_items = [
        (T('hist_mean_label'), f"{fmt_c(_rev_data.mean(), CUR)}"),
        (T('hist_min'), f"{fmt_c(_rev_data.min(), CUR)}"),
        ("25%ile", f"{fmt_c(_rev_data.quantile(0.25), CUR)}"),
        (T('hist_median'), f"{fmt_c(_rev_data.quantile(0.50), CUR)}"),
        ("75%ile", f"{fmt_c(_rev_data.quantile(0.75), CUR)}"),
        (T('hist_max'), f"{fmt_c(_rev_data.max(), CUR)}"),
    ]
    _stats_html = ''.join(
        f'<span style="margin-right:1.5em;"><span style="color:#666;font-size:0.68rem;">{lbl}</span>'
        f'<br><span style="color:#999;font-size:0.78rem;">{val}</span></span>'
        for lbl, val in _stats_items
    )
    st.markdown(
        f'<div style="display:flex;flex-wrap:nowrap;gap:0;padding:2px 0 8px 0;'
        f'border-top:1px solid rgba(255,255,255,0.06);">{_stats_html}</div>',
        unsafe_allow_html=True,
    )
    st.plotly_chart(_hist_fig, use_container_width=True)
    if n_dormant == 0 and n_corrected == 0 and n_excluded == 0 and n_outlier == 0:
        st.success(T('main_data_loaded', n=f"{n_input:,}"))

    if len(df) < 10:
        st.error(T('main_data_too_few'))
        st.stop()

except Exception as e:
    st.error(T('main_data_error', e=str(e)))
    st.stop()

# ══════════════════════════════════════════════════════════════
# Run analysis
# ══════════════════════════════════════════════════════════════

# ── オフセット設定 ──
if business_type == BIZ_SPOT:
    # 都度購入型：dormancy_daysをオフセットとして使用
    # t=0〜dormancy_daysはS(t)=1.0で確定なのでWeibullフィットから切り離す
    ltv_offset_days = dormancy_days or 180
elif prorate_cancel:
    ltv_offset_days = 0
elif billing_cycle == BILLING_CALENDAR_MONTHLY:
    ltv_offset_days = 30.44
elif billing_cycle == BILLING_ANNUAL_365:
    ltv_offset_days = 365
elif billing_cycle == BILLING_CUSTOM_DAYS:
    ltv_offset_days = custom_cycle_days or 30
else:
    ltv_offset_days = 30.44

# オフセット適用：durationからオフセットを引いてフィッティング
km_df_raw = _compute_km_df(df)  # オフセット前のKM（グラフ・Excel表示用）
if ltv_offset_days > 0:
    df_fit = df.copy()
    df_fit['duration'] = df_fit['duration'] - ltv_offset_days
    # オフセット後にduration≤0になった顧客は打ち切り扱い（最低契約期間内）
    df_fit.loc[df_fit['duration'] <= 0, 'event'] = 0
    df_fit.loc[df_fit['duration'] <= 0, 'duration'] = 1
    df_fit['duration'] = df_fit['duration'].clip(lower=1)
    km_df = _compute_km_df(df_fit)
else:
    km_df = km_df_raw

k, lam, r2, fit_df = _fit_weibull_df(km_df)

if k is None:
    st.error(T('main_weibull_fail'))
    st.stop()

# LTV計算（オフセット分を加算）
def ltv_inf_offset(k, lam, arpu, offset_days):
    surv_int = lam * gamma(1 + 1/k)
    return (surv_int + offset_days) * arpu, surv_int

def ltv_horizon_offset(k, lam, arpu, h, offset_days):
    h_adj = max(h - offset_days, 0)
    if h_adj == 0:
        return offset_days * arpu
    x = (h_adj / lam) ** k
    return (lam * gamma(1 + 1/k) * gammainc(1 + 1/k, x) + offset_days) * arpu

def ltv_horizon_spot(k, lam, arpu_0d, arpu_long, h, dorm):
    """都度購入型専用LTVホライズン計算
    h: 総ホライズン（初回購入からの日数）
    dorm: dormancy_days
    """
    h_short = min(h, dorm)
    h_long  = max(h - dorm, 0)
    ltv_s   = h_short * arpu_0d
    if h_long > 0:
        x     = (h_long / lam) ** k
        ltv_l = lam * gamma(1 + 1/k) * gammainc(1 + 1/k, x) * arpu_long
    else:
        ltv_l = 0
    return ltv_s + ltv_l

if business_type == BIZ_SPOT:
    # LTV = LTV_short（固定部分）+ LTV_long（Weibull積分部分）
    _dorm_off = dormancy_days or 180
    _ltv_short_rev = _dorm_off * arpu_0_dorm   # t=0〜dormancy_daysの固定積分
    _ltv_long_rev, _surv_long = ltv_inf_offset(k, lam, arpu_long, 0)  # t=180以降
    ltv_rev  = _ltv_short_rev + _ltv_long_rev
    surv_int = _dorm_off + _surv_long
else:
    ltv_rev, surv_int = ltv_inf_offset(k, lam, arpu_daily, ltv_offset_days)  # 売上ベース
if business_type == BIZ_SPOT:
    _gp_short  = _dorm_off * (arpu_0_dorm * gpm)
    _gp_long_v, _ = ltv_inf_offset(k, lam, arpu_long * gpm, 0)
    ltv_val = _gp_short + _gp_long_v
else:
    ltv_val, _ = ltv_inf_offset(k, lam, gp_daily, ltv_offset_days)  # 粗利ベース
cac_upper = ltv_val / cac_n

# ══════════════════════════════════════════════════════════════
# Metrics
# ══════════════════════════════════════════════════════════════

st.markdown(f"<div class='section-title'>{T('main_summary_title')}</div>", unsafe_allow_html=True)
m1, m2, m3, m4, m5 = st.columns(5)

if k < 1.0:
    k_desc = T('summary_k_early')
else:
    k_desc = T('summary_k_late')
_lam_unit = T('common_days')
metrics = [
    (f"{fmt_c(ltv_rev, CUR)}", "LTV∞",       T('summary_rev_basis')),
    (f"{fmt_c(cac_upper, CUR)}", "CAC Ceiling" if get_lang() == 'en' else "CAC上限",  f"{cac_label} {T('summary_cac_gp_basis')}"),
    (f"{k:.3f}",           "Weibull k", f"{k_desc}"),
    (f"{lam + ltv_offset_days:.1f}{_lam_unit}", "Weibull λ", T('summary_k_desc_long')),
    (f"{r2:.3f}",          "R²",        T('summary_r2_note')),
]
for col, (val, title, desc) in zip([m1,m2,m3,m4,m5], metrics):
    with col:
        st.markdown(
            f"<div class='metric-card'>"
            f"<div class='metric-value'>{val}</div>"
            f"<div class='metric-label' style='font-size:0.82rem; color:#56b4d3; margin-top:6px; letter-spacing:0.05em;'>{title}</div>"
            f"<div class='metric-label' style='font-size:0.65rem; color:#444; margin-top:3px; letter-spacing:0.03em; line-height:1.4;'>{desc}</div>"
            f"</div>",
            unsafe_allow_html=True
        )

# R² warning
if r2 < 0.85:
    st.warning(T('summary_r2_warning', r2=r2))

# ── サマリー解説ボックス ──────────────────────────────────────

# 単発離脱率の計算
if business_type == BIZ_SPOT:
    # 初回購入後に1度も再購入せず離脱した顧客の割合
    churn_period = dormancy_days if dormancy_days else 365
    if 'last_purchase_date' in df.columns and df['last_purchase_date'].notna().any():
        # first=last かつ 観測完結（基準日-last >= dormancy_days）= 単発購入・離脱確定
        _gap = (df['last_purchase_date'] - df['start_date']).dt.days.fillna(-1)
        _days_since = (today - df['last_purchase_date']).dt.days.fillna(0)
        single_churn_rate = ((_gap == 0) & (_days_since >= churn_period)).sum() / len(df) * 100
    else:
        single_churn_rate = ((df['event'] == 1) & (df['duration'] <= churn_period)).sum() / len(df) * 100
    period_label = T('common_days_unit', n=churn_period)
else:
    # 最初の契約期間のみで解約した割合（実データから直接計算）
    if billing_cycle == BILLING_ANNUAL_365:
        _min_c = 365
    elif billing_cycle == BILLING_CUSTOM_DAYS and custom_cycle_days:
        _min_c = custom_cycle_days
    else:
        _min_c = 30
    churn_period = _min_c
    _dur_col = "duration_raw" if "duration_raw" in df.columns else "duration"
    single_churn_rate = ((df["event"] == 1) & (df[_dur_col] <= churn_period)).sum() / len(df) * 100
    period_label = T('common_days_unit', n=churn_period)

_lam_val = lam + ltv_offset_days
_lam_y = _lam_val / 365
_ltv_str = fmt_c(ltv_rev, CUR)
_cac_str = fmt_c(cac_upper, CUR)

if business_type == BIZ_SPOT:
    if k < 1.0:
        k_summary = T('conclusion_spot_early_churn', k=k, period=period_label, rate=single_churn_rate, lam=_lam_val, lam_y=_lam_y, ltv=_ltv_str, cac=_cac_str)
    else:
        k_summary = T('conclusion_spot_late_churn', k=k, period=period_label, rate=single_churn_rate, lam=_lam_val, lam_y=_lam_y, ltv=_ltv_str, cac=_cac_str)
else:  # サブスク
    if k < 1.0:
        k_summary = T('conclusion_sub_early_churn', k=k, rate=single_churn_rate, lam=_lam_val, lam_y=_lam_y, ltv=_ltv_str, cac=_cac_str)
    else:
        k_summary = T('conclusion_sub_late_churn', k=k, rate=single_churn_rate, lam=_lam_val, lam_y=_lam_y, ltv=_ltv_str, cac=_cac_str)

if r2 >= 0.95:
    r2_summary = T('conclusion_r2_high', r2=r2)
elif r2 >= 0.85:
    r2_summary = T('conclusion_r2_mid', r2=r2)
else:
    r2_summary = T('conclusion_r2_low', r2=r2)

summary_text = f"{k_summary}{r2_summary}"

st.markdown(f"""
<div style='margin: 16px 0 4px 0; line-height: 1.9;'>
    <div style='
        font-size: 0.78rem;
        font-weight: 600;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: #56b4d3;
        border-bottom: 1px solid #56b4d3;
        padding-bottom: 4px;
        margin-bottom: 10px;
        display: inline-block;
    '>{T('summary_conclusion')}</div>
    <div style='font-size: 0.95rem; color: #a8dadc; letter-spacing: 0.01em;'>{summary_text}</div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# Charts
# ══════════════════════════════════════════════════════════════

st.markdown(f"<div class='section-title'>{T('chart_reliability_title')}</div>", unsafe_allow_html=True)
c1, c2 = st.columns(2)

# ── グラフ描画用データ準備 ──
if business_type == BIZ_SPOT:
    # 都度購入型：
    # durationには既にdormancy_daysが含まれているので、
    # KM・Weibullともにオフセット加算不要
    # Weibullはオフセット後のkm_dfでフィットしたk/lamを使用
    # t_smooth（オフセット後の時間軸）でS(t)を計算し、そのまま描画
    km_t_plot     = np.concatenate([[0], km_df_raw['t'].values])
    km_s_plot     = np.concatenate([[1.0], km_df_raw['S'].values])
    # Weibull：t=ltv_offset_daysから開始（それ以前は線なし）
    t_smooth_plot = np.linspace(ltv_offset_days, km_df_raw['t'].max() * 1.3, 600)
    S_wei_plot    = weibull_s(t_smooth_plot - ltv_offset_days, k, lam)
else:
    # サブスク：km_df（オフセット後）のt軸を+ltv_offset_daysで元スケールに戻す
    # これによりt=ltv_offset_days未満のデータが除外され、正しくt=30から落ちる
    t_smooth      = np.linspace(1, km_df['t'].max() * 1.3, 600)
    S_wei         = weibull_s(t_smooth, k, lam)
    km_t_plot     = np.concatenate([[0, ltv_offset_days], km_df['t'].values + ltv_offset_days])
    km_s_plot     = np.concatenate([[1.0, 1.0], km_df['S'].values])
    t_smooth_plot = t_smooth + ltv_offset_days
    S_wei_plot    = S_wei

with c1:
    fig, ax = plt.subplots(figsize=(6, 3.8))
    ax.step(km_t_plot, km_s_plot, where='post', color=ACCENT, lw=1.8, label='KM Curve (Observed)')
    ax.plot(t_smooth_plot, S_wei_plot, color=ACCENT2, lw=1.5, ls='--', label='Weibull Fit')
    ax.fill_between(t_smooth_plot, S_wei_plot, alpha=0.06, color=ACCENT2)
    ax.set(xlabel='Days', ylabel='Survival Rate S(t)', ylim=(0,1.05),
           xlim=(0, t_smooth_plot.max()))
    ax.legend(fontsize=8, framealpha=0.15)
    ax.grid(True, alpha=0.25)
    ax.set_title('Survival Curve', color='#ccc', fontsize=10, pad=8)
    fig.tight_layout()
    st.pyplot(fig)
    plt.close()

with c1:
    st.caption(T('chart_survival_caption'))

with c2:
    fig, ax = plt.subplots(figsize=(6, 3.8))
    if fit_df is not None:
        ax.scatter(fit_df['ln_t'], fit_df['ln_neg_ln_S'], color=ACCENT, s=18, alpha=0.7, label='Observed')
        x_r = np.linspace(fit_df['ln_t'].min(), fit_df['ln_t'].max(), 200)
        b   = np.log(-np.log(weibull_s(1, k, lam) + 1e-15))
        ax.plot(x_r, k*x_r + b, color=ACCENT2, lw=1.5, ls='--', label=f'Regression Line (R²={r2:.3f})')
        ax.annotate(f'y = {k:.4f}x + {b:.4f}',
                    xy=(0.05,0.93), xycoords='axes fraction', color='#777', fontsize=8)
    ax.set(xlabel='ln(t)', ylabel='ln(−ln(S(t)))')
    ax.legend(fontsize=8, framealpha=0.15)
    ax.grid(True, alpha=0.25)
    ax.set_title('Weibull Linearization Plot', color='#ccc', fontsize=10, pad=8)
    fig.tight_layout()
    st.pyplot(fig)
    plt.close()

with c2:
    st.caption(T('chart_linearization_caption'))

# Save chart images for export
fig1, ax1 = plt.subplots(figsize=(7, 4))
ax1.step(km_t_plot, km_s_plot, where='post', color=ACCENT, lw=2, label='KM Curve')
ax1.plot(t_smooth_plot, S_wei_plot, color=ACCENT2, lw=1.8, ls='--', label='Weibull Fit')
ax1.fill_between(t_smooth_plot, S_wei_plot, alpha=0.07, color=ACCENT2)
ax1.set(xlabel='Days', ylabel='Survival Rate S(t)', ylim=(0,1.05))
ax1.legend(fontsize=9, framealpha=0.15)
ax1.grid(True, alpha=0.25)
ax1.set_title('Survival Curve (KM × Weibull)', color='#ccc', fontsize=11, pad=10)
fig1.tight_layout()
buf1 = io.BytesIO(); fig1.savefig(buf1, format='png', dpi=150, bbox_inches='tight'); buf1.seek(0)
plt.close()

fig2, ax2 = plt.subplots(figsize=(7, 4))
if fit_df is not None:
    ax2.scatter(fit_df['ln_t'], fit_df['ln_neg_ln_S'], color=ACCENT, s=22, alpha=0.75)
    x_r = np.linspace(fit_df['ln_t'].min(), fit_df['ln_t'].max(), 200)
    b   = np.log(-np.log(weibull_s(1, k, lam) + 1e-15))
    ax2.plot(x_r, k*x_r+b, color=ACCENT2, lw=1.8, ls='--', label=f'R²={r2:.3f}')
    ax2.annotate(f'y = {k:.4f}x + {b:.4f}', xy=(0.05,0.93), xycoords='axes fraction', color='#777', fontsize=9)
ax2.set(xlabel='ln(t)', ylabel='ln(−ln(S(t)))')
ax2.legend(fontsize=9, framealpha=0.15)
ax2.grid(True, alpha=0.25)
ax2.set_title('Weibull Linearization Plot', color='#ccc', fontsize=11, pad=10)
fig2.tight_layout()
buf2 = io.BytesIO(); fig2.savefig(buf2, format='png', dpi=150, bbox_inches='tight'); buf2.seek(0)
plt.close()

# ══════════════════════════════════════════════════════════════
# Horizon table
# ══════════════════════════════════════════════════════════════

st.markdown(f"<div class='section-title'>{T('chart_interim_ltv_title')}</div>", unsafe_allow_html=True)

horizons = [180, 365, 730, 1095, 1825]  # 180日・1年・2年・3年・5年

# ── 折れ線グラフ（180日〜5年、λを縦線で表示）──────────────
# グラフ用の細かい点を生成（滑らかな曲線）
# λをグラフ最大値として扱う（5年超の場合はλが最大）
lam_actual = lam + ltv_offset_days  # 実際のλ位置（オフセット込み）
x_max = max(1825, round(lam_actual) + 100) if lam_actual > 1825 else 1825

t_range = list(range(1, x_max + 1, max(1, x_max // 300)))
if business_type == BIZ_SPOT:
    _dorm_off = dormancy_days or 180
    rev_line = [ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, t, _dorm_off) for t in t_range]
    gp_line  = [ltv_horizon_spot(k, lam, arpu_0_dorm*gpm, arpu_long*gpm, t, _dorm_off) for t in t_range]
else:
    rev_line = [ltv_horizon_offset(k, lam, arpu_daily, t, ltv_offset_days) for t in t_range]
    gp_line  = [ltv_horizon_offset(k, lam, gp_daily,   t, ltv_offset_days) for t in t_range]
cac_line = [v / cac_n for v in gp_line]

lam_int = round(lam_actual)

_col_ltv, = st.columns([1])
with _col_ltv:
    fig_ltv = go.Figure()
    # 線グラフ（マーカーなし）
    fig_ltv.add_trace(go.Scatter(x=t_range, y=rev_line, name=T('chart_ltv_rev'), mode='lines', line=dict(color='#56b4d3', width=2), showlegend=True))
    fig_ltv.add_trace(go.Scatter(x=t_range, y=gp_line,  name=T('chart_ltv_gp'), mode='lines', line=dict(color='#a8dadc', width=2, dash='dash'), showlegend=True))
    fig_ltv.add_trace(go.Scatter(x=t_range, y=cac_line, name=T('chart_cac_cap'),    mode='lines', line=dict(color='#4a7a8a', width=1.5, dash='dot'), showlegend=True))

    # 特定ポイントにマーカー（180日,1年,2年,3年,4年,5年,λ）
    _marker_days = sorted(set([180, 365, 730, 1095, 1460, 1825, round(lam_actual)]))
    for _trace_data, _color, _name in [(rev_line, '#56b4d3', T('chart_ltv_rev')), (gp_line, '#a8dadc', T('chart_ltv_gp')), (cac_line, '#4a7a8a', T('chart_cac_cap'))]:
        _mx = [d for d in _marker_days if d <= max(t_range)]
        _my = []
        for d in _mx:
            # t_rangeから最も近いインデックスを探す
            _closest = min(range(len(t_range)), key=lambda i: abs(t_range[i] - d))
            _my.append(_trace_data[_closest])
        fig_ltv.add_trace(go.Scatter(x=_mx, y=_my, mode='markers', marker=dict(color=_color, size=6, symbol='circle'), showlegend=False, hoverinfo='skip'))

    fig_ltv.add_hline(y=ltv_rev, line_dash='dot', line_color='#56b4d3', line_width=1, opacity=0.4,
        annotation_text=f'LTV∞ {fmt_c(ltv_rev, CUR)}', annotation_position='right',
        annotation_font=dict(color='#56b4d3', size=10))
    fig_ltv.add_shape(type='line', x0=lam_actual, x1=lam_actual, y0=0, y1=ltv_rev,
        line=dict(color='#a8dadc', width=1.5, dash='dash'), layer='above')
    fig_ltv.add_annotation(x=lam_actual, y=0.85 if k < 1.0 else 0.35, yref='paper',
        text=T('chart_lam_days', n=lam_int), showarrow=False,
        font=dict(color='#a8dadc', size=10), xanchor='center', yanchor='middle',
        bgcolor='#111820', borderpad=2)
    tick_vals = [180, 365, 730, 1095, 1460, 1825]
    _d = T('chart_days_suffix')
    _y = T('chart_year_suffix')
    tick_text = [f'180{_d}', f'1{_y}', f'2{_y}', f'3{_y}', f'4{_y}', f'5{_y}']
    fig_ltv.update_layout(
        paper_bgcolor='#111820', plot_bgcolor='#111820',
        height=300, margin=dict(t=40, b=50, l=70, r=120),
        font=dict(color='#ccc', size=10),
        legend=dict(orientation='h', y=1.08, x=0, font=dict(size=10), bgcolor='rgba(0,0,0,0)'),
        xaxis=dict(title=T('chart_duration'), gridcolor='#1a3040', tickvals=tick_vals, ticktext=tick_text, tickfont=dict(color='#888'), range=[0, x_max + 50]),
        yaxis=dict(title=T('chart_amount'), gridcolor='#1a3040', tickfont=dict(color='#888'), tickformat=',', tickprefix=''),
    )
    st.plotly_chart(fig_ltv, use_container_width=True)

# ── 日本語フォント動的探索ヘルパー ──
def _find_jp_font_path():
    """環境に依存せず日本語フォントパスを返す。見つからなければNone。"""
    import os
    _candidates = [
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Medium.ttc',
        '/usr/share/fonts/opentype/ipafont-gothic/ipagp.ttf',
        '/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf',
        '/usr/share/fonts/truetype/fonts-japanese-gothic.ttf',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc',
    ]
    for _p in _candidates:
        if os.path.exists(_p):
            return _p
    try:
        import matplotlib.font_manager as _fmx
        _all = _fmx.findSystemFonts()
        for _kw in ['CJK', 'ipagp', 'ipag', 'Japanese', 'Gothic']:
            for _fp in _all:
                if _kw.lower() in _fp.lower():
                    return _fp
    except Exception:
        pass
    return None

_JP_FONT_PATH = _find_jp_font_path()

# matplotlibにフォントを明示登録（Streamlit Cloud対応）
try:
    import matplotlib as _mpl_init
    import matplotlib.font_manager as _fm_init
    _IPA_TTF = '/usr/share/fonts/truetype/fonts-japanese-gothic.ttf'
    import os as _os_init
    if _os_init.path.exists(_IPA_TTF):
        _fm_init.fontManager.addfont(_IPA_TTF)
        _JP_FONT_PATH = _IPA_TTF
    elif _JP_FONT_PATH:
        _fm_init.fontManager.addfont(_JP_FONT_PATH)
    _JP_FONT_NAME = _fm_init.FontProperties(fname=_JP_FONT_PATH).get_name() if _JP_FONT_PATH else None
except Exception:
    _JP_FONT_NAME = None



# λ時点と99%到達日数を逆算
try:
    if business_type == BIZ_SPOT:
        _dorm_off = dormancy_days or 180
        days_99 = brentq(
            lambda h: ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, h, _dorm_off) / ltv_rev - 0.99,
            1, 365000
        )
    else:
        days_99 = brentq(
            lambda h: ltv_horizon_offset(k, lam, arpu_daily, h, ltv_offset_days) / ltv_rev - 0.99,
            1, 365000  # 上限1000年に拡大
        )
except Exception:
    # 上限でも99%に届かない場合は上限値で近似
    days_99 = 365000

def fmt_horizon(days):
    _d = T('chart_days_suffix')
    if days < 365:
        return f'{int(days)}{_d}'
    elif days % 365 == 0:
        return T('chart_year_days', y=int(days//365), d=f'{int(days):,}')
    else:
        return T('chart_year_days', y=f'{days/365:.1f}', d=f'{int(days):,}')

# テーブルデータ構築
tbl_rows = []
for h in horizons:
    if business_type == BIZ_SPOT:
        _dorm_off = dormancy_days or 180
        lh_rev = ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, h, _dorm_off)
        lh_gp  = ltv_horizon_spot(k, lam, arpu_0_dorm*gpm, arpu_long*gpm, h, _dorm_off)
    else:
        lh_rev = ltv_horizon_offset(k, lam, arpu_daily, h, ltv_offset_days)
        lh_gp  = ltv_horizon_offset(k, lam, gp_daily,   h, ltv_offset_days)
    tbl_rows.append({
        'horizon':    fmt_horizon(h),
        'ltv_rev':    f'{fmt_c(lh_rev, CUR)}',
        'ltv_gp':     f'{fmt_c(lh_gp, CUR)}',
        'cac_cap':    f'{fmt_c(lh_gp/cac_n, CUR)}',
        'pct':        f'{lh_rev/ltv_rev*100:.1f}%',
        '_type': 'normal',
    })

# λ行
if business_type == BIZ_SPOT:
    _dorm_off = dormancy_days or 180
    lam_rev = ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, lam + _dorm_off, _dorm_off)
    lam_gp  = ltv_horizon_spot(k, lam, arpu_0_dorm*gpm, arpu_long*gpm, lam + _dorm_off, _dorm_off)
else:
    lam_rev  = ltv_horizon_offset(k, lam, arpu_daily, lam + ltv_offset_days, ltv_offset_days)
    lam_gp   = ltv_horizon_offset(k, lam, gp_daily,   lam + ltv_offset_days, ltv_offset_days)
lam_pct  = lam_rev / ltv_rev * 100
tbl_rows.append({
    'horizon':    T('tbl_lam_row', n=f'{round(lam + ltv_offset_days):,}'),
    'ltv_rev':    f'{fmt_c(lam_rev, CUR)}',
    'ltv_gp':     f'{fmt_c(lam_gp, CUR)}',
    'cac_cap':    f'{fmt_c(lam_gp/cac_n, CUR)}',
    'pct':        f'{lam_pct:.1f}%',
    '_type': 'lambda',
})

# 99%到達行
if business_type == BIZ_SPOT:
    _dorm_off = dormancy_days or 180
    rev_99 = ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, days_99, _dorm_off)
    gp_99  = ltv_horizon_spot(k, lam, arpu_0_dorm*gpm, arpu_long*gpm, days_99, _dorm_off)
else:
    rev_99 = ltv_horizon_offset(k, lam, arpu_daily, days_99, ltv_offset_days)
    gp_99  = ltv_horizon_offset(k, lam, gp_daily,   days_99, ltv_offset_days)
tbl_rows.append({
    'horizon':    T('tbl_99pct_row', n=f'{int(days_99):,}'),
    'ltv_rev':    f'{fmt_c(rev_99, CUR)}',
    'ltv_gp':     f'{fmt_c(gp_99, CUR)}',
    'cac_cap':    f'{fmt_c(gp_99/cac_n, CUR)}',
    'pct':        '99.0%',
    '_type': '99pct',
})

# LTV∞行
tbl_rows.append({
    'horizon':    'LTV∞',
    'ltv_rev':    f'{fmt_c(ltv_rev, CUR)}',
    'ltv_gp':     f'{fmt_c(ltv_val, CUR)}',
    'cac_cap':    f'{fmt_c(ltv_val/cac_n, CUR)}',
    'pct':        '100%',
    '_type': 'ltv_inf',
})

# HTML テーブルで描画
ACCENT   = '#56b4d3'
BG_HEAD  = '#0d1f2d'
BG_ROW1  = '#0d1520'
BG_ROW2  = '#0d1520'
SEP_COLOR = '#1a3a4a'  # λ・99%行の区切り線色

html_rows = ''
for i, row in enumerate(tbl_rows):
    rtype = row['_type']
    if rtype == 'normal':
        bg    = BG_ROW1 if i % 2 == 0 else BG_ROW2
        color = '#c8d0d8'
        border_top = ''
    elif rtype == 'lambda':
        bg    = BG_HEAD
        color = '#a8dadc'
        border_top = f"border-top:1px solid {SEP_COLOR};"
    else:  # 99pct or ltv_inf
        bg    = BG_HEAD
        color = '#a8dadc'
        border_top = f"border-top:1px solid {SEP_COLOR};"

    html_rows += f"<tr style='background:{bg}; {border_top}'>"
    html_rows += f"<td style='text-align:left; padding:8px 14px; color:{color}; font-size:0.85rem; width:28%;'>{row['horizon']}</td>"
    for col in ['ltv_rev', 'ltv_gp', 'cac_cap', 'pct']:
        html_rows += f"<td style='text-align:right; padding:8px 14px; color:{color}; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{row[col]}</td>"
    html_rows += '</tr>'

tbl_html = f"""
<table style='width:100%; border-collapse:collapse; margin-top:4px; table-layout:fixed;'>
  <colgroup>
    <col style='width:28%;'>
    <col style='width:18%;'>
    <col style='width:18%;'>
    <col style='width:18%;'>
    <col style='width:18%;'>
  </colgroup>
  <thead>
    <tr style='background:{BG_HEAD};'>
      <th style='text-align:left; padding:9px 14px; color:{ACCENT}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT};'>{T('tbl_horizon')}</th>
      <th style='text-align:right; padding:9px 14px; color:{ACCENT}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT};'>{T('tbl_ltv_rev')}</th>
      <th style='text-align:right; padding:9px 14px; color:{ACCENT}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT};'>{T('tbl_ltv_gp')}</th>
      <th style='text-align:right; padding:9px 14px; color:{ACCENT}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT};'>{T('tbl_cac_cap')}</th>
      <th style='text-align:right; padding:9px 14px; color:{ACCENT}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT};'>{T('tbl_pct_ltv')}</th>
    </tr>
  </thead>
  <tbody>
    {html_rows}
  </tbody>
</table>
"""
st.markdown(tbl_html, unsafe_allow_html=True)


# 解釈ガイドを自動生成
if business_type == BIZ_SPOT:
    _dorm_off = dormancy_days or 180
    ltv_1y = ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, 365,  _dorm_off)
    ltv_2y = ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, 730,  _dorm_off)
    ltv_3y = ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, 1095, _dorm_off)
else:
    ltv_1y  = ltv_horizon_offset(k, lam, arpu_daily, 365,  ltv_offset_days)
    ltv_2y  = ltv_horizon_offset(k, lam, arpu_daily, 730,  ltv_offset_days)
    ltv_3y  = ltv_horizon_offset(k, lam, arpu_daily, 1095, ltv_offset_days)
pct_1y  = ltv_1y  / ltv_rev * 100
pct_2y  = ltv_2y  / ltv_rev * 100
pct_3y  = ltv_3y  / ltv_rev * 100

# CAC回収期間を逆算（売上ベース・粗利ベース両方）
def recover_str(days):
    _d = T('chart_days_suffix')
    if days >= 365:
        return T('chart_year_days', y=f'{days/365:.1f}', d=f'{int(days):,}')
    return f"{int(days)}{_d}"

try:
    cac_recover_rev = brentq(
        lambda h: (ltv_horizon_spot(k, lam, arpu_0_dorm, arpu_long, h, dormancy_days or 180) if business_type == BIZ_SPOT else ltv_horizon_offset(k, lam, arpu_daily, h, ltv_offset_days)) - cac_upper,
        1, 36500
    )
    cac_recover_rev_str = recover_str(cac_recover_rev)
except Exception:
    cac_recover_rev_str = "N/A"

try:
    cac_recover_gp = brentq(
        lambda h: (ltv_horizon_spot(k, lam, arpu_0_dorm*gpm, arpu_long*gpm, h, dormancy_days or 180) if business_type == BIZ_SPOT else ltv_horizon_offset(k, lam, gp_daily, h, ltv_offset_days)) - cac_upper,
        1, 36500
    )
    cac_recover_gp_str = recover_str(cac_recover_gp)
except Exception:
    cac_recover_gp_str = "N/A"

# 後方互換用
cac_recover_days = cac_recover_rev if cac_recover_rev_str != "N/A" else None
cac_recover_str  = cac_recover_rev_str

# λの解釈
lam_display = lam + ltv_offset_days if business_type == BIZ_SPOT else lam
_lam_int = round(lam_display)
_lam_yr = lam_display / 365
if lam_display < 365:
    lam_desc = T('insight_lam_short', lam=_lam_int, y=_lam_yr)
elif lam_display < 730:
    lam_desc = T('insight_lam_mid', lam=_lam_int, y=_lam_yr, yl=1, yh=2)
else:
    lam_desc = T('insight_lam_long', lam=_lam_int, y=_lam_yr)

# k の解釈
if k < 0.8:
    k_desc = T('insight_k_early_strong', k=k, acq=acq_label)
elif k < 1.0:
    k_desc = T('insight_k_early_mild', k=k, acq=acq_label)
elif k < 1.5:
    k_desc = T('insight_k_late_mild', k=k, acq=acq_label)
else:
    k_desc = T('insight_k_late_strong', k=k, acq=acq_label)

# R² の解釈（insight用）
if r2 >= 0.95:
    _r2_insight = T('insight_r2_high', r2=r2)
elif r2 >= 0.85:
    _r2_insight = T('insight_r2_mid', r2=r2)
else:
    _r2_insight = T('insight_r2_low', r2=r2)

_insight_ltv_inf = T('insight_ltv_inf', ltv=fmt_c(ltv_rev, CUR))
_insight_1y = T('insight_pct_years', label=T('insight_1y'), pct=pct_1y, val=fmt_c(ltv_1y, CUR))
_insight_2y = T('insight_pct_years', label=T('insight_2y'), pct=pct_2y, val=fmt_c(ltv_2y, CUR))
_insight_3y = T('insight_pct_years', label=T('insight_3y'), pct=pct_3y, val=fmt_c(ltv_3y, CUR))
_insight_cac = T('insight_cac_recover', cac=fmt_c(cac_upper, CUR), rev_str=cac_recover_rev_str, gp_str=cac_recover_gp_str)
_insight_design = T('insight_cac_design', lam=f'{round(lam_actual):,}', y=lam_actual/365, gp=fmt_c(lam_gp, CUR))

insight_html = f"""
<div style='background:#0d1f2d; border:1px solid #1a3a4a; border-radius:10px; padding:18px 20px; margin-top:40px; line-height:1.9; font-size:0.85rem; color:#ccc;'>
  <div style='font-size:0.65rem; font-weight:600; text-transform:uppercase; letter-spacing:0.12em; color:#3a6a7a; margin-bottom:12px;'>{T('insight_title')}</div>
  <div>・{lam_desc}</div>
  <div>・{k_desc}</div>
  <div>・{_insight_ltv_inf}</div>
  <div style='margin-top:8px;'>
    {_insight_1y}<br>{_insight_2y}<br>{_insight_3y}
  </div>
  <div style='margin-top:8px;'>・{_insight_cac}</div>
  <div style='margin-top:12px; padding-top:10px; border-top:1px solid #1a3a4a;'>
    <span style='color:#56b4d3; font-weight:600;'>{_insight_design}</span>
  </div>
</div>
"""
st.markdown(insight_html, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# AI Prompt Generator
# ══════════════════════════════════════════════════════════════

_UPGRADE_TITLE = 'Advanced 機能' if get_lang() == 'ja' else 'Advanced Feature'
_UPGRADE_DESC = 'セグメント別LTV∞分析を利用するにはAdvancedへのアップグレードが必要です。' if get_lang() == 'ja' else 'Upgrade to Advanced to unlock Segment-level LTV∞ Analysis.'
_UPGRADE_BTN = 'Advanced にアップグレード' if get_lang() == 'ja' else 'Upgrade to Advanced'
_UPGRADE_HTML = f"""
<div style='background: linear-gradient(135deg, #0d1f2d 0%, #142838 100%); border: 1px solid #1a4a5a;
    border-radius: 12px; padding: 28px 32px; margin: 16px 0 24px 0; text-align: center;'>
  <div style='font-size: 1.1rem; color: #56b4d3; font-weight: 600; margin-bottom: 8px;'>{_UPGRADE_TITLE}</div>
  <div style='font-size: 0.85rem; color: #8aa; line-height: 1.6; margin-bottom: 16px;'>
    {_UPGRADE_DESC}
  </div>
  <a href='https://ltv-analyzer.com' target='_blank'
     style='display: inline-block; background: #56b4d3; color: #0a1020; padding: 10px 32px;
     border-radius: 6px; font-weight: 600; font-size: 0.85rem; text-decoration: none;'>{_UPGRADE_BTN}</a>
</div>
"""

st.markdown(f"<div class='section-title'>{T('section_ai_prompt')}</div>", unsafe_allow_html=True)

st.markdown(f"<div class='help-box'>{T('prompt_help_box')}</div>", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs([T('prompt_tab1'), T('prompt_tab2'), T('prompt_tab3')])

dormancy_label = T('common_none') + " (end_date)" if dormancy_days is None else T('common_days_unit', n=dormancy_days)
churned_count  = int(df['event'].sum())
active_count   = int((df['event']==0).sum())
churn_rate     = churned_count / len(df) * 100
# ビジネスタイプで「契約」「初回購入」を切り替え
acq_label  = T('common_first_purchase') if business_type == BIZ_SPOT else T('common_acquisition')
date_label = T('common_first_purchase_date') if business_type == BIZ_SPOT else T('common_start_date')

_biz_disp = T('biz_spot') if business_type == BIZ_SPOT else T('biz_subscription')
_lam_prompt = lam+ltv_offset_days if business_type==BIZ_SPOT else lam

if get_lang() == 'en':
    k_pattern = f"early-churn (high churn right after {acq_label})" if k < 1 else f"increasing-churn (churn grows over time)"
    prompt_base = f"""I used an LTV analysis tool (Kaplan-Meier × Weibull model) and obtained the following results.

[Results]
- Business type: {_biz_disp} / Dormancy: {dormancy_label}
- Customers: {len(df):,} (churned: {churned_count:,} / active: {active_count:,} / churn rate: {churn_rate:.1f}%)
- Daily ARPU (revenue): {fmt_c(arpu_daily, CUR, 2)} / Daily GP: {fmt_c(gp_daily, CUR, 2)} / GPM: {gpm:.1%}
- LTV∞ (revenue): {fmt_c(ltv_rev, CUR)} / LTV∞ (gross profit): {fmt_c(ltv_val, CUR)}
- CAC cap ({cac_label}): {fmt_c(cac_upper, CUR)}
- Weibull k (shape): {k:.4f} → {k_pattern}
- Weibull λ (scale): {_lam_prompt:.1f}d / R² (fit): {r2:.4f}
- Method: Kaplan-Meier + Weibull survival analysis"""

    with tab1:
        p1 = prompt_base + f"""

[Questions — use the numbers above in your answers]
1. What does k={k:.4f} tell us about this business's churn pattern and payback feasibility? Interpret from the perspective that k<1 means "long payback, heavy tail from few long-term customers" and k>1 means "relatively short payback."
2. It takes {int(days_99):,} days ({days_99/365:.1f} years) to reach 99% of LTV∞. How should this affect CAC design? What's a realistic payback target?
3. With a churn rate of {churn_rate:.1f}% ({churned_count:,} churned / {active_count:,} active), how would you assess this business's health?
4. Is R²={r2:.4f} acceptable for Weibull analysis? What are the reliability limitations at this level?"""
        st.code(p1, language=None)

    with tab2:
        p2 = prompt_base + f"""

[Questions — derive actionable decisions from the numbers above]
1. If using interim LTV (GP) at λ={lam_display:.0f}d (~{lam_display/365:.1f}yr) = {fmt_c(lam_gp, CUR)} as CAC cap basis with LTV:CAC={cac_n:.1f}:1, what should the CAC target be? Is this appropriate?
2. What's the risk of using LTV∞ directly for CAC decisions with k={k:.4f}? Compare CAC caps at 1yr ({fmt_c(ltv_1y*gpm/cac_n, CUR)}), 2yr ({fmt_c(ltv_2y*gpm/cac_n, CUR)}), λ={round(lam_display)}d ({fmt_c(lam_gp/cac_n, CUR)}) — which is most practical?
3. Given λ={lam_display:.0f}d, when should retention campaigns be deployed after {acq_label}?
4. What retention strategies and timing are most effective for a {k_pattern} pattern?"""
        st.code(p2, language=None)

    with tab3:
        _q4_en = "4. Analysis uses end_date for churn. How would missing or delayed churn data affect LTV estimates?" if dormancy_days is None else f"4. Is the dormancy threshold of {dormancy_label} appropriate? Describe a sensitivity analysis to determine the optimal threshold."
        p3 = prompt_base + f"""

[Questions — assess model reliability, limitations, and improvements]
1. With {len(df):,} records and {churned_count:,} churn events, what confidence interval can we expect for Weibull estimates? What sample size is recommended?
2. Does k={k:.4f} satisfy Weibull's monotone hazard assumption? What are typical violations and how to check?
3. How can R²={r2:.4f} be improved? Discuss data volume, observation period, and outlier handling.
{_q4_en}
"""
        st.code(p3, language=None)

else:
    k_pattern = f"初期集中型（{acq_label}直後の離脱が多い）" if k < 1 else "逓増型（時間とともに離脱が増える）"
    prompt_base = f"""私はLTV分析ツール（Kaplan-Meier法 × Weibullモデル）を使い、以下の結果を得ました。

【分析結果】
・ビジネスタイプ: {_biz_disp} / 休眠判定: {dormancy_label}
・顧客数: {len(df):,}件（解約済み: {churned_count:,}件 / 継続中: {active_count:,}件 / 解約率: {churn_rate:.1f}%）
・Daily ARPU（売上）: {fmt_c(arpu_daily, CUR, 2)} / Daily GP（粗利）: {fmt_c(gp_daily, CUR, 2)} / GPM: {gpm:.1%}
・LTV∞（売上ベース）: {fmt_c(ltv_rev, CUR)} / LTV∞（粗利ベース）: {fmt_c(ltv_val, CUR)}
・CAC上限（{cac_label}）: {fmt_c(cac_upper, CUR)}
・Weibull k（形状）: {k:.4f} → {k_pattern}
・Weibull λ（尺度）: {_lam_prompt:.1f}日 / R²（フィット精度）: {r2:.4f}
・分析手法: Kaplan-Meier法 + Weibullモデルによる生存分析"""

    with tab1:
        p1 = prompt_base + f"""

【質問】―― 以下の数値を具体的に使って答えてください ――
1. k={k:.4f}という値はこのビジネスの離脱パターンと投資回収のしやすさについて何を示していますか？k<1は「LTV∞の回収に長期間かかる・少数の超長期顧客の影響が大きい」、k>1は「比較的短期にLTV∞を回収できる」という観点で解釈してください。
2. 99%到達まで{int(days_99):,}日（{days_99/365:.1f}年）かかるという事実は、このビジネスのCAC設計にどう影響しますか？現実的な回収期間として何年を目安にすべきですか？
3. 解約率{churn_rate:.1f}%（解約{churned_count:,}件・継続{active_count:,}件）という比率から、このビジネスの健全性をどう評価しますか？
4. R²={r2:.4f}のフィット精度はWeibull分析として許容範囲ですか？この値が示す信頼性の限界を教えてください。"""
        st.code(p1, language=None)

    with tab2:
        p2 = prompt_base + f"""

【質問】―― 上記の数値から直接導ける意思決定を具体的に答えてください ――
1. このビジネスのλ={lam_display:.0f}日（約{lam_display/365:.1f}年）時点の暫定LTV（粗利）{fmt_c(lam_gp, CUR)}をCAC上限の基準とした場合、LTV:CAC={cac_n:.1f}:1の設定でCACの目安はいくらですか？またこの設定は適切ですか？
2. k={k:.4f}のビジネスにおいて、LTV∞をそのままCAC判断に使うリスクを説明してください。λ日基準・1年基準・2年基準それぞれのCAC上限（1年:{fmt_c(ltv_1y*gpm/cac_n, CUR)}、2年:{fmt_c(ltv_2y*gpm/cac_n, CUR)}、λ={round(lam_display)}日:{fmt_c(lam_gp/cac_n, CUR)})を比較して、最も実務的な基準はどれですか？
3. λ={lam_display:.0f}日を踏まえると、{acq_label}後何日目にリテンション施策を打つのが最も効果的ですか？
4. {k_pattern}に対して最も効果的なリテンション施策のタイミングと種類を教えてください。"""
        st.code(p2, language=None)

    with tab3:
        p3 = prompt_base + f"""

【質問】―― モデルの信頼性・限界・改善策を具体的に答えてください ――
1. {len(df):,}件・解約{churned_count:,}件のサンプルサイズでWeibull推定の信頼区間はどの程度ですか？十分なサンプル数の目安を教えてください。
2. k={k:.4f}はWeibull分布の単調ハザード率の仮定を満たしていますか？仮定が崩れる典型例とチェック方法を教えてください。
3. R²={r2:.4f}を改善するにはどうすればよいですか？データ量・期間・外れ値処理の観点から具体的に教えてください。
{"4. 解約日ベースで分析していますが、解約データの欠損や遅延がある場合にLTV推定にどんな影響が出ますか？" if dormancy_days is None else f"4. 休眠判定{dormancy_label}の設定はこのビジネスに適切ですか？最適な判定日数を決める感度分析の手順を教えてください。"}
"""
        st.code(p3, language=None)

# ══════════════════════════════════════════════════════════════
# Pre-compute Segment Analysis (shared by all outputs)
# ══════════════════════════════════════════════════════════════
all_seg_results = {}  # {seg_col: seg_df}  全出力で共有
all_seg_details = {}  # {seg_col: [dict, ...]}  ソート済みseg_results

if segment_cols_input.strip():
    _pre_seg_cols = [c.strip() for c in segment_cols_input.split(',') if c.strip() and c.strip() in df.columns]
    for _pre_sc in _pre_seg_cols[:5]:
        _pre_seg_values = df[_pre_sc].dropna().unique()
        if len(_pre_seg_values) > 50:
            continue
        _pre_results = []
        for _pre_sv in sorted(_pre_seg_values):
            _pre_df = df[df[_pre_sc] == _pre_sv].copy()
            if len(_pre_df) < 10 or _pre_df['event'].sum() < 5:
                continue
            try:
                if ltv_offset_days > 0:
                    _pre_df_fit = _pre_df.copy()
                    _pre_df_fit['duration'] = _pre_df_fit['duration'] - ltv_offset_days
                    _pre_df_fit.loc[_pre_df_fit['duration'] <= 0, 'event'] = 0
                    _pre_df_fit.loc[_pre_df_fit['duration'] <= 0, 'duration'] = 1
                    _pre_df_fit['duration'] = _pre_df_fit['duration'].clip(lower=1)
                    _pre_km = _compute_km_df(_pre_df_fit)
                else:
                    _pre_km = _compute_km_df(_pre_df)
                _pre_k, _pre_lam, _pre_r2, _ = _fit_weibull_df(_pre_km)
                if _pre_k is None:
                    continue
                if business_type == BIZ_SPOT and 'last_purchase_date' in _pre_df.columns:
                    _pre_dorm = dormancy_days or 180
                    _gap = (_pre_df['last_purchase_date'] - _pre_df['start_date']).dt.days.fillna(-1)
                    _days = (today - _pre_df['last_purchase_date']).dt.days.fillna(0)
                    _single_m = (_gap == 0) & (_days >= _pre_dorm)
                    _long_m   = (_gap > 0)
                    _single_d = _pre_df[_single_m]
                    _long_d   = _pre_df[_long_m]
                    _arpu_short = _single_d['revenue_total'].sum() / len(_single_d) / _pre_dorm if len(_single_d) > 0 else 0.0
                    _arpu_long  = _long_d['revenue_total'].sum() / _long_d['duration'].sum() if len(_long_d) > 0 else _pre_df['revenue_total'].sum() / _pre_df['duration'].sum()
                    _n = len(_single_d) + len(_long_d)
                    _w_s = len(_single_d) / _n if _n > 0 else 0.5
                    _w_l = len(_long_d) / _n if _n > 0 else 0.5
                    _arpu_0_dorm = _arpu_short * _w_s + _arpu_long * _w_l
                    _arpu = _arpu_long
                else:
                    _arpu = _pre_df['arpu_daily'].mean()
                    _arpu_long = _arpu
                    _arpu_0_dorm = _arpu
                    _pre_dorm = ltv_offset_days
                _gp = _arpu * gpm
                if business_type == BIZ_SPOT:
                    _ltv_short = _pre_dorm * _arpu_0_dorm
                    _ltv_long, _ = ltv_inf_offset(_pre_k, _pre_lam, _arpu_long, 0)
                    _ltv_r = _ltv_short + _ltv_long
                    _gp_short = _pre_dorm * (_arpu_0_dorm * gpm)
                    _gp_long, _ = ltv_inf_offset(_pre_k, _pre_lam, _arpu_long * gpm, 0)
                    _ltv_g = _gp_short + _gp_long
                else:
                    _ltv_r, _ = ltv_inf_offset(_pre_k, _pre_lam, _arpu, ltv_offset_days)
                    _ltv_g, _ = ltv_inf_offset(_pre_k, _pre_lam, _gp, ltv_offset_days)
                _pre_results.append({
                    'segment': _pre_sv, 'n_customers': len(_pre_df),
                    'ltv_rev': _ltv_r, 'ltv_gp': _ltv_g,
                    'cac_cap': _ltv_g / cac_n,
                    'total_potential': _ltv_r * len(_pre_df),
                    'k': _pre_k, 'lam_days': _pre_lam + ltv_offset_days, 'λ_raw': _pre_lam,
                    'arpu_s': _arpu, 'arpu_long_s': _arpu_long, 'arpu_0_dorm_s': _arpu_0_dorm,
                    'R²': _pre_r2,
                    'acq_efficiency': (_ltv_g / cac_input) if cac_known else None,
                    'priority_score': _ltv_r * len(_pre_df),
                })
            except Exception:
                continue
        if _pre_results:
            _pre_df_result = pd.DataFrame(_pre_results).sort_values('ltv_rev', ascending=False).reset_index(drop=True)
            all_seg_results[_pre_sc] = _pre_df_result
            all_seg_details[_pre_sc] = sorted(_pre_results, key=lambda x: x['ltv_rev'], reverse=True)

# ══════════════════════════════════════════════════════════════
# Export buttons
# ══════════════════════════════════════════════════════════════

st.markdown(f"<div class='section-title'>{T('section_export')}</div>", unsafe_allow_html=True)

# ── Excel export ─────────────────────────────────────────────
if True:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference, Series

        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = 'Summary'
        hdr_fill = PatternFill('solid', start_color='1a1a1a', end_color='1a1a1a')
        hdr_font = Font(name='Calibri', bold=True, color='56B4D3', size=11)
        val_font = Font(name='Calibri', size=11, color='111111')

        ws['A1'] = T('excel_summary_title')
        ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='56B4D3')
        if client_name:
            ws['A2'] = f'{T("excel_client")}: {client_name}'
        if analyst_name:
            ws['A3'] = f'{T("excel_analyst")}: {analyst_name}'

        summary_data = [
            ('', ''),
            (T('excel_summary_title'), ''),
            (T('excel_customers'), len(df)),
            (T('excel_churned_active'), int(df['event'].sum())),
            (f'{T("excel_daily_arpu")} ({cur_symbol(CUR)})', round(arpu_daily, 2)),
            (T('excel_gpm'), f'{gpm:.1%}'),
            (f'{T("excel_daily_gp")} ({cur_symbol(CUR)})', round(gp_daily, 2)),
            (f'{T("excel_ltv_rev")} ({cur_symbol(CUR)})', round(ltv_rev, 0)),
            (f'{T("excel_ltv_gp")} ({cur_symbol(CUR)})', round(ltv_val, 0)),
            (f'{T("excel_cac_cap")} ({cac_label}) ({cur_symbol(CUR)})', round(cac_upper, 0)),
            ('', ''),
            ('Weibull Parameters', ''),
            (T('excel_weibull_k'), round(k, 4)),
            (T('excel_weibull_lam'), round(lam + ltv_offset_days if business_type == BIZ_SPOT else lam, 2)),
            (T('excel_r2'), round(r2, 4)),
        ]
        for i, (label, val) in enumerate(summary_data, start=5):
            ws.cell(i, 1, label).font = Font(name='Calibri', bold=(str(label) == T('excel_summary_title') or str(label) == 'Weibull Parameters'), color='333333', size=10)
            ws.cell(i, 2, val).font   = Font(name='Calibri', size=10, color='111111')
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 20

        # KM sheet
        ws2 = wb.create_sheet(T('excel_sheet_km'))
        ws2.append(['t (days)', 'S(t) KM Observed', 'S(t) Weibull Fit'])
        # t=0：生存率100%
        ws2.append([0, 1.0, 1.0])
        if business_type == BIZ_SPOT:
            # 都度購入型：オフセット後のkm_dfを使い、t軸を元に戻す
            ws2.append([int(ltv_offset_days), 1.0, 1.0])
            for _, row in km_df.iterrows():
                t_orig = int(row['t'] + ltv_offset_days)
                ws2.append([t_orig, round(row['S'], 6), round(float(weibull_s(row['t'], k, lam)), 6)])
        else:
            # サブスク：km_df（オフセット後）のt軸を+ltv_offset_daysで元スケールに戻す
            # t=0はヘッダー直後に既に追加済みなので重複しない
            if ltv_offset_days > 0:
                ws2.append([int(ltv_offset_days), 1.0, round(float(weibull_s(0, k, lam)), 6)])
            for _, row in km_df.iterrows():
                t      = row['t']
                t_plot = int(t + ltv_offset_days)
                ws2.append([t_plot, round(row['S'], 6), round(float(weibull_s(t, k, lam)), 6)])

        # Horizon sheet
        ws3 = wb.create_sheet(T('excel_sheet_interim'))
        ws3.append([T('tbl_horizon'), f'{T("tbl_ltv_rev")} ({cur_symbol(CUR)})', f'{T("tbl_pct_ltv")}', f'{T("tbl_cac_cap")} ({cur_symbol(CUR)})'])
        for h in horizons:
            if business_type == BIZ_SPOT:
                # 都度購入型：LTV_short（固定）+ LTV_long（Weibull積分）
                _dorm_off = dormancy_days or 180
                _h_short  = min(h, _dorm_off)
                _h_long   = max(h - _dorm_off, 0)
                _lh_short = _h_short * arpu_0_dorm
                _lh_long  = ltv_horizon_offset(k, lam, arpu_long, _h_long, 0) if _h_long > 0 else 0
                lh = _lh_short + _lh_long
            else:
                lh = ltv_horizon_offset(k, lam, arpu_daily, h, ltv_offset_days)
            ws3.append([h, round(lh, 0), round(lh/ltv_val*100, 1), round(lh/cac_n, 0)])
        # λ行
        ws3.append([T('tbl_lam_row', n=f'{round(lam + ltv_offset_days):,}'), round(lam_rev, 0), round(lam_rev/ltv_val*100, 1), round(lam_gp/cac_n, 0)])
        # 99%到達行
        ws3.append([f'99% ({int(days_99):,}{T("chart_days_suffix")})', round(rev_99, 0), 99.0, round(gp_99/cac_n, 0)])
        # LTV∞行
        ws3.append(['LTV∞', round(ltv_rev, 0), 100.0, round(ltv_val/cac_n, 0)])

        # セグメント別シート追加
        if segment_cols_input.strip():
            seg_cols_xl = [c.strip() for c in segment_cols_input.split(',') if c.strip() and c.strip() in df.columns]
            for sc in seg_cols_xl:
                # ── セグメント概要シート（all_seg_resultsから参照）──
                ws_seg = wb.create_sheet(f'SEG_{sc}'[:31])
                ws_seg.append([T('seg_tbl_segment'), T('seg_tbl_n'), T('seg_tbl_ltv_rev'), T('seg_tbl_ltv_gp'), T('seg_tbl_cac_cap'), 'k', T('seg_tbl_lam'), 'R²', 'ROI'])
                if sc in all_seg_results:
                    for _, _r in all_seg_results[sc].iterrows():
                        eff = round(_r['acq_efficiency'], 2) if _r['acq_efficiency'] is not None else '-'
                        ws_seg.append([_r['segment'], _r['n_customers'], round(_r['ltv_rev'],0), round(_r['ltv_gp'],0), round(_r['cac_cap'],0), round(_r['k'],4), round(_r['lam_days'],1), round(_r['R²'],4), eff])
                ws_seg.column_dimensions['A'].width = 20

                # ── セグメント別暫定LTV詳細シート（all_seg_resultsから参照）──
                ws_seg_hor = wb.create_sheet(f'SEG_{sc}_LTV'[:31])
                hor_header = [T('seg_tbl_segment'), T('tbl_horizon'), T('tbl_ltv_rev'), T('tbl_ltv_gp'), T('tbl_cac_cap'), T('tbl_pct_ltv')]
                ws_seg_hor.append(hor_header)
                hor_points = [180, 365, 730, 1095, 1825]
                if sc in all_seg_results:
                    for _, _r in all_seg_results[sc].iterrows():
                        k_s    = _r['k']
                        lam_s  = _r['λ_raw']
                        arpu_s = _r['arpu_s']
                        arpu_long_s  = _r['arpu_long_s']
                        arpu_0_dorm_s = _r['arpu_0_dorm_s']
                        gp_s   = arpu_s * gpm
                        ltv_inf_s = _r['ltv_rev']
                        lam_s_actual = _r['lam_days']
                        _dorm_s = dormancy_days or 180 if business_type == BIZ_SPOT else ltv_offset_days
                        sv = str(_r['segment'])
                        try:
                            # 通常ホライズン
                            for h in hor_points:
                                label = fmt_horizon(h)
                                if business_type == BIZ_SPOT:
                                    lh_r = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, h, _dorm_s)
                                    lh_g = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, h, _dorm_s)
                                else:
                                    lh_r = ltv_horizon_offset(k_s, lam_s, arpu_s, h, ltv_offset_days)
                                    lh_g = ltv_horizon_offset(k_s, lam_s, gp_s,   h, ltv_offset_days)
                                pct  = round(lh_r / ltv_inf_s * 100, 1) if ltv_inf_s > 0 else 0
                                ws_seg_hor.append([sv, label, round(lh_r,0), round(lh_g,0), round(lh_g/cac_n,0), pct])
                            # λ行
                            if business_type == BIZ_SPOT:
                                lam_r = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, lam_s_actual, _dorm_s)
                                lam_g = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, lam_s_actual, _dorm_s)
                            else:
                                lam_r = ltv_horizon_offset(k_s, lam_s, arpu_s, lam_s_actual, ltv_offset_days)
                                lam_g = ltv_horizon_offset(k_s, lam_s, gp_s,   lam_s_actual, ltv_offset_days)
                            lam_pct = round(lam_r / ltv_inf_s * 100, 1) if ltv_inf_s > 0 else 0
                            ws_seg_hor.append([sv, T('tbl_lam_row', n=str(int(lam_s_actual))), round(lam_r,0), round(lam_g,0), round(lam_g/cac_n,0), lam_pct])
                            # 99%到達行
                            try:
                                if business_type == BIZ_SPOT:
                                    days_99_s = brentq(lambda h: ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, h, _dorm_s) / ltv_inf_s - 0.99, 1, 500000)
                                    r99_r = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, days_99_s, _dorm_s)
                                    r99_g = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, days_99_s, _dorm_s)
                                else:
                                    days_99_s = brentq(lambda h: ltv_horizon_offset(k_s, lam_s, arpu_s, h, ltv_offset_days) / ltv_inf_s - 0.99, 1, 500000)
                                    r99_r = ltv_horizon_offset(k_s, lam_s, arpu_s, days_99_s, ltv_offset_days)
                                    r99_g = ltv_horizon_offset(k_s, lam_s, gp_s,   days_99_s, ltv_offset_days)
                                ws_seg_hor.append([sv, T('tbl_99pct_row', n=f'{int(days_99_s):,}'), round(r99_r,0), round(r99_g,0), round(r99_g/cac_n,0), 99.0])
                            except Exception:
                                r99_approx = ltv_inf_s * 0.99
                                g99_approx = r99_approx * gpm
                                ws_seg_hor.append([sv, 'LTV∞ 99%', round(r99_approx,0), round(g99_approx,0), round(g99_approx/cac_n,0), 99.0])
                            # LTV∞行
                            ltv_g_s = _r['ltv_gp']
                            ws_seg_hor.append([sv, 'LTV∞', round(ltv_inf_s,0), round(ltv_g_s,0), round(ltv_g_s/cac_n,0), 100.0])
                            # 空行で区切り
                            ws_seg_hor.append([''] * 6)
                        except Exception:
                            continue
                ws_seg_hor.column_dimensions['A'].width = 20
                ws_seg_hor.column_dimensions['B'].width = 18

        xl_buf = io.BytesIO()
        wb.save(xl_buf)
        xl_buf.seek(0)
        import base64 as _b64
        _xl_b64 = _b64.b64encode(xl_buf.read()).decode()
        _fn_xl = f"LTV_Analysis_{client_name or 'report'}.xlsx"
        _xl_href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{_xl_b64}" download="{_fn_xl}" class="dl-btn">.xlsx</a>'
        _xl_html = _xl_href
    except Exception as e:
        _xl_html = f'<span class="dl-btn-err">.xlsx Error</span>'

# ── PowerPoint export ─────────────────────────────────────────
if True:
    try:
        import os as _os, sys as _sys
        _here = _os.path.dirname(_os.path.abspath(__file__))
        _sys.path.insert(0, _here)
        from pptx_export import generate_pptx
        from pptx import Presentation

        _TMPL_PATH = _os.path.join(_here, 'LTV-analyzer.pptx')

        _billing_disp = billing_cycle_display_text if 'billing_cycle_display_text' in dir() else T('biz_spot')

        # S4ガイドデータを組み立て
        _lam_display = lam + ltv_offset_days if business_type == BIZ_SPOT else lam
        _s4_guide = {
            'lam_desc': lam_desc,
            'k_desc': k_desc,
            'ltv_rev': ltv_rev,
            'pct_1y': pct_1y, 'ltv_1y': ltv_1y,
            'pct_2y': pct_2y, 'ltv_2y': ltv_2y,
            'pct_3y': pct_3y, 'ltv_3y': ltv_3y,
            'cac_upper': cac_upper,
            'cac_recover_rev_str': cac_recover_rev_str,
            'cac_recover_gp_str': cac_recover_gp_str,
            'lam_actual_round': round(_lam_display),
            'lam_years': _lam_display / 365,
            'lam_gp': lam_gp,
            'lam_meaning': T('insight_cac_design', lam=f'{round(_lam_display):,}', y=_lam_display/365, gp=fmt_c(lam_gp, CUR)) if business_type == BIZ_SPOT else T('insight_cac_design', lam=f'{round(_lam_display):,}', y=_lam_display/365, gp=fmt_c(lam_gp, CUR)),
        }

        # 異常値処理の表示文字列
        _outlier_parts = []
        if outlier_upper_pct > 0:
            _outlier_parts.append(T('hist_upper_pct', pct=outlier_upper_pct))
        if outlier_lower_pct > 0:
            _outlier_parts.append(T('hist_lower_pct', pct=outlier_lower_pct))
        _join = ', ' if get_lang() == 'en' else '、'
        _outlier_label = _join.join(_outlier_parts) if _outlier_parts else T('excel_none')

        pptx_buf = generate_pptx(
            tmpl_path=_TMPL_PATH,
            k=k, lam=lam, lam_actual=lam_actual, r2=r2,
            arpu_daily=arpu_daily, gpm=gpm, gp_daily=gp_daily,
            cac_n=cac_n, cac_upper=cac_upper,
            ltv_rev=ltv_rev, lam_rev=lam_rev, lam_gp=lam_gp,
            rev_99=rev_99, gp_99=gp_99, days_99=days_99,
            t_range=t_range, rev_line=rev_line, gp_line=gp_line,
            cac_line=cac_line, x_max=x_max,
            tbl_rows=tbl_rows, horizons=horizons,
            ltv_offset_days=ltv_offset_days,
            ltv_horizon_offset=ltv_horizon_offset,
            ltv_horizon_spot=ltv_horizon_spot,
            business_type=business_type,
            dormancy_days=dormancy_days if 'dormancy_days' in dir() else None,
            buf1=buf1, buf2=buf2,
            df=df, client_name=client_name, analyst_name=analyst_name,
            billing_cycle_display=_billing_disp,
            k_summary=k_summary, r2_summary=r2_summary,
            cac_label=cac_label,
            cac_recover_rev_str=cac_recover_rev_str,
            cac_recover_gp_str=cac_recover_gp_str,
            segment_cols_input=segment_cols_input,
            _compute_km_df=_compute_km_df,
            _fit_weibull_df=_fit_weibull_df,
            ltv_inf=ltv_inf,
            fmt_horizon=fmt_horizon,
            s4_guide_data=_s4_guide,
            report_title=report_title,
            arpu_0_dorm=arpu_0_dorm if 'arpu_0_dorm' in dir() else arpu_daily,
            arpu_long=arpu_long if 'arpu_long' in dir() else arpu_daily,
            outlier_label=_outlier_label,
            all_seg_results=all_seg_results,
            cur=CUR,
        )

        import base64 as _b64
        _pp_b64 = _b64.b64encode(pptx_buf.read()).decode()
        _fn_pp  = f"LTV_Analysis_{client_name or 'report'}.pptx"
        _pp_href = (f'<a href="data:application/vnd.openxmlformats-officedocument'
                    f'.presentationml.presentation;base64,{_pp_b64}" '
                    f'download="{_fn_pp}" class="dl-btn">.pptx</a>')
        _pp_html = _pp_href
    except ImportError as _ie:
        _pp_html = f'<span class="dl-btn-err">.pptx Not supported: {str(_ie)[:80]}</span>'
    except Exception as e:
        import traceback as _tb
        _tb_str = _tb.format_exc().replace('\n', ' | ')[-300:]
        _pp_html = (f'<span class="dl-btn-err">.pptx Error: {str(e)[:150]}'
                    f'<br><small style="font-size:0.6rem;opacity:0.7">{_tb_str}</small></span>')

# ── PDF export ────────────────────────────────────────────────
if True:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Image as RLImage, Table as RLTable,
                                         TableStyle, PageBreak, KeepTogether)
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))

        # ── デザイントークン ──
        _BG       = rl_colors.HexColor('#0E1117')
        _BG2      = rl_colors.HexColor('#161B22')
        _BG3      = rl_colors.HexColor('#1C2333')
        _ACCENT   = rl_colors.HexColor('#56b4d3')
        _ACCENT2  = rl_colors.HexColor('#a8dadc')
        _TEXT     = rl_colors.HexColor('#c8ccd0')
        _TEXT_DIM = rl_colors.HexColor('#6e7681')
        _GRID     = rl_colors.HexColor('#2a3040')
        _WHITE    = rl_colors.HexColor('#e6edf3')
        _GOLD     = rl_colors.HexColor('#c8b89a')

        CONTENT_W = 16.5 * cm  # グラフ・テーブル統一幅
        COL_GAP   = 0.4 * cm
        HALF_W    = (CONTENT_W - COL_GAP) / 2

        pdf_buf = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buf, pagesize=A4,
                                leftMargin=2 * cm, rightMargin=2 * cm,
                                topMargin=1.5 * cm, bottomMargin=1.5 * cm)

        # ── スタイル ──
        s_title  = ParagraphStyle('T',  fontName='HeiseiMin-W3', fontSize=20, leading=26,
                                   textColor=_ACCENT, spaceAfter=4)
        s_sub    = ParagraphStyle('Su', fontName='HeiseiMin-W3', fontSize=11, leading=15,
                                   textColor=_TEXT_DIM, spaceAfter=2)
        s_chap   = ParagraphStyle('Ch', fontName='HeiseiMin-W3', fontSize=14, leading=20,
                                   textColor=_ACCENT, spaceBefore=0, spaceAfter=8)
        s_h3     = ParagraphStyle('H3', fontName='HeiseiMin-W3', fontSize=10, leading=14,
                                   textColor=_ACCENT2, spaceBefore=10, spaceAfter=4)
        s_body   = ParagraphStyle('B',  fontName='HeiseiMin-W3', fontSize=9, leading=13,
                                   textColor=_TEXT, spaceAfter=3)
        s_small  = ParagraphStyle('Sm', fontName='HeiseiMin-W3', fontSize=7.5, leading=11,
                                   textColor=_TEXT_DIM, spaceAfter=2)
        s_label  = ParagraphStyle('Lb', fontName='HeiseiMin-W3', fontSize=8, leading=12,
                                   textColor=_ACCENT, spaceAfter=1, spaceBefore=4)
        s_prompt = ParagraphStyle('Pr', fontName='HeiseiMin-W3', fontSize=7.5, leading=11,
                                   textColor=_TEXT, spaceAfter=1, leftIndent=8, rightIndent=8)
        s_rec    = ParagraphStyle('Rc', fontName='HeiseiMin-W3', fontSize=9, leading=13,
                                   textColor=_ACCENT, backColor=_BG3,
                                   borderPadding=8, spaceAfter=6, spaceBefore=8,
                                   leftIndent=8, rightIndent=8)

        # テーブルスタイルヘルパー
        def _dark_tbl_style(has_title_col=True):
            s = [
                ('BACKGROUND',    (0, 0), (-1, 0), _BG3),
                ('TEXTCOLOR',     (0, 0), (-1, 0), _ACCENT),
                ('TEXTCOLOR',     (0, 1), (-1, -1), _TEXT),
                ('FONTNAME',      (0, 0), (-1, -1), 'HeiseiMin-W3'),
                ('FONTSIZE',      (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS',(0, 1), (-1, -1), [_BG, _BG2]),
                ('GRID',          (0, 0), (-1, -1), 0.3, _GRID),
                ('LEFTPADDING',   (0, 0), (-1, -1), 6),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
                ('TOPPADDING',    (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN',         (0, 0), (0, -1), 'LEFT'),
                ('ALIGN',         (1, 0), (-1, -1), 'RIGHT'),
            ]
            if has_title_col:
                s.append(('TEXTCOLOR', (0, 1), (0, -1), _GOLD))
            return TableStyle(s)

        story = []

        # ═══════════════════════════════════════════════════════════
        # Chapter 1: 表紙
        # ═══════════════════════════════════════════════════════════
        story.append(Spacer(1, 4 * cm))
        _cover_sub = report_title if report_title else 'Kaplan–Meier × Weibull Model'
        story.append(Paragraph(_cover_sub, s_title))
        story.append(Paragraph(T('pdf_title'), s_sub))
        story.append(Spacer(1, 1.5 * cm))
        import datetime as _dt_pdf
        if client_name:
            story.append(Paragraph(client_name, s_body))
        story.append(Paragraph(_dt_pdf.date.today().strftime(T("pdf_date_format")), s_body))
        if analyst_name:
            story.append(Paragraph(analyst_name, s_body))

        # ═══════════════════════════════════════════════════════════
        # Chapter 2: 分析結果サマリー
        # ═══════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph(T('pdf_chapter_summary'), s_chap))
        story.append(Spacer(1, 0.3 * cm))

        # 結論テキスト
        _k_desc_pdf = T("excel_early_churn") if k < 1 else T("excel_late_churn")
        story.append(Paragraph(T('pdf_conclusion'), s_h3))
        story.append(Paragraph(summary_text, s_body))
        story.append(Spacer(1, 0.8 * cm))

        # サマリーテーブル（異常値除外行追加）
        _outlier_parts_pdf = []
        if outlier_upper_pct > 0:
            _outlier_parts_pdf.append(f'{T("hist_upper_pct", pct=outlier_upper_pct)}')
        if outlier_lower_pct > 0:
            _outlier_parts_pdf.append(f'{T("hist_lower_pct", pct=outlier_lower_pct)}')
        _outlier_label_pdf = (", " if get_lang()=="en" else "、").join(_outlier_parts_pdf) if _outlier_parts_pdf else T("excel_none")

        _sum_data = [
            [T('excel_metric'), ''],
            [T('excel_ltv_rev'), f'{fmt_c(ltv_rev, CUR)}'],
            [T('excel_ltv_gp'), f'{fmt_c(ltv_val, CUR)}'],
            [f'{T("excel_cac_cap")} ({cac_label})', f'{fmt_c(cac_upper, CUR)}'],
            [T('excel_weibull_k'), f'{k:.4f}  →  {_k_desc_pdf}'],
            [T('excel_weibull_lam'), f'{lam_display:.1f}{T("chart_days_suffix")} (~{lam_display/365:.1f}{T("chart_year_suffix")})'],
            [T('excel_r2'), f'{r2:.4f}  →  {T("excel_r2_good") if r2 >= 0.9 else "△"}'],
            [T('excel_customers'), f'{len(df):,}'],
            [T('excel_outlier'), f'{n_outlier:,} ({_outlier_label_pdf})'],
            [T('excel_churned_active'), f'{int(df["event"].sum()):,} / {int((df["event"]==0).sum()):,}'],
            [T('excel_daily_arpu'), f'{fmt_c(arpu_daily, CUR, 2)}'],
            [T('excel_gpm'), f'{gpm:.1%}'],
            [T('excel_biz_type'), T('biz_spot') if business_type == BIZ_SPOT else T('biz_subscription')],
        ]
        if business_type == BIZ_SPOT:
            _sum_data.append([T('excel_dormancy'), dormancy_label])
        else:
            _prorate_val = 'ON' if (prorate_cancel if 'prorate_cancel' in dir() else False) else 'OFF'
            _sum_data.append([T('excel_prorate'), _prorate_val])
        _col_w = CONTENT_W / 2
        _sum_t = RLTable(_sum_data, colWidths=[_col_w, _col_w])
        _sum_style = _dark_tbl_style(has_title_col=True)
        # ヘッダー行を結合して中央配置、結論と同じフォントサイズ
        _sum_style.add('SPAN', (0, 0), (1, 0))
        _sum_style.add('ALIGN', (0, 0), (1, 0), 'CENTER')
        _sum_style.add('FONTSIZE', (0, 0), (1, 0), 10)
        _sum_style.add('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        _sum_t.setStyle(_sum_style)
        story.append(_sum_t)

        # ═══════════════════════════════════════════════════════════
        # Chapter 3: モデル信頼性
        # ═══════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph(T('pdf_chapter_reliability'), s_chap))
        story.append(Spacer(1, 0.3 * cm))

        story.append(Paragraph('Survival Curve', s_h3))
        buf1.seek(0)
        story.append(RLImage(buf1, width=CONTENT_W, height=CONTENT_W * 0.48))
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(
            T('chart_survival_caption') + f' k={k:.3f}, λ={lam_display:.1f}{T("chart_days_suffix")}',
            s_small
        ))
        story.append(Spacer(1, 0.8 * cm))

        story.append(Paragraph('Weibull Linearization Plot', s_h3))
        buf2.seek(0)
        story.append(RLImage(buf2, width=CONTENT_W, height=CONTENT_W * 0.48))
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(
            T('chart_linearization_caption'),
            s_small
        ))

        # ═══════════════════════════════════════════════════════════
        # Chapter 4: Interim LTV
        # ═══════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph(T('pdf_chapter_interim'), s_chap))
        story.append(Spacer(1, 0.3 * cm))

        # LTV推移グラフ（matplotlib dark）
        import matplotlib.pyplot as plt_pdf
        import numpy as np_pdf

        # PDF用グラフの日本語フォント設定
        if _JP_FONT_NAME:
            plt_pdf.rcParams['font.family'] = _JP_FONT_NAME
        elif _JP_FONT_PATH:
            from matplotlib.font_manager import FontProperties as _FP_pdf
            _jp_fp = _FP_pdf(fname=_JP_FONT_PATH)
            plt_pdf.rcParams['font.family'] = _jp_fp.get_name()

        fig_ltv_pdf, ax_lp = plt_pdf.subplots(figsize=(10, 4.5))
        fig_ltv_pdf.patch.set_facecolor('#0E1117')
        ax_lp.set_facecolor('#0E1117')

        ax_lp.plot(t_range, rev_line, color='#56b4d3', lw=2, label=T('chart_ltv_rev'))
        ax_lp.plot(t_range, gp_line,  color='#a8dadc', lw=2, ls='--', label=T('chart_ltv_gp'))
        ax_lp.plot(t_range, cac_line, color='#4a7a8a', lw=1.5, ls=':', label=T('chart_cac_cap'))
        ax_lp.axhline(y=ltv_rev, color='#56b4d3', lw=1, ls=':', alpha=0.4)
        ax_lp.axvline(x=lam_actual, color='#a8dadc', lw=1.5, ls='--', alpha=0.5)

        # マーカー
        _marker_days_pdf = sorted(set([180, 365, 730, 1095, 1825]))
        for _td, _tc, _tl in [(rev_line, '#56b4d3', 'rev'), (gp_line, '#a8dadc', 'gp')]:
            _mx = [d for d in _marker_days_pdf if d <= max(t_range)]
            _my = [_td[min(range(len(t_range)), key=lambda i: abs(t_range[i] - d))] for d in _mx]
            ax_lp.scatter(_mx, _my, color=_tc, s=25, zorder=5)

        ax_lp.annotate(f'LTV∞ {fmt_c(ltv_rev, CUR)}', xy=(max(t_range)*0.7, ltv_rev),
                       fontsize=8, color='#56b4d3', va='bottom')
        ax_lp.annotate(T('chart_lam_days', n=round(lam_actual)), xy=(lam_actual, max(rev_line)*0.5),
                       fontsize=8, color='#a8dadc', ha='center',
                       bbox=dict(boxstyle='round,pad=0.2', fc='#0E1117', ec='none'))

        _tick_v = [180, 365, 730, 1095, 1460, 1825]
        _tick_l = [f'180{T("chart_days_suffix")}', f'1{T("chart_year_suffix")}', f'2{T("chart_year_suffix")}', f'3{T("chart_year_suffix")}', f'4{T("chart_year_suffix")}', f'5{T("chart_year_suffix")}']
        ax_lp.set_xticks([v for v in _tick_v if v <= max(t_range)])
        ax_lp.set_xticklabels([l for v, l in zip(_tick_v, _tick_l) if v <= max(t_range)])

        if _JP_FONT_NAME:
            ax_lp.set_xlabel(T('chart_duration'), fontsize=9, color='#888',
                            fontfamily=_JP_FONT_NAME)
            ax_lp.set_ylabel(T('chart_amount'), fontsize=9, color='#888',
                            fontfamily=_JP_FONT_NAME)
        else:
            ax_lp.set_xlabel('Days', fontsize=9, color='#888')
            ax_lp.set_ylabel('Amount', fontsize=9, color='#888')

        ax_lp.tick_params(colors='#888', labelsize=8)
        ax_lp.yaxis.set_major_formatter(plt_pdf.FuncFormatter(lambda x, _: f'{fmt_c(x, CUR)}'))
        ax_lp.grid(True, alpha=0.15, color='#2a3040')
        ax_lp.legend(fontsize=8, facecolor='#0E1117', edgecolor='#2a3040',
                    labelcolor='#ccc', loc='upper left')
        for spine in ax_lp.spines.values():
            spine.set_color('#2a3040')

        fig_ltv_pdf.tight_layout()
        buf_ltv_pdf = io.BytesIO()
        fig_ltv_pdf.savefig(buf_ltv_pdf, format='png', dpi=150,
                           facecolor='#0E1117', bbox_inches='tight')
        buf_ltv_pdf.seek(0)
        plt_pdf.close(fig_ltv_pdf)

        story.append(RLImage(buf_ltv_pdf, width=CONTENT_W, height=CONTENT_W * 0.42))
        story.append(Spacer(1, 0.6 * cm))

        # 暫定LTVテーブル（チャプタータイトルと重複するためh3は省略）
        _ltv_hdr = [T('tbl_horizon'), T('tbl_ltv_rev'), T('tbl_ltv_gp'), T('tbl_cac_cap'), T('tbl_pct_ltv')]
        _ltv_tdata = [_ltv_hdr]
        for row in tbl_rows:
            _ltv_tdata.append([
                row['horizon'], row['ltv_rev'], row['ltv_gp'],
                row['cac_cap'], row['pct'],
            ])

        _title_cw = 5.5 * cm
        _data_cw  = (CONTENT_W - _title_cw) / 4
        _ltv_t = RLTable(_ltv_tdata, colWidths=[_title_cw] + [_data_cw] * 4)
        _ltv_t.setStyle(_dark_tbl_style(has_title_col=True))
        story.append(_ltv_t)
        story.append(Spacer(1, 0.5 * cm))

        # 解釈ガイド
        story.append(Paragraph(T('insight_title'), s_h3))
        _guide_lines = [
            lam_desc,
            k_desc,
            _insight_ltv_inf,
            f'{_insight_1y} {_insight_2y} {_insight_3y}',
            _insight_cac,
            _insight_design,
        ]
        for gl in _guide_lines:
            story.append(Paragraph(f'・{gl}', s_small))

        # ═══════════════════════════════════════════════════════════
        # Chapter 5+6: セグメント別分析
        # ═══════════════════════════════════════════════════════════
        if segment_cols_input.strip():
            seg_cols_pdf = [c.strip() for c in segment_cols_input.split(',')
                           if c.strip() and c.strip() in df.columns]
            for sc in seg_cols_pdf:
                story.append(PageBreak())
                story.append(Paragraph(f'{T("section_segment")}: {sc}', s_chap))
                story.append(Spacer(1, 0.3 * cm))

                seg_vals = df[sc].dropna().unique()
                pdf_rows = [[T('seg_tbl_segment'), T('seg_tbl_n'), T('seg_tbl_ltv_rev'), T('seg_tbl_ltv_gp'),
                             T('seg_tbl_cac_cap'), 'k', T('seg_tbl_lam'), 'R²']]
                best_pdf = None
                avg_ltv_pdf = []
                _seg_results = []  # 加重平均用
                # all_seg_resultsから参照（独自Weibullフィットしない）
                if sc in all_seg_results:
                    for _, _r in all_seg_results[sc].iterrows():
                        pdf_rows.append([str(_r['segment']), f'{int(_r["n_customers"]):,}',
                                        f'{fmt_c(_r["ltv_rev"], CUR)}', f'{fmt_c(_r["ltv_gp"], CUR)}',
                                        f'{fmt_c(_r["cac_cap"], CUR)}',
                                        f'{_r["k"]:.3f}', f'{_r['lam_days']:.1f}', f'{_r["R²"]:.3f}'])
                        _seg_results.append({'n': int(_r['n_customers']), 'ltv_r': _r['ltv_rev'],
                                            'ltv_g': _r['ltv_gp'], 'cac': _r['cac_cap']})
                        avg_ltv_pdf.append(_r['ltv_rev'])
                        if best_pdf is None or _r['ltv_rev'] > best_pdf['ltv_r']:
                            best_pdf = {'seg': str(_r['segment']), 'ltv_r': _r['ltv_rev'],
                                       'ltv_g': _r['ltv_gp'], 'cac': _r['cac_cap']}

                # TOP PICK
                if best_pdf and avg_ltv_pdf:
                    avg_pp = sum(avg_ltv_pdf) / len(avg_ltv_pdf)
                    prem = (best_pdf['ltv_r'] - avg_pp) / avg_pp * 100
                    story.append(Paragraph(f'{sc}: LTV∞', s_h3))
                    story.append(Paragraph(
                        f"TOP PICK {best_pdf['seg']}",
                        s_label
                    ))
                    story.append(Paragraph(
                        f"LTV∞ (Rev): {fmt_c(best_pdf['ltv_r'], CUR)} (vs avg +{prem:.1f}%)"
                        f" | CAC Ceiling {fmt_c(best_pdf['cac'], CUR)}",
                        s_small
                    ))
                    story.append(Spacer(1, 0.4 * cm))

                # LTV∞バーチャート
                if len(pdf_rows) > 1:
                    import matplotlib.pyplot as plt_seg_bar
                    _seg_names = [r[0] for r in pdf_rows[1:]]
                    _seg_ltvs = [float(r[2].replace(cur_symbol(CUR), '').replace(',', '')) for r in pdf_rows[1:]]
                    # 降順ソートして上位10件に制限
                    _sorted_idx = sorted(range(len(_seg_ltvs)), key=lambda i: _seg_ltvs[i], reverse=True)[:10]
                    _seg_names = [_seg_names[i] for i in _sorted_idx]
                    _seg_ltvs = [_seg_ltvs[i] for i in _sorted_idx]
                    # barh用に昇順に戻す（下が低い→上が高い）
                    _seg_names = _seg_names[::-1]
                    _seg_ltvs = _seg_ltvs[::-1]

                    _n_bars = len(_seg_names)
                    _fig_h = max(2.0, _n_bars * 0.6)
                    _fig_w = 10
                    fig_bar, ax_bar = plt_seg_bar.subplots(figsize=(_fig_w, _fig_h))
                    fig_bar.patch.set_facecolor('#0E1117')
                    ax_bar.set_facecolor('#0E1117')
                    # TOP PICKは濃い青、それ以外はダークティール
                    _bar_colors = ['#56b4d3' if _seg_names[i] == best_pdf['seg'] else '#2a4a5a'
                                   for i in range(len(_seg_names))] if best_pdf else ['#2a4a5a'] * len(_seg_names)
                    bars = ax_bar.barh(range(len(_seg_names)), _seg_ltvs, color=_bar_colors, height=0.6, zorder=3)
                    ax_bar.set_yticks(range(len(_seg_names)))
                    ax_bar.set_yticklabels(_seg_names, fontsize=8, color='#ccc')
                    ax_bar.tick_params(colors='#888', labelsize=8)
                    ax_bar.xaxis.set_major_formatter(plt_seg_bar.FuncFormatter(lambda x, _: f'{fmt_c(x, CUR)}'))
                    ax_bar.xaxis.tick_top()
                    ax_bar.xaxis.set_label_position('top')
                    ax_bar.set_axisbelow(True)
                    ax_bar.grid(True, axis='x', alpha=0.15, color='#2a3040', zorder=0)
                    for sp in ax_bar.spines.values():
                        sp.set_color('#2a3040')
                    # 値ラベル（棒グラフ内の右寄せ、白太字）
                    for bar, val in zip(bars, _seg_ltvs):
                        ax_bar.text(bar.get_width() - max(_seg_ltvs) * 0.01,
                                   bar.get_y() + bar.get_height() / 2,
                                   f'{fmt_c(val, CUR)}', va='center', ha='right',
                                   fontsize=8, color='#ffffff', fontweight='bold')
                    fig_bar.tight_layout()
                    buf_bar = io.BytesIO()
                    fig_bar.savefig(buf_bar, format='png', dpi=120, facecolor='#0E1117', bbox_inches='tight')
                    buf_bar.seek(0)
                    plt_seg_bar.close(fig_bar)
                    # 縦横比をfigと一致させる
                    _rl_bar_h = CONTENT_W * (_fig_h / _fig_w)
                    story.append(RLImage(buf_bar, width=CONTENT_W, height=_rl_bar_h))
                    _total_segs = len(pdf_rows) - 1
                    if _total_segs > 10:
                        story.append(Paragraph(
                            T('pdf_note_bar_chart', max=10, total=_total_segs),
                            s_small
                        ))
                    story.append(Spacer(1, 0.4 * cm))

                # サマリーテーブル（上位10件 + 加重平均行）
                story.append(Paragraph(f'{sc}: {T("pptx_summary")}', s_h3))
                pdf_rows_show = ([pdf_rows[0]] +
                                sorted(pdf_rows[1:],
                                       key=lambda x: float(x[2].replace(cur_symbol(CUR), '').replace(',', '')),
                                       reverse=True)[:10])
                # 加重平均行
                if _seg_results:
                    _total_n = sum(s['n'] for s in _seg_results)
                    _wavg_r = sum(s['ltv_r'] * s['n'] for s in _seg_results) / _total_n
                    _wavg_g = sum(s['ltv_g'] * s['n'] for s in _seg_results) / _total_n
                    _wavg_c = sum(s['cac'] * s['n'] for s in _seg_results) / _total_n
                    pdf_rows_show.append([T('seg_weighted_avg'), f'{_total_n:,}',
                                         f'{fmt_c(_wavg_r, CUR)}', f'{fmt_c(_wavg_g, CUR)}',
                                         f'{fmt_c(_wavg_c, CUR)}', '—', '—', '—'])

                _seg_title_cw = 3.8 * cm
                _seg_data_cw = (CONTENT_W - _seg_title_cw) / 7
                t_seg = RLTable(pdf_rows_show,
                               colWidths=[_seg_title_cw] + [_seg_data_cw] * 7)
                _seg_style = _dark_tbl_style(has_title_col=True)
                _seg_style.add('FONTSIZE', (0, 0), (-1, -1), 7)
                # 加重平均行をハイライト
                if _seg_results:
                    _last_row = len(pdf_rows_show) - 1
                    _seg_style.add('BACKGROUND', (0, _last_row), (-1, _last_row), _BG3)
                    _seg_style.add('TEXTCOLOR', (0, _last_row), (-1, _last_row), _ACCENT2)
                t_seg.setStyle(_seg_style)
                story.append(t_seg)
                story.append(Spacer(1, 0.2 * cm))

                # NOTE
                _n_segs = len(pdf_rows) - 1
                story.append(Paragraph(
                    T('pdf_note_summary_table', max=10, total=_n_segs),
                    s_small
                ))
                story.append(Spacer(1, 0.4 * cm))

                # ── セグメント詳細 ──
                story.append(PageBreak())
                story.append(Paragraph(f'{T("pdf_chapter_segment")}: {sc}', s_chap))
                story.append(Spacer(1, 0.6 * cm))

                _seg_detail_count = 0
                # all_seg_resultsのLTV∞降順で出力（all_seg_detailsから参照）
                _pdf_seg_list = all_seg_details.get(sc, [])
                for _pdf_sr in _pdf_seg_list:
                    sv = _pdf_sr['segment']
                    df_sv2 = df[df[sc] == sv]
                    if len(df_sv2) < 10 or df_sv2['event'].sum() < 5:
                        continue
                    try:
                        # パラメータはall_seg_resultsから取得（再計算しない）
                        k_sv2    = _pdf_sr['k']
                        lam_sv2  = _pdf_sr['λ_raw']
                        r2_sv2   = _pdf_sr['R²']
                        arpu_sv2 = _pdf_sr['arpu_s']
                        arpu_long_sv2  = _pdf_sr['arpu_long_s']
                        arpu_0_dorm_sv2 = _pdf_sr['arpu_0_dorm_s']
                        ltv_inf_sv2 = _pdf_sr['ltv_rev']
                        lam_sv2_actual = _pdf_sr['lam_days']
                        _dorm_sv2 = dormancy_days or 180 if business_type == BIZ_SPOT else ltv_offset_days

                        # グラフ描画用にKMは計算（数値には使わない）
                        if ltv_offset_days > 0:
                            _df_sv2_fit = df_sv2.copy()
                            _df_sv2_fit['duration'] = _df_sv2_fit['duration'] - ltv_offset_days
                            _df_sv2_fit.loc[_df_sv2_fit['duration'] <= 0, 'event'] = 0
                            _df_sv2_fit.loc[_df_sv2_fit['duration'] <= 0, 'duration'] = 1
                            _df_sv2_fit['duration'] = _df_sv2_fit['duration'].clip(lower=1)
                            km_sv2 = _compute_km_df(_df_sv2_fit)
                        else:
                            km_sv2 = _compute_km_df(df_sv2)

                        story.append(Paragraph(
                            f'{str(sv)}　（{len(df_sv2):,}件 / LTV∞ {fmt_c(ltv_inf_sv2, CUR)} '
                            f'/ k={k_sv2:.3f} / λ={lam_sv2:.1f}日 / R²={r2_sv2:.3f}）',
                            s_label
                        ))

                        # 2グラフ横並び（dark theme）
                        import matplotlib.pyplot as plt_pdf_seg
                        import numpy as np_pdf_seg
                        fig_2g, (ax_g1, ax_g2) = plt_pdf_seg.subplots(1, 2, figsize=(12, 4))
                        fig_2g.patch.set_facecolor('#0E1117')
                        ax_g1.set_facecolor('#0E1117')
                        ax_g2.set_facecolor('#0E1117')

                        # 左：生存曲線
                        ax_g1.step(km_sv2['t'], km_sv2['S'], color='#56b4d3', lw=1.5,
                                  label='KM Curve')
                        t_r2 = km_sv2['t'].values
                        ax_g1.plot(t_r2,
                                  [float(weibull_s(t, k_sv2, lam_sv2)) for t in t_r2],
                                  '--', color='#a8dadc', lw=1.2, label='Weibull Fit')
                        ax_g1.set_xlabel('Days', fontsize=9, color='#888')
                        ax_g1.set_ylabel('S(t)', fontsize=9, color='#888')
                        ax_g1.tick_params(colors='#888', labelsize=8)
                        ax_g1.legend(fontsize=7, facecolor='#0E1117', edgecolor='#2a3040',
                                    labelcolor='#ccc')
                        ax_g1.grid(True, alpha=0.15, color='#2a3040')
                        ax_g1.set_title(f'Survival: {str(sv)}', fontsize=9, color='#ccc')
                        for sp in ax_g1.spines.values():
                            sp.set_color('#2a3040')

                        # 右：Weibull直線化
                        km_fit_p = km_sv2[km_sv2['S'] > 0]
                        ln_t_p = np_pdf_seg.log(km_fit_p['t'].values.astype(float) + 1e-10)
                        ln_neg_p = np_pdf_seg.log(
                            -np_pdf_seg.log(km_fit_p['S'].values.astype(float) + 1e-15))
                        valid_p = np_pdf_seg.isfinite(ln_t_p) & np_pdf_seg.isfinite(ln_neg_p)
                        ln_t_p, ln_neg_p = ln_t_p[valid_p], ln_neg_p[valid_p]
                        if len(ln_t_p) > 1:
                            slope_p, int_p, _, _, _ = __import__('scipy').stats.linregress(
                                ln_t_p, ln_neg_p)
                            x_lp2 = np_pdf_seg.linspace(ln_t_p.min(), ln_t_p.max(), 100)
                            ax_g2.scatter(ln_t_p, ln_neg_p, color='#56b4d3', s=12, alpha=0.7)
                            ax_g2.plot(x_lp2, slope_p * x_lp2 + int_p, '--',
                                      color='#a8dadc', lw=1.5, label=f'R²={r2_sv2:.3f}')
                            ax_g2.annotate(f'y={slope_p:.3f}x+{int_p:.3f}',
                                          xy=(0.05, 0.93), xycoords='axes fraction',
                                          fontsize=7, color='#888')
                        ax_g2.set_xlabel('ln(t)', fontsize=9, color='#888')
                        ax_g2.set_ylabel('ln(−ln(S(t)))', fontsize=9, color='#888')
                        ax_g2.tick_params(colors='#888', labelsize=8)
                        ax_g2.legend(fontsize=7, facecolor='#0E1117', edgecolor='#2a3040',
                                    labelcolor='#ccc')
                        ax_g2.grid(True, alpha=0.15, color='#2a3040')
                        ax_g2.set_title(f'Weibull Lin.: {str(sv)}', fontsize=9, color='#ccc')
                        for sp in ax_g2.spines.values():
                            sp.set_color('#2a3040')

                        fig_2g.tight_layout(pad=1.5)
                        buf_2g = io.BytesIO()
                        fig_2g.savefig(buf_2g, format='png', dpi=120,
                                      facecolor='#0E1117', bbox_inches='tight')
                        buf_2g.seek(0)
                        plt_pdf_seg.close(fig_2g)

                        # 横並び：左グラフ左端 = 単一グラフ左端、右グラフ右端 = 単一グラフ右端
                        story.append(RLImage(buf_2g, width=CONTENT_W,
                                            height=CONTENT_W * 0.28))
                        story.append(Spacer(1, 0.15 * cm))

                        # 暫定LTVテーブル
                        hor_data2 = [[T('tbl_horizon'), T('tbl_ltv_rev'), 'LTV∞ %',
                                     T('tbl_cac_cap'), T('tbl_pct_ltv')]]
                        for h in horizons:
                            if business_type == BIZ_SPOT:
                                lh_sv2 = ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2, arpu_long_sv2, h, _dorm_sv2)
                                lh_gp_sv2 = ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2*gpm, arpu_long_sv2*gpm, h, _dorm_sv2)
                            else:
                                lh_sv2 = ltv_horizon_offset(k_sv2, lam_sv2, arpu_sv2, h, ltv_offset_days)
                                lh_gp_sv2 = lh_sv2 * gpm
                            label_h = f'{h}{T("chart_days_suffix")}' if h < 365 else f'{h // 365}{T("chart_year_suffix")}'
                            _pct_sv2 = lh_sv2 / ltv_inf_sv2 * 100 if ltv_inf_sv2 > 0 else 0
                            hor_data2.append([
                                label_h, f'{fmt_c(lh_sv2, CUR)}',
                                f'{_pct_sv2:.1f}%',
                                f'{fmt_c(lh_gp_sv2 / cac_n, CUR)}',
                                f'{_pct_sv2:.1f}%',
                            ])
                        # λ行
                        if business_type == BIZ_SPOT:
                            _lh_lam_sv2 = ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2, arpu_long_sv2, lam_sv2_actual, _dorm_sv2)
                            _lg_lam_sv2 = ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2*gpm, arpu_long_sv2*gpm, lam_sv2_actual, _dorm_sv2)
                        else:
                            _lh_lam_sv2 = ltv_horizon_offset(k_sv2, lam_sv2, arpu_sv2, lam_sv2_actual, ltv_offset_days)
                            _lg_lam_sv2 = _lh_lam_sv2 * gpm
                        _pct_lam_sv2 = _lh_lam_sv2 / ltv_inf_sv2 * 100 if ltv_inf_sv2 > 0 else 0
                        hor_data2.append([
                            T('tbl_lam_row', n=f'{round(lam_sv2_actual):,}'), f'{fmt_c(_lh_lam_sv2, CUR)}',
                            f'{_pct_lam_sv2:.1f}%',
                            f'{fmt_c(_lg_lam_sv2 / cac_n, CUR)}',
                            f'{_pct_lam_sv2:.1f}%',
                        ])
                        # 99%到達行
                        try:
                            if business_type == BIZ_SPOT:
                                _d99_sv2 = brentq(
                                    lambda hh: ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2, arpu_long_sv2, hh, _dorm_sv2) / ltv_inf_sv2 - 0.99,
                                    1, 500000)
                                _lh99_sv2 = ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2, arpu_long_sv2, _d99_sv2, _dorm_sv2)
                                _lg99_sv2 = ltv_horizon_spot(k_sv2, lam_sv2, arpu_0_dorm_sv2*gpm, arpu_long_sv2*gpm, _d99_sv2, _dorm_sv2)
                            else:
                                _d99_sv2 = brentq(
                                    lambda hh: ltv_horizon_offset(k_sv2, lam_sv2, arpu_sv2, hh, ltv_offset_days) / ltv_inf_sv2 - 0.99,
                                    1, 500000)
                                _lh99_sv2 = ltv_horizon_offset(k_sv2, lam_sv2, arpu_sv2, _d99_sv2, ltv_offset_days)
                                _lg99_sv2 = _lh99_sv2 * gpm
                            hor_data2.append([
                                T('tbl_99pct_row', n=f'{int(_d99_sv2):,}'), f'{fmt_c(_lh99_sv2, CUR)}',
                                '99.0%',
                                f'{fmt_c(_lg99_sv2 / cac_n, CUR)}',
                                '99.0%',
                            ])
                        except Exception:
                            _r99_approx = ltv_inf_sv2 * 0.99
                            hor_data2.append([
                                T('tbl_99pct_row', n='~'), f'{fmt_c(_r99_approx, CUR)}',
                                '99.0%',
                                f'{fmt_c(_r99_approx * gpm / cac_n, CUR)}',
                                '99.0%',
                            ])
                        # LTV∞行
                        _ltv_g_sv2 = _pdf_sr['ltv_gp']
                        hor_data2.append([
                            'LTV∞', f'{fmt_c(ltv_inf_sv2, CUR)}',
                            '100.0%',
                            f'{fmt_c(_ltv_g_sv2 / cac_n, CUR)}',
                            '100.0%',
                        ])
                        _seg_ltv_title_cw = 4 * cm
                        _seg_ltv_data_cw = (CONTENT_W - _seg_ltv_title_cw) / 4
                        t_sv2 = RLTable(hor_data2,
                                       colWidths=[_seg_ltv_title_cw] + [_seg_ltv_data_cw] * 4)
                        _sv2_style = _dark_tbl_style(has_title_col=True)
                        _sv2_style.add('FONTSIZE', (0, 0), (-1, -1), 7)
                        _sv2_style.add('TOPPADDING', (0, 0), (-1, -1), 2)
                        _sv2_style.add('BOTTOMPADDING', (0, 0), (-1, -1), 2)
                        t_sv2.setStyle(_sv2_style)
                        story.append(t_sv2)
                        story.append(Spacer(1, 1.0 * cm))
                        _seg_detail_count += 1
                        # 2項目ごとに改ページ
                        if _seg_detail_count % 2 == 0:
                            story.append(PageBreak())
                    except Exception:
                        continue

        def _pdf_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('HeiseiMin-W3', 7)
            canvas.setFillColor(rl_colors.HexColor('#555555'))
            _footer_text = f'Copyright © LTV-analyzer All rights reserved.  |  Page {doc.page}'
            canvas.drawCentredString(A4[0] / 2, 1.0 * cm, _footer_text)
            # ページ背景
            canvas.setFillColor(_BG)
            canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
            canvas.restoreState()

        def _pdf_on_page(canvas, doc):
            # 背景を先に描画
            canvas.saveState()
            canvas.setFillColor(_BG)
            canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
            canvas.restoreState()
            # フッター
            canvas.saveState()
            canvas.setFont('HeiseiMin-W3', 7)
            canvas.setFillColor(rl_colors.HexColor('#555555'))
            _footer_text = f'Copyright \u00a9 LTV-analyzer All rights reserved.  |  Page {doc.page}'
            canvas.drawCentredString(A4[0] / 2, 1.0 * cm, _footer_text)
            canvas.restoreState()

        doc.build(story, onFirstPage=_pdf_on_page, onLaterPages=_pdf_on_page)
        pdf_buf.seek(0)
        import base64 as _b64
        _pdf_b64 = _b64.b64encode(pdf_buf.read()).decode()
        _fn_pdf = f"LTV_Analysis_{client_name or 'report'}.pdf"
        _pdf_href = (f'<a href="data:application/pdf;base64,{_pdf_b64}" '
                     f'download="{_fn_pdf}" class="dl-btn">.pdf</a>')
        _pdf_html = _pdf_href
    except ImportError:
        _pdf_html = '<span class="dl-btn-err">.pdf Not supported</span>'
    except Exception as e:
        import traceback as _tb_pdf
        _tb_pdf_str = _tb_pdf.format_exc().replace('\n', ' | ')[-300:]
        _pdf_html = (f'<span class="dl-btn-err">.pdf Error: {str(e)[:150]}'
                     f'<br><small style="font-size:0.6rem;opacity:0.7">{_tb_pdf_str}</small></span>')

# ── 3ボタンまとめて表示 ───────────────────────────────────────
st.markdown(f"""
<style>
.dl-row {{ display:flex; gap:8px; align-items:center; margin-top:4px; }}
a.dl-btn {{
    display:inline-flex; align-items:center; justify-content:center;
    width:72px; height:30px;
    background:#0d1f2d; color:#a8dadc;
    border:1.5px solid #56b4d3; border-radius:6px;
    font-size:0.78rem; font-weight:600; letter-spacing:0.04em;
    text-decoration:none;
    transition:background 0.2s, color 0.2s;
}}
a.dl-btn:hover {{ background:#56b4d3; color:#0d1f2d; }}
a.dl-btn:focus, a.dl-btn:focus-visible, a.dl-btn:active {{
    outline: none !important;
    box-shadow: none !important;
    border: 1.5px solid #56b4d3 !important;
    background: #0d1f2d;
    color: #a8dadc;
}}
.dl-btn-err {{ font-size:0.75rem; color:#888; }}
</style>
<div class="dl-row">
  {_xl_html if '_xl_html' in dir() else ''}
  {_pp_html if '_pp_html' in dir() else ''}
  {_pdf_html if '_pdf_html' in dir() else ''}
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# Segment Analysis (Advanced)
# ══════════════════════════════════════════════════════════════

if APP_MODE == 'standard':
    st.markdown(f"<div class='section-title'>{T('section_segment')}</div>", unsafe_allow_html=True)
    st.markdown(_UPGRADE_HTML, unsafe_allow_html=True)
elif segment_cols_input.strip():
    seg_cols = [c.strip() for c in segment_cols_input.split(',') if c.strip()]
    valid_seg_cols = [c for c in seg_cols if c in df.columns]
    invalid_seg_cols = [c for c in seg_cols if c not in df.columns]

    if invalid_seg_cols:
        st.warning(T('seg_col_not_found', cols=str(invalid_seg_cols)))

    MAX_SEG_COLS = 5
    if len(valid_seg_cols) > MAX_SEG_COLS:
        st.warning(T('seg_max_cols', max=MAX_SEG_COLS, n=len(valid_seg_cols)))
        valid_seg_cols = valid_seg_cols[:MAX_SEG_COLS]

    if valid_seg_cols:
        st.markdown(f"<div class='section-title'>{T('section_segment')}</div>", unsafe_allow_html=True)

        # セグメント一覧リンク
        _seg_links = " &nbsp;|&nbsp; ".join(
            f"<a href='#{c}_anchor' style='color:#56b4d3; font-size:0.78rem; text-decoration:none;'>{c}</a>"
            for c in valid_seg_cols
        )
        st.markdown(f"<div style='margin-bottom:12px; font-size:0.78rem; color:#888;'>{_seg_links}</div>", unsafe_allow_html=True)

        for seg_col in valid_seg_cols:
            st.markdown(f"<div id='{seg_col}_anchor' style='position:relative; top:-120px; visibility:hidden;'></div><div style='font-size:0.82rem; font-weight:600; color:#a8dadc; margin:16px 0 4px 0; letter-spacing:0.05em;'>{seg_col}</div>", unsafe_allow_html=True)

            seg_values = df[seg_col].dropna().unique()
            if len(seg_values) > 50:
                st.warning(T('seg_too_many', col=seg_col, n=len(seg_values)))
                continue
            elif len(seg_values) > 20:
                st.info(T('seg_many_values', col=seg_col, n=len(seg_values)))

            # Pre-computeの結果を使用（計算は1回だけ）
            if seg_col not in all_seg_results:
                st.caption(T('seg_insufficient_data'))
                continue

            seg_results = all_seg_details[seg_col]
            seg_df = all_seg_results[seg_col]

            # Top Pick（タイトルとグラフの間）
            if seg_results:
                best_seg    = seg_df.iloc[0]
                avg_ltv     = seg_df['ltv_rev'].mean()
                premium     = (best_seg['ltv_rev'] - avg_ltv) / avg_ltv * 100
                cac_best    = best_seg['cac_cap']
                cac_avg_seg = (seg_df['cac_cap'] * seg_df['n_customers']).sum() / seg_df['n_customers'].sum()
                cac_diff    = cac_best - cac_avg_seg
                cac_str     = f"CAC Ceiling {fmt_c(cac_best, CUR)} (vs segment avg: {fmt_c(abs(cac_diff), CUR)}{'higher' if cac_diff >= 0 else 'lower'})"
                st.markdown(f"""
<div style='background:#0d1f2d; border:1px solid #1a3a4a; border-left:3px solid #56b4d3; border-radius:8px; padding:10px 16px; margin-bottom:8px; font-size:0.82rem; color:#ccc;'>
  <span style='font-size:0.65rem; font-weight:600; text-transform:uppercase; letter-spacing:0.12em; color:#7ab4c4;'>Top Pick</span>　<b style='color:#a8dadc;'>{best_seg['segment']}</b>　
  <span style='color:#888; margin:0 6px;'>|</span>
  LTV∞ (Rev): <b style='color:#a8dadc;'>{fmt_c(best_seg['ltv_rev'], CUR)}</b>（vs avg +{premium:.1f}%）
  <span style='color:#888; margin:0 6px;'>|</span>
  {cac_str}
</div>""", unsafe_allow_html=True)

            # Plotlyで横棒グラフ（PPTXと統一）
            n_seg = len(seg_df)
            bar_colors = [ACCENT if i == 0 else ACCENT2 for i in range(n_seg)]
            fig_height = max(300, n_seg * 35 + 100)

            fig_plotly = go.Figure()
            seg_names = seg_df['segment'].astype(str).tolist()[::-1]
            seg_vals  = seg_df['ltv_rev'].tolist()[::-1]
            bar_colors_rev = bar_colors[::-1]
            fig_plotly.add_trace(go.Bar(
                y=seg_names,
                x=seg_vals,
                orientation='h',
                marker=dict(color=bar_colors_rev),
                text=[f'{fmt_c(v, CUR)}' for v in seg_vals],
                textposition='inside',
                insidetextanchor='end',
                textfont=dict(color='white', size=10),
                hovertemplate='%{y}: ' + cur_symbol(CUR) + '%{x:,.0f}<extra></extra>',
            ))
            fig_plotly.update_yaxes(
                tickfont=dict(color='#aaa', size=10 if n_seg <= 20 else 8),
                gridcolor='#1a3040',
                linecolor='#1a3040',
            )
            fig_plotly.update_xaxes(
                title_text=f'LTV∞ ({cur_symbol(CUR)})',
                tickfont=dict(color='#888'),
                gridcolor='#1a3040',
                tickprefix=cur_symbol(CUR),
                tickformat=',',
                showgrid=True,
                side='top',
            )
            fig_plotly.update_layout(
                title_text='',
                paper_bgcolor='#111820',
                plot_bgcolor='#111820',
                height=fig_height,
                showlegend=False,
                margin=dict(t=30, b=40, l=120, r=30),
                font=dict(color='#ccc', size=12),
            )
            st.plotly_chart(fig_plotly, use_container_width=True)

            # 結果テーブル
            display_df = seg_df.copy()
            display_df = display_df.drop(columns=['total_potential', 'priority_score', 'λ_raw', 'arpu_s', 'arpu_long_s', 'arpu_0_dorm_s'], errors='ignore')
            if not cac_known:
                display_df = display_df.drop(columns=['acq_efficiency'], errors='ignore')

            # セグメント加重平均LTV∞を計算してテーブル上に表示
            total_n = seg_df['n_customers'].sum()
            weighted_ltv_rev = (seg_df['ltv_rev'] * seg_df['n_customers']).sum() / total_n
            weighted_ltv_gp  = (seg_df['ltv_gp'] * seg_df['n_customers']).sum() / total_n
            weighted_cac     = weighted_ltv_gp / cac_n
            diff_pct = (weighted_ltv_rev - ltv_rev) / ltv_rev * 100
            diff_str = f"+{diff_pct:.1f}%" if diff_pct >= 0 else f"{diff_pct:.1f}%"
            diff_color = "#a8dadc" if diff_pct >= 0 else "#e8a0a0"

            # HTMLテーブルで描画（数字右寄せ・列幅均等）
            ACCENT   = '#56b4d3'
            BG_HEAD  = '#0d1f2d'
            BG_ROW1  = '#0d1520'
            BG_ROW2  = '#0a1018'
            num_cols = ['n_customers', 'ltv_rev', 'ltv_gp', 'cac_cap', 'k', 'lam_days', 'R²']
            if cac_known:
                num_cols.append('acq_efficiency')
            cols = [c for c in display_df.columns]
            n_cols = len(cols)
            seg_col_w = 20
            num_col_w = round((100 - seg_col_w) / (n_cols - 1), 1)

            # ヘッダー
            seg_html_rows = ''
            for i, row in display_df.iterrows():
                bg = BG_ROW1 if i % 2 == 0 else BG_ROW2
                seg_html_rows += f"<tr style='background:{bg};'>"
                for col in cols:
                    val = row[col]
                    # 数値フォーマット
                    if col == 'n_customers':
                        val = f'{int(val):,}'
                        align = 'right'
                    elif col in ['ltv_rev', 'ltv_gp', 'cac_cap']:
                        val = f'{fmt_c(val, CUR)}'
                        align = 'right'
                    elif col == 'k':
                        val = f'{val:.3f}'
                        align = 'right'
                    elif col == 'lam_days':
                        val = f'{val:.1f}'
                        align = 'right'
                    elif col == 'R²':
                        val = f'{val:.3f}'
                        align = 'right'
                    elif col == 'acq_efficiency':
                        val = f'{val:.2f}x' if val else '-'
                        align = 'right'
                    else:
                        align = 'left'
                    w = f'{seg_col_w}%' if col == 'segment' else f'{num_col_w}%'
                    seg_html_rows += f"<td style='text-align:{align}; padding:8px 14px; color:#c8d0d8; font-size:0.85rem; font-variant-numeric:tabular-nums; width:{w};'>{val}</td>"
                seg_html_rows += '</tr>'

            header_html = ''
            _col_display = {
                'segment': T('seg_tbl_segment'), 'n_customers': T('seg_tbl_n'),
                'ltv_rev': T('seg_tbl_ltv_rev'), 'ltv_gp': T('seg_tbl_ltv_gp'),
                'cac_cap': T('seg_tbl_cac_cap'), 'k': 'k', 'lam_days': T('seg_tbl_lam'),
                'R²': 'R²', 'acq_efficiency': 'ROI',
            }
            for col in cols:
                align = 'left' if col == 'segment' else 'right'
                w = f'{seg_col_w}%' if col == 'segment' else f'{num_col_w}%'
                header_html += f"<th style='text-align:{align}; padding:9px 14px; color:{ACCENT}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT}; width:{w};'>{_col_display.get(col, col)}</th>"

            # 加重平均行
            avg_row_vals = {
                'segment': T('seg_weighted_avg'),
                'n_customers':      f'{total_n:,}',
                'ltv_rev': f'{fmt_c(weighted_ltv_rev, CUR)}',
                'ltv_gp': f'{fmt_c(weighted_ltv_gp, CUR)}',
                'cac_cap': f'{fmt_c(weighted_cac, CUR)}',
                'k': '—', 'lam_days': '—', 'R²': '—',
            }
            avg_row_html = "<tr style='background:#0d1f2d; border-top:1px solid #1a3a4a;'>"
            for col in cols:
                align = 'left' if col == 'segment' else 'right'
                w = f'{seg_col_w}%' if col == 'segment' else f'{num_col_w}%'
                val = avg_row_vals.get(col, '—')
                avg_row_html += f"<td style='text-align:{align}; padding:8px 14px; color:#a8dadc; font-size:0.82rem; font-variant-numeric:tabular-nums; width:{w};'>{val}</td>"
            avg_row_html += '</tr>'

            seg_tbl_html = f"""
<table style='width:100%; border-collapse:collapse; margin-top:4px; table-layout:fixed;'>
  <colgroup>{''.join([f'<col style="width:{seg_col_w}%;">' if c == 'segment' else f'<col style="width:{num_col_w}%;">' for c in cols])}</colgroup>
  <thead><tr style='background:{BG_HEAD};'>{header_html}</tr></thead>
  <tbody>{seg_html_rows}{avg_row_html}</tbody>
</table>"""
            st.markdown(seg_tbl_html, unsafe_allow_html=True)

            # NOTEのみ（加重平均はテーブル内に移動）
            st.markdown(f"""
<div style='background:#0a1520; border:1px solid #1a3040; border-radius:8px; padding:10px 18px; margin-top:6px; font-size:0.78rem; color:#888; line-height:1.7;'>
  <span style='font-size:0.65rem; font-weight:600; text-transform:uppercase; letter-spacing:0.1em; color:#3a6a7a;'>Note</span> — {T('seg_note', ltv=fmt_c(ltv_rev, CUR), diff=diff_str)}
</div>""", unsafe_allow_html=True)



            # ── 全セグメントの詳細分析（上位N件のみブラウザ表示）──
            seg_results_display = seg_results[:seg_display_limit]
            remaining = len(seg_results) - len(seg_results_display)
            st.markdown(f"<div style='font-size:0.78rem; color:#888; margin:12px 0 4px 0;'>{T('seg_detail_limit', limit=seg_display_limit, total=len(seg_results))}</div>", unsafe_allow_html=True)
            if remaining > 0:
                st.caption(T("sidebar_display_caption").split(chr(10))[1].strip() if get_lang() == "en" else "エクスポートされる各ファイルには全項目出力されます。サイドバーの「ブラウザ表示件数」を増やすと追加表示できます。")
            for sr in seg_results_display:
                sv           = sr['segment']
                k_s          = sr['k']
                lam_s        = sr['λ_raw']         # Weibull計算用生値
                lam_s_disp   = sr['lam_days']         # 表示用（オフセット込み）
                r2_s         = sr['R²']
                n_s          = sr['n_customers']
                arpu_s       = sr['arpu_s']
                arpu_long_s  = sr['arpu_long_s']
                arpu_0_dorm_s= sr['arpu_0_dorm_s']
                gp_s         = arpu_s * gpm
                df_sv        = df[df[seg_col] == sv]
                _dorm_s      = dormancy_days or 180 if business_type == BIZ_SPOT else ltv_offset_days
                if business_type == BIZ_SPOT:
                    _ltv_short_s = _dorm_s * arpu_0_dorm_s
                    _ltv_long_s, _ = ltv_inf_offset(k_s, lam_s, arpu_long_s, 0)
                    ltv_inf_s = _ltv_short_s + _ltv_long_s
                else:
                    ltv_inf_s = ltv_inf_offset(k_s, lam_s, arpu_s, ltv_offset_days)[0]
                is_best = (sv == seg_df.iloc[0]['segment'])

                with st.expander(
                    f"{sv}{' ·  Priority' if is_best else ''}  ·  {n_s:,} customers  ·  LTV∞ {fmt_c(sr['ltv_rev'], CUR)}  ·  k={k_s:.3f}  ·  λ={lam_s:.1f}d  ·  R²={r2_s:.3f}",
                    expanded=is_best
                ):
                    try:
                        # オフセット処理したkm_svを使用
                        if ltv_offset_days > 0:
                            df_sv_fit = df_sv.copy()
                            df_sv_fit['duration'] = df_sv_fit['duration'] - ltv_offset_days
                            df_sv_fit.loc[df_sv_fit['duration'] <= 0, 'event'] = 0
                            df_sv_fit.loc[df_sv_fit['duration'] <= 0, 'duration'] = 1
                            df_sv_fit['duration'] = df_sv_fit['duration'].clip(lower=1)
                            km_sv_fit = _compute_km_df(df_sv_fit)
                        else:
                            km_sv_fit = _compute_km_df(df_sv)
                        km_sv = _compute_km_df(df_sv)  # 表示用（オフセットなし）

                        # ── グラフ2枚（全体と同じ）──
                        st.markdown(f"<div style='font-size:0.75rem; color:#7ab4c4; font-weight:600; text-transform:uppercase; letter-spacing:0.08em; margin-bottom:6px;'>{T('chart_reliability_title')}</div>", unsafe_allow_html=True)
                        col_g1, col_g2 = st.columns(2)

                        with col_g1:
                            # KM曲線：km_sv_fit（オフセット後）のt軸を+ltv_offset_daysで元スケールに戻す
                            # t=0〜ltv_offset_daysはS=1.0フラット
                            _km_sv_t = [0, ltv_offset_days] + [t + ltv_offset_days for t in km_sv_fit['t'].tolist()]
                            _km_sv_s = [1.0, 1.0] + km_sv_fit['S'].tolist()
                            fig_sv1 = go.Figure()
                            fig_sv1.add_trace(go.Scatter(
                                x=_km_sv_t, y=_km_sv_s,
                                mode='lines', name='KM Curve (Observed)',
                                line=dict(color=ACCENT, width=2, shape='hv')
                            ))
                            # Weibull：t=ltv_offset_daysスタート（日割りONはt=0）
                            t_max_s = int(km_sv_fit['t'].max() + ltv_offset_days)
                            _wei_start = int(ltv_offset_days)
                            t_range_s = list(range(_wei_start, t_max_s + 30, max(1, t_max_s // 200)))
                            fig_sv1.add_trace(go.Scatter(
                                x=t_range_s,
                                y=[float(weibull_s(max(t - ltv_offset_days, 0), k_s, lam_s)) for t in t_range_s],
                                mode='lines', name=f'Weibull Fit',
                                line=dict(color=ACCENT2, width=1.5, dash='dash')
                            ))
                            fig_sv1.update_layout(
                                title=dict(text='Survival Curve', font=dict(color='#ccc', size=12)),
                                paper_bgcolor='#111820', plot_bgcolor='#111820',
                                height=300, margin=dict(t=40, b=50, l=50, r=10),
                                font=dict(color='#ccc', size=10),
                                legend=dict(font=dict(size=9), bgcolor='rgba(0,0,0,0)'),
                                xaxis=dict(title='Days', gridcolor='#1a3040', tickfont=dict(color='#888')),
                                yaxis=dict(title='Survival Rate S(t)', gridcolor='#1a3040', tickfont=dict(color='#888'), range=[0, 1.05]),
                            )
                            st.plotly_chart(fig_sv1, use_container_width=True)
                            st.caption(f"Survival Curve: KM observed (solid) + Weibull fit (dashed). k={k_s:.3f}, λ={lam_s:.1f}{T('chart_days_suffix')}")

                        with col_g2:
                            # Weibull直線化プロット
                            import numpy as np_sv
                            km_fit = km_sv[km_sv['S'] > 0].copy()
                            ln_t_s = np_sv.log(km_fit['t'].values.astype(float) + 1e-10)
                            ln_neg_ln_S_s = np_sv.log(-np_sv.log(km_fit['S'].values.astype(float) + 1e-15))
                            valid_s = np_sv.isfinite(ln_t_s) & np_sv.isfinite(ln_neg_ln_S_s)
                            ln_t_s = ln_t_s[valid_s]
                            ln_neg_ln_S_s = ln_neg_ln_S_s[valid_s]
                            slope_s, intercept_s, _, _, _ = __import__('scipy').stats.linregress(ln_t_s, ln_neg_ln_S_s)
                            x_line_s = [float(ln_t_s.min()), float(ln_t_s.max())]
                            y_line_s = [slope_s * x + intercept_s for x in x_line_s]

                            fig_sv2 = go.Figure()
                            fig_sv2.add_trace(go.Scatter(
                                x=ln_t_s.tolist(), y=ln_neg_ln_S_s.tolist(),
                                mode='markers', name='Observed',
                                marker=dict(color=ACCENT, size=5, opacity=0.7)
                            ))
                            fig_sv2.add_trace(go.Scatter(
                                x=x_line_s, y=y_line_s,
                                mode='lines', name=f'Regression Line (R²={r2_s:.3f})',
                                line=dict(color=ACCENT2, width=1.5, dash='dash')
                            ))
                            fig_sv2.add_annotation(
                                x=0.05, y=0.93, xref='paper', yref='paper',
                                text=f'y = {slope_s:.4f}x + {intercept_s:.4f}',
                                showarrow=False, font=dict(color='#777', size=9)
                            )
                            fig_sv2.update_layout(
                                title=dict(text='Weibull Linearization Plot', font=dict(color='#ccc', size=12)),
                                paper_bgcolor='#111820', plot_bgcolor='#111820',
                                height=300, margin=dict(t=40, b=50, l=50, r=10),
                                font=dict(color='#ccc', size=10),
                                legend=dict(font=dict(size=9), bgcolor='rgba(0,0,0,0)'),
                                xaxis=dict(title='ln(t)', gridcolor='#1a3040', tickfont=dict(color='#888')),
                                yaxis=dict(title='ln(−ln(S(t)))', gridcolor='#1a3040', tickfont=dict(color='#888')),
                            )
                            st.plotly_chart(fig_sv2, use_container_width=True)
                            st.caption(f"Weibull linearization: R²={r2_s:.3f} (closer to 1.0 = better fit)")

                        # ── 暫定LTVテーブル（全体と同仕様）──
                        st.markdown(f"<div style='font-size:0.75rem; color:#7ab4c4; font-weight:600; text-transform:uppercase; letter-spacing:0.08em; margin-top:16px; margin-bottom:6px;'>{T('chart_interim_ltv_title')}</div>", unsafe_allow_html=True)

                        # グラフ
                        lam_s_actual = lam_s_disp  # 表示用λ（オフセット込み済み）
                        lam_s_int = round(lam_s_actual)
                        x_max_s = max(1825, lam_s_int + 100) if lam_s_actual > 1825 else 1825
                        t_range_s = list(range(1, x_max_s + 1, max(1, x_max_s // 300)))
                        if business_type == BIZ_SPOT:
                            rev_line_s = [ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, t, _dorm_s) for t in t_range_s]
                            gp_line_s  = [ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, t, _dorm_s) for t in t_range_s]
                        else:
                            rev_line_s = [ltv_horizon_offset(k_s, lam_s, arpu_s, t, ltv_offset_days) for t in t_range_s]
                            gp_line_s  = [ltv_horizon_offset(k_s, lam_s, arpu_s * gpm, t, ltv_offset_days) for t in t_range_s]
                        cac_line_s = [v / cac_n for v in gp_line_s]
                        if business_type == BIZ_SPOT:
                            ltv_inf_s_offset = ltv_inf_s
                        else:
                            ltv_inf_s_offset, _ = ltv_inf_offset(k_s, lam_s, arpu_s, ltv_offset_days)

                        fig_hor_s = go.Figure()
                        fig_hor_s.add_trace(go.Scatter(x=t_range_s, y=rev_line_s, name=T('chart_ltv_rev'), mode='lines', line=dict(color='#56b4d3', width=2)))
                        fig_hor_s.add_trace(go.Scatter(x=t_range_s, y=gp_line_s,  name=T('chart_ltv_gp'), mode='lines', line=dict(color='#a8dadc', width=2, dash='dash')))
                        fig_hor_s.add_trace(go.Scatter(x=t_range_s, y=cac_line_s, name=T('chart_cac_cap'),    mode='lines', line=dict(color='#4a7a8a', width=1.5, dash='dot')))
                        fig_hor_s.add_hline(y=ltv_inf_s_offset, line_dash='dot', line_color='#56b4d3', line_width=1, opacity=0.4,
                            annotation_text=f'LTV∞ {fmt_c(ltv_inf_s_offset, CUR)}', annotation_position='right', annotation_font=dict(color='#56b4d3', size=10))
                        fig_hor_s.add_shape(type='line', x0=lam_s_actual, x1=lam_s_actual, y0=0, y1=ltv_inf_s_offset,
                            line=dict(color='#a8dadc', width=1.5, dash='dash'), layer='above')
                        fig_hor_s.add_annotation(x=lam_s_actual, y=0.85 if k_s < 1.0 else 0.35, yref='paper',
                            text=T('chart_lam_days', n=lam_s_int), showarrow=False, font=dict(color='#a8dadc', size=10),
                            xanchor='center', yanchor='middle', bgcolor='#111820', borderpad=2)

                        # プロット点（全体分析と同じ分岐）
                        plot_pts_s = sorted(set([p for p in [180, 365, 730, 1095, 1460, 1825, lam_s_int] if p <= x_max_s]))
                        for arpu_v, color in [(arpu_s, '#56b4d3'), (arpu_s * gpm, '#a8dadc'), (arpu_s * gpm / cac_n, '#4a7a8a')]:
                            px_s = [p for p in plot_pts_s]
                            if business_type == BIZ_SPOT:
                                if arpu_v == arpu_s * gpm / cac_n:
                                    py_s = [ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, p, _dorm_s) / cac_n for p in plot_pts_s]
                                elif arpu_v == arpu_s * gpm:
                                    py_s = [ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, p, _dorm_s) for p in plot_pts_s]
                                else:
                                    py_s = [ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, p, _dorm_s) for p in plot_pts_s]
                            else:
                                if arpu_v == arpu_s * gpm / cac_n:
                                    py_s = [ltv_horizon_offset(k_s, lam_s, arpu_s * gpm, p, ltv_offset_days) / cac_n for p in plot_pts_s]
                                else:
                                    py_s = [ltv_horizon_offset(k_s, lam_s, arpu_v, p, ltv_offset_days) for p in plot_pts_s]
                            fig_hor_s.add_trace(go.Scatter(x=px_s, y=py_s, mode='markers',
                                marker=dict(color=color, size=5, line=dict(color='#111820', width=1)),
                                showlegend=False, hovertemplate='%{x}' + T('chart_days_suffix') + ': ' + cur_symbol(CUR) + '%{y:,.0f}<extra></extra>'))

                        tick_vals_s = [180, 365, 730, 1095, 1460, 1825]
                        tick_text_s = [f'180{T("chart_days_suffix")}', f'1{T("chart_year_suffix")}', f'2{T("chart_year_suffix")}', f'3{T("chart_year_suffix")}', f'4{T("chart_year_suffix")}', f'5{T("chart_year_suffix")}']
                        fig_hor_s.update_layout(
                            paper_bgcolor='#111820', plot_bgcolor='#111820',
                            height=280, margin=dict(t=40, b=50, l=70, r=120),
                            font=dict(color='#ccc', size=10),
                            legend=dict(orientation='h', y=1.08, x=0, font=dict(size=10), bgcolor='rgba(0,0,0,0)'),
                            xaxis=dict(title=T('chart_duration'), gridcolor='#1a3040', tickvals=tick_vals_s, ticktext=tick_text_s, tickfont=dict(color='#888')),
                            yaxis=dict(title=T('chart_amount'), gridcolor='#1a3040', tickfont=dict(color='#888'), tickformat=',', tickprefix=''),
                        )
                        st.plotly_chart(fig_hor_s, use_container_width=True)

                        # テーブル
                        ACCENT_S = '#56b4d3'; BG_HEAD_S = '#0d1f2d'; BG_ROW1_S = '#0d1520'; BG_ROW2_S = '#0a1018'
                        SEP_S = '#1a3a4a'
                        all_horizons_s = [180, 365, 730, 1095, 1825]
                        if business_type == BIZ_SPOT:
                            lam_s_rev  = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, lam_s_actual, _dorm_s)
                            lam_s_gp   = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, lam_s_actual, _dorm_s)
                            rev_99_s_d = brentq(lambda h: ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, h, _dorm_s) / ltv_inf_s_offset - 0.99, 1, 365000)
                            rev_99_s   = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, rev_99_s_d, _dorm_s)
                            gp_99_s    = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, rev_99_s_d, _dorm_s)
                        else:
                            lam_s_rev  = ltv_horizon_offset(k_s, lam_s, arpu_s, lam_s_actual, ltv_offset_days)
                            lam_s_gp   = ltv_horizon_offset(k_s, lam_s, arpu_s * gpm, lam_s_actual, ltv_offset_days)
                            rev_99_s_d = brentq(lambda h: ltv_horizon_offset(k_s, lam_s, arpu_s, h, ltv_offset_days) / ltv_inf_s_offset - 0.99, 1, 365000)
                            rev_99_s   = ltv_horizon_offset(k_s, lam_s, arpu_s, rev_99_s_d, ltv_offset_days)
                            gp_99_s    = ltv_horizon_offset(k_s, lam_s, arpu_s * gpm, rev_99_s_d, ltv_offset_days)

                        hor_html_rows = ''
                        for idx_h, h in enumerate(all_horizons_s):
                            if business_type == BIZ_SPOT:
                                lh_rev_s = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s, arpu_long_s, h, _dorm_s)
                                lh_gp_s  = ltv_horizon_spot(k_s, lam_s, arpu_0_dorm_s*gpm, arpu_long_s*gpm, h, _dorm_s)
                            else:
                                lh_rev_s = ltv_horizon_offset(k_s, lam_s, arpu_s, h, ltv_offset_days)
                                lh_gp_s  = ltv_horizon_offset(k_s, lam_s, arpu_s * gpm, h, ltv_offset_days)
                            label = fmt_horizon(h)
                            bg = BG_ROW1_S if idx_h % 2 == 0 else BG_ROW2_S
                            pct = lh_rev_s / ltv_inf_s_offset * 100
                            hor_html_rows += f"<tr style='background:{bg};'><td style='text-align:left; padding:8px 14px; color:#c8d0d8; font-size:0.85rem; width:28%;'>{label}</td><td style='text-align:right; padding:8px 14px; color:#c8d0d8; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(lh_rev_s, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#c8d0d8; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(lh_gp_s, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#c8d0d8; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(lh_gp_s/cac_n, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#c8d0d8; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{pct:.1f}%</td></tr>"

                        # λ行
                        lam_pct_s = lam_s_rev / ltv_inf_s_offset * 100
                        hor_html_rows += f"<tr style='background:{BG_HEAD_S}; border-top:1px solid {SEP_S};'><td style='text-align:left; padding:8px 14px; color:#a8dadc; font-size:0.85rem; width:28%;'>λ {lam_s_int}日</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(lam_s_rev, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(lam_s_gp, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(lam_s_gp/cac_n, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{lam_pct_s:.1f}%</td></tr>"

                        # 99%行
                        hor_html_rows += f"<tr style='background:{BG_HEAD_S}; border-top:1px solid {SEP_S};'><td style='text-align:left; padding:8px 14px; color:#a8dadc; font-size:0.85rem; width:28%;'>{T('tbl_99pct_row', n=f'{int(rev_99_s_d):,}')}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(rev_99_s, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(gp_99_s, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(gp_99_s/cac_n, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>99.0%</td></tr>"

                        # LTV∞行
                        _ltv_gp_s_tbl = ltv_inf_s_offset * gpm if business_type == BIZ_SPOT else ltv_inf_offset(k_s, lam_s, arpu_s * gpm, ltv_offset_days)[0]
                        hor_html_rows += f"<tr style='background:{BG_HEAD_S}; border-top:1px solid {SEP_S};'><td style='text-align:left; padding:8px 14px; color:#a8dadc; font-size:0.85rem; width:28%;'>LTV∞</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(ltv_inf_s_offset, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(_ltv_gp_s_tbl, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>{fmt_c(_ltv_gp_s_tbl/cac_n, CUR)}</td><td style='text-align:right; padding:8px 14px; color:#a8dadc; font-size:0.85rem; font-variant-numeric:tabular-nums; width:18%;'>100%</td></tr>"

                        hor_tbl_html = f"""
<table style='width:100%; border-collapse:collapse; margin-top:4px; table-layout:fixed;'>
  <colgroup><col style='width:28%;'><col style='width:18%;'><col style='width:18%;'><col style='width:18%;'><col style='width:18%;'></colgroup>
  <thead><tr style='background:{BG_HEAD_S};'>
    <th style='text-align:left; padding:9px 14px; color:{ACCENT_S}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT_S};'>'{T('tbl_horizon')}'</th>
    <th style='text-align:right; padding:9px 14px; color:{ACCENT_S}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT_S};'>{T('tbl_ltv_rev')}</th>
    <th style='text-align:right; padding:9px 14px; color:{ACCENT_S}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT_S};'>{T('tbl_ltv_gp')}</th>
    <th style='text-align:right; padding:9px 14px; color:{ACCENT_S}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT_S};'>{T('tbl_cac_cap')}</th>
    <th style='text-align:right; padding:9px 14px; color:{ACCENT_S}; font-size:0.8rem; font-weight:600; border-bottom:2px solid {ACCENT_S};'>{T('tbl_pct_ltv')}</th>
  </tr></thead>
  <tbody>{hor_html_rows}</tbody>
</table>"""
                        st.markdown(hor_tbl_html, unsafe_allow_html=True)


                    except Exception as e_sv:
                        st.caption(f"Chart error: {e_sv}")

else:
    st.markdown(f"<div class='section-title'>{T('section_segment')}</div>", unsafe_allow_html=True)
    st.info(T('seg_hint_info'))
    st.markdown(T('seg_hint_howto'))

# ══════════════════════════════════════════════════════════════
# Data preview
# ══════════════════════════════════════════════════════════════

with st.expander("Raw Data" if get_lang() == "en" else "読み込んだデータを確認"):
    st.write(f"Records: {len(df):,} / Churned: {df['event'].sum():,} / Active: {(df['event']==0).sum():,} / Daily ARPU: {fmt_c(arpu_daily, CUR, 2)}")
    st.dataframe(
        df[['customer_id','start_date','end_date','duration','event','arpu_daily']].head(30),
        hide_index=True
    )

st.markdown("---")
st.markdown("<p style='color:#333; font-size:0.82rem; text-align:center;'>LTV Analyzer — KM × Weibull Model — Built for marketing analytics professionals</p>", unsafe_allow_html=True)
